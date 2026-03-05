# -*- coding: utf-8 -*-
"""
Java Grading Gallery — v2.2
(Pastel + per-student EXACTLY 2 questions + code comments/highlight + PDF + Summary tab + LabID + class stats + histogram)

Key features kept / improved:
- Submissions DB: scan/commit folders, store LabID per student, store file content optionally.
- Robust student ID/name detection:
  - From Java file headers (comment scanning)
  - Optional folder-name patterns (regex) + path fallback
- Flexible scan patterns:
  - Provide a list of glob patterns (e.g., ["*.java", "*.txt"]) AND/OR an include-regex for filenames.
- Grading DB: rubric CSV loading, per-question rubric scoring, rationale, per-student assessed questions (exactly two).
- Code comments tool (like “PDF comment/highlight” concept):
  - Highlight in app
  - In PDF: (a) annotated code snapshot with injected comment blocks AND (b) “highlighted lines” rendering (background for lines with comments).
- Summary tab: full class summary table including LabID + Q1/Q2 IDs + totals + overall.
- Status: class avg/min/max + suggested curve factor + histogram chart and curve overlay.
  - Includes a curve factor control (scale) to preview adjusted distribution.
- Export:
  - Excel (all)
  - Summary PDF
  - Student PDF (single)
  - Batch export PDFs for ALL students into a folder
- AI auto grade is separated into a separate class (optional, can be left empty).
  - No references/branding in UI text. The button is “Auto Grade (optional)” and it safely no-ops if disabled.

Dependencies:
- tkinter, sqlite3, openpyxl
- reportlab (optional for PDF)
- matplotlib (optional for histogram)
"""

import csv
import os
import re
import sqlite3
import hashlib
import math
import random
import textwrap
from pathlib import Path
from datetime import datetime
import json
import io
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import tkinter.font as tkfont

from openpyxl import Workbook

# PDF (optional)
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Preformatted, PageBreak
    )
    from reportlab.lib import colors
except Exception:
    SimpleDocTemplate = None

# Matplotlib (optional, for histogram UI)
try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
except Exception:
    FigureCanvasTkAgg = None
    Figure = None


# =============================================================================
# Config / Constants
# =============================================================================

SUBMISSIONS_DB = "submissions.sqlite"

DEFAULT_THEME = "Grade strictly. If unclear, give 0 and explain why."

ID_DIGITS_RE = re.compile(r"\b\d{5,12}\b")
POSITIVE_COMMENT_RE = re.compile(
    r"\b(good|great|nice|well done|correct|excellent|clean|solid|perfect)\b",
    re.IGNORECASE,
)


# =============================================================================
# Utility
# =============================================================================

def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def sha256_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8", errors="ignore")).hexdigest()

def read_file_text(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="ignore")


def extract_numeric_id(raw: str) -> str:
    """
    Keep only the numeric student ID portion.
    """
    text = (raw or "").strip()
    if not text:
        return ""
    m = ID_DIGITS_RE.search(text)
    return m.group(0) if m else ""

def is_full_student(student_id: str) -> bool:
    return (student_id or "").strip().lower() == "full"


def is_mistake_focused_comment(text: str) -> bool:
    msg = (text or "").strip()
    if not msg:
        return False
    if POSITIVE_COMMENT_RE.search(msg) and not any(k in msg.lower() for k in ["missing", "error", "bug", "incorrect", "wrong", "fix", "issue", "fail"]):
        return False
    return True


# =============================================================================
# 1) Student detection from Java header comments
# =============================================================================

def clean_comment_line(line: str) -> str:
    s = line.strip()
    s = s.replace("/*", "").replace("*/", "")
    s = re.sub(r"^\s*//\s*", "", s)
    s = re.sub(r"^\s*\*\s*", "", s)
    return s.strip()

def extract_value(line: str, want: str):
    s = clean_comment_line(line)

    if want == "id":
        s2 = re.sub(r"stud\w*\s*(number|id|no|num)\s*[:=\-]?\s*", "", s, flags=re.IGNORECASE)
        m = ID_DIGITS_RE.search(s2)
        return m.group(0) if m else None

    if want == "name":
        s2 = re.sub(r"stud\w*\s*name\s*[:=\-]?\s*", "", s, flags=re.IGNORECASE).strip()
        if not s2:
            return None
        if ID_DIGITS_RE.search(s2):
            return None
        return s2

    return None

def extract_student_info_from_file(file_path: Path, max_lines: int = 200):
    """
    Tries to parse student id/name from top-of-file comments.
    """
    student_id = None
    student_name = None
    expect_next = None  # "id" or "name"

    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            for i, raw in enumerate(f):
                if i > max_lines:
                    break

                stripped = raw.strip()
                if not stripped:
                    continue

                if not (stripped.startswith("//") or stripped.startswith("/*") or stripped.startswith("*")):
                    continue

                line = clean_comment_line(stripped)
                if not line:
                    continue

                low = line.lower()

                if expect_next == "id" and student_id is None:
                    m = ID_DIGITS_RE.search(line)
                    if m:
                        student_id = m.group(0)
                        expect_next = None
                        continue

                if expect_next == "name" and student_name is None:
                    if "student" not in low and "name" not in low and "number" not in low and "id" not in low:
                        student_name = line.strip()
                        expect_next = None
                        continue

                if student_id is None and ("student" in low or "stud" in low) and (
                    "number" in low or "id" in low or "num" in low or "no" in low
                ):
                    v = extract_value(line, "id")
                    if v:
                        student_id = v
                    else:
                        expect_next = "id"

                if student_name is None and ("student" in low or "stud" in low) and "name" in low:
                    v = extract_value(line, "name")
                    if v:
                        student_name = v
                    else:
                        expect_next = "name"

                if student_id and student_name:
                    break
    except Exception:
        pass

    return student_id, student_name


# =============================================================================
# 2) Folder-name / path-based detection (flexibility you asked for)
# =============================================================================

def try_extract_from_folder_name(folder: Path, id_regex: str, name_regex: str):
    """
    id_regex / name_regex are optional regex strings with a capturing group.
    Examples:
      id_regex   = r"(\\d{7,10})"
      name_regex = r"([A-Za-z]+\\s+[A-Za-z]+)"
    """
    fid = None
    fname = None
    text = str(folder)

    try:
        if id_regex:
            m = re.search(id_regex, text)
            if m:
                fid = m.group(1) if m.groups() else m.group(0)
    except Exception:
        pass

    try:
        if name_regex:
            m = re.search(name_regex, text)
            if m:
                fname = m.group(1) if m.groups() else m.group(0)
    except Exception:
        pass

    return fid, fname


def _compile_optional_regex(pattern: str, ignore_case: bool = False):
    raw = (pattern or "").strip()
    if not raw:
        return None
    flags = re.IGNORECASE if ignore_case else 0
    try:
        return re.compile(raw, flags)
    except Exception:
        return None

def infer_student_for_folder(
    folder: Path,
    file_globs: list[str],
    include_filename_regex: str,
    exclude_filename_regex: str,
    filename_regex_ignore_case: bool,
    folder_id_regex: str,
    folder_name_regex: str,
):
    """
    - Finds candidate files via file_globs (glob patterns) AND optional include_filename_regex.
    - Extracts student info from files (primary)
    - Falls back to folder-based regex extraction if needed.
    """
    detected_id = None
    detected_name = None

    # Collect files by globs
    files = []
    for g in (file_globs or ["*.java"]):
        files.extend(list(folder.rglob(g)))

    # Optional filename regex filter
    include_re = _compile_optional_regex(include_filename_regex, ignore_case=filename_regex_ignore_case)
    exclude_re = _compile_optional_regex(exclude_filename_regex, ignore_case=filename_regex_ignore_case)
    if include_re:
        files = [p for p in files if include_re.search(p.name)]
    if exclude_re:
        files = [p for p in files if not exclude_re.search(p.name)]

    # Prefer .java content for header scan, but allow other files for listing
    java_like = [p for p in files if p.suffix.lower() == ".java"]
    scan_candidates = java_like if java_like else files

    for jf in scan_candidates:
        sid, name = extract_student_info_from_file(jf)
        if sid and not detected_id:
            detected_id = sid
        if name and not detected_name:
            detected_name = name
        if detected_id and detected_name:
            break

    # Optional folder regex fallback (disabled by default; folder names are often machine IDs).
    folder_id = ""
    folder_name = ""
    if folder_id_regex or folder_name_regex:
        folder_id, folder_name = try_extract_from_folder_name(folder, folder_id_regex, folder_name_regex)
        if not detected_id and folder_id:
            detected_id = folder_id
        if not detected_name and folder_name:
            detected_name = folder_name

    final_id = None
    final_name = None

    if detected_id or detected_name:
        if detected_id:
            final_name = (detected_name or "").strip()
            final_id = build_student_key(detected_id, final_name)
        else:
            final_name = (detected_name or "").strip()
            final_id = f"NAME:{final_name}" if final_name else ""

    return final_id, final_name, (detected_id or ""), (detected_name or ""), [str(p) for p in files]


def build_student_key(student_id: str, student_name: str) -> str:
    """
    Build a stable student key.
    Numeric IDs are always stored as numeric-only keys.
    """
    raw_sid = (student_id or "").strip()
    if raw_sid.startswith("NAME:"):
        return raw_sid
    if raw_sid.lower() == "full":
        return raw_sid

    sid = extract_numeric_id(raw_sid)
    sname = re.sub(r"\s+", " ", (student_name or "").strip())

    if not sid:
        return f"NAME:{sname}" if sname else ""
    return sid


def has_required_student_fields(student_id: str, student_name: str) -> bool:
    """
    Treat rows as valid student-like records when:
    - numeric student ID + non-empty student name, OR
    - the special FULL aggregate row.
    """
    raw_sid = (student_id or "").strip()
    sname = re.sub(r"\s+", " ", (student_name or "").strip())

    if raw_sid.lower() == "full":
        return True

    sid = extract_numeric_id(raw_sid)
    if not sid or not sname:
        return False
    if sname.lower() in {"unknown student", "unknown"}:
        return False
    return True


class FolderScannerBase:
    """
    Base scanner so folder scanning can be customized via inheritance later.
    """
    def __init__(self, root_folder: Path, file_globs: list[str], include_filename_regex: str,
                 exclude_filename_regex: str, filename_regex_ignore_case: bool,
                 folder_id_regex: str, folder_name_regex: str):
        self.root_folder = root_folder
        self.file_globs = file_globs or ["*.java"]
        self.include_filename_regex = include_filename_regex
        self.exclude_filename_regex = exclude_filename_regex
        self.filename_regex_ignore_case = filename_regex_ignore_case
        self.folder_id_regex = folder_id_regex
        self.folder_name_regex = folder_name_regex

    def collect_folders(self) -> list[Path]:
        folders: list[Path] = []
        seen: set[str] = set()

        for folder in [self.root_folder] + [p for p in self.root_folder.rglob("*") if p.is_dir()]:
            has_match = False
            for g in self.file_globs:
                if any(folder.glob(g)):
                    has_match = True
                    break
            if not has_match:
                continue

            key = str(folder)
            if key in seen:
                continue
            seen.add(key)
            folders.append(folder)

        return folders

    def detect_folder(self, folder: Path):
        return infer_student_for_folder(
            folder,
            file_globs=self.file_globs,
            include_filename_regex=self.include_filename_regex,
            exclude_filename_regex=self.exclude_filename_regex,
            filename_regex_ignore_case=self.filename_regex_ignore_case,
            folder_id_regex=self.folder_id_regex,
            folder_name_regex=self.folder_name_regex,
        )


class DefaultFolderScanner(FolderScannerBase):
    pass


# =============================================================================
# 3) Submissions DB
# =============================================================================

def db_connect(db_path: Path) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    con.execute("PRAGMA foreign_keys = ON;")
    return con

def submissions_db_init(con: sqlite3.Connection):
    con.executescript("""
    CREATE TABLE IF NOT EXISTS app_meta (
        meta_key TEXT PRIMARY KEY,
        meta_value TEXT
    );

    CREATE TABLE IF NOT EXISTS students (
        student_id TEXT PRIMARY KEY,
        student_name TEXT NOT NULL,
        lab_id TEXT,
        folder_path TEXT,
        included INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS files (
        file_path TEXT PRIMARY KEY,
        student_id TEXT,
        source_folder TEXT,
        detected_id TEXT,
        detected_name TEXT,
        file_hash TEXT,
        file_content TEXT,
        last_seen TEXT,
      FOREIGN KEY(student_id) REFERENCES students(student_id) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS regex_profiles (
        profile_name TEXT PRIMARY KEY,
        settings_json TEXT NOT NULL,
        updated_at TEXT
    );

    CREATE TABLE IF NOT EXISTS scan_sessions (
        session_id INTEGER PRIMARY KEY AUTOINCREMENT,
        committed_at TEXT,
        root_folder TEXT,
        lab_id TEXT,
        regex_profile_name TEXT,
        regex_settings_json TEXT,
        session_payload_json TEXT
    );
    """)
    # Safe migration
    try:
        cols = [r[1] for r in con.execute("PRAGMA table_info(students)").fetchall()]
        if "lab_id" not in cols:
            con.execute("ALTER TABLE students ADD COLUMN lab_id TEXT;")
        if "included" not in cols:
            con.execute("ALTER TABLE students ADD COLUMN included INTEGER NOT NULL DEFAULT 1;")
    except Exception:
        pass
    con.commit()

def sub_meta_set(con: sqlite3.Connection, key: str, value: str):
    con.execute("""
      INSERT INTO app_meta(meta_key, meta_value)
      VALUES(?, ?)
      ON CONFLICT(meta_key) DO UPDATE SET meta_value=excluded.meta_value
    """, (key, value))
    con.commit()

def sub_meta_get(con: sqlite3.Connection, key: str, default: str = "") -> str:
    row = con.execute("SELECT meta_value FROM app_meta WHERE meta_key=?", (key,)).fetchone()
    return row[0] if row and row[0] is not None else default

def upsert_regex_profile(con: sqlite3.Connection, profile_name: str, payload: dict):
    con.execute(
        """
        INSERT INTO regex_profiles(profile_name, settings_json, updated_at)
        VALUES(?, ?, ?)
        ON CONFLICT(profile_name) DO UPDATE SET
          settings_json=excluded.settings_json,
          updated_at=excluded.updated_at
        """,
        (profile_name, json.dumps(payload, ensure_ascii=False), now_ts()),
    )
    con.commit()

def load_regex_profile(con: sqlite3.Connection, profile_name: str) -> dict | None:
    row = con.execute("SELECT settings_json FROM regex_profiles WHERE profile_name=?", (profile_name,)).fetchone()
    if not row:
        return None
    try:
        return json.loads(row[0])
    except Exception:
        return None

def list_regex_profiles(con: sqlite3.Connection) -> list[str]:
    rows = con.execute("SELECT profile_name FROM regex_profiles ORDER BY profile_name").fetchall()
    return [r[0] for r in rows]

def commit_scan_session(con: sqlite3.Connection, root_folder: str, lab_id: str, profile_name: str, profile_payload: dict, session_payload: dict):
    con.execute(
        """
        INSERT INTO scan_sessions(committed_at, root_folder, lab_id, regex_profile_name, regex_settings_json, session_payload_json)
        VALUES(?, ?, ?, ?, ?, ?)
        """,
        (
            now_ts(),
            root_folder,
            lab_id,
            profile_name,
            json.dumps(profile_payload, ensure_ascii=False),
            json.dumps(session_payload, ensure_ascii=False),
        ),
    )
    con.commit()

def upsert_student(con: sqlite3.Connection, student_id: str, student_name: str, lab_id: str | None, folder_path: str | None, included: bool = True, commit: bool = True):
    con.execute("""
    INSERT INTO students(student_id, student_name, lab_id, folder_path, included)
    VALUES(?, ?, ?, ?, ?)
    ON CONFLICT(student_id) DO UPDATE SET
      student_name=excluded.student_name,
      lab_id=COALESCE(excluded.lab_id, students.lab_id),
      folder_path=COALESCE(excluded.folder_path, students.folder_path),
      included=excluded.included
    """, (student_id, student_name, lab_id, folder_path, 1 if included else 0))
    if commit:
        con.commit()


def set_student_included(con: sqlite3.Connection, student_id: str, included: bool, commit: bool = True):
    con.execute("UPDATE students SET included=? WHERE student_id=?", (1 if included else 0, student_id))
    if commit:
        con.commit()

def upsert_file(con: sqlite3.Connection, file_path: str, student_id: str | None,
                source_folder: str | None,
                detected_id: str | None, detected_name: str | None,
                file_hash: str | None, file_content: str | None,
                commit: bool = True):
    con.execute("""
    INSERT INTO files(file_path, student_id, source_folder, detected_id, detected_name, file_hash, file_content, last_seen)
    VALUES(?, ?, ?, ?, ?, ?, ?, ?)
    ON CONFLICT(file_path) DO UPDATE SET
      student_id=excluded.student_id,
      source_folder=excluded.source_folder,
      detected_id=excluded.detected_id,
      detected_name=excluded.detected_name,
      file_hash=excluded.file_hash,
      file_content=excluded.file_content,
      last_seen=excluded.last_seen
    """, (file_path, student_id, source_folder, detected_id, detected_name, file_hash, file_content, now_ts()))
    if commit:
        con.commit()

def get_students(con: sqlite3.Connection):
    cur = con.execute("""
    SELECT s.student_id, s.student_name, COALESCE(s.lab_id,''),
           (SELECT COUNT(*) FROM files f WHERE f.student_id = s.student_id) AS file_count
    FROM students s
    WHERE LOWER(s.student_id)='full' OR s.included=1
    ORDER BY CASE WHEN LOWER(s.student_id)='full' OR LOWER(s.student_name)='full' THEN 0 ELSE 1 END,
             s.student_id
    """)
    return cur.fetchall()

def get_student_files(con: sqlite3.Connection, student_id: str):
    cur = con.execute("""
    SELECT file_path FROM files
    WHERE student_id = ?
    ORDER BY file_path
    """, (student_id,))
    return [r[0] for r in cur.fetchall()]

def get_file_content(con: sqlite3.Connection, file_path: str) -> str | None:
    row = con.execute("SELECT file_content FROM files WHERE file_path=?", (file_path,)).fetchone()
    return row[0] if row else None

def merge_student_code(sub_con: sqlite3.Connection, student_id: str) -> str:
    files = get_student_files(sub_con, student_id)
    parts = []
    for fp in files:
        content = get_file_content(sub_con, fp)
        if content is None:
            try:
                content = Path(fp).read_text(encoding="utf-8", errors="ignore")
            except Exception:
                content = ""
        parts.append(f"\n\n// ===== FILE: {Path(fp).name} =====\n{content}")
    return "\n".join(parts)


# =============================================================================
# 4) Grading DB (rubric + assignments + notes + code comments)
# =============================================================================

def grading_db_init(con: sqlite3.Connection):
    con.executescript("""
    PRAGMA foreign_keys = ON;

    CREATE TABLE IF NOT EXISTS meta(
      meta_key TEXT PRIMARY KEY,
      meta_value TEXT
    );

    CREATE TABLE IF NOT EXISTS rubric_questions(
      question_id TEXT PRIMARY KEY,
      question_title TEXT,
      sub_id TEXT
    );

    CREATE TABLE IF NOT EXISTS rubric_columns(
      question_id TEXT NOT NULL,
      col_key TEXT NOT NULL,
      col_group TEXT,
      col_text TEXT NOT NULL,
      col_max REAL NOT NULL,
      col_order INTEGER NOT NULL,
      PRIMARY KEY(question_id, col_key),
      FOREIGN KEY(question_id) REFERENCES rubric_questions(question_id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS rubric_scores(
      student_id TEXT NOT NULL,
      question_id TEXT NOT NULL,
      col_key TEXT NOT NULL,
      points REAL,
      note TEXT,
      updated_at TEXT,
      PRIMARY KEY(student_id, question_id, col_key)
    );

    CREATE TABLE IF NOT EXISTS student_notes(
      student_id TEXT NOT NULL,
      question_id TEXT NOT NULL,
      rationale TEXT,
      overall_grade REAL,
      updated_at TEXT,
      PRIMARY KEY(student_id, question_id)
    );

    CREATE TABLE IF NOT EXISTS student_assignments(
      student_id TEXT PRIMARY KEY,
      q1 TEXT,
      q2 TEXT,
      updated_at TEXT
    );

    CREATE TABLE IF NOT EXISTS grading_progress(
      student_id TEXT PRIMARY KEY,
      graded INTEGER NOT NULL DEFAULT 0,
      reviewed INTEGER NOT NULL DEFAULT 1,
      first_graded_at TEXT,
      last_updated_at TEXT,
      last_question_id TEXT
    );

    CREATE TABLE IF NOT EXISTS code_comments(
      comment_id INTEGER PRIMARY KEY AUTOINCREMENT,
      student_id TEXT NOT NULL,
      file_path TEXT NOT NULL,
      start_index TEXT NOT NULL,
      end_index TEXT NOT NULL,
      comment_text TEXT NOT NULL,
      color TEXT,
      created_at TEXT
    );
    """)
    try:
        cols = [r[1] for r in con.execute("PRAGMA table_info(rubric_questions)").fetchall()]
        if "sub_id" not in cols:
            con.execute("ALTER TABLE rubric_questions ADD COLUMN sub_id TEXT;")
    except Exception:
        pass
    try:
        cols = [r[1] for r in con.execute("PRAGMA table_info(grading_progress)").fetchall()]
        if "reviewed" not in cols:
            con.execute("ALTER TABLE grading_progress ADD COLUMN reviewed INTEGER NOT NULL DEFAULT 1;")
    except Exception:
        pass
    con.commit()

def meta_set(con: sqlite3.Connection, key: str, value: str):
    con.execute("""
      INSERT INTO meta(meta_key, meta_value)
      VALUES(?, ?)
      ON CONFLICT(meta_key) DO UPDATE SET meta_value=excluded.meta_value
    """, (key, value))
    con.commit()

def meta_get(con: sqlite3.Connection, key: str, default: str = "") -> str:
    row = con.execute("SELECT meta_value FROM meta WHERE meta_key=?", (key,)).fetchone()
    return row[0] if row and row[0] is not None else default

def wipe_rubric(con: sqlite3.Connection):
    con.execute("DELETE FROM rubric_scores")
    con.execute("DELETE FROM student_notes")
    con.execute("DELETE FROM rubric_columns")
    con.execute("DELETE FROM rubric_questions")
    con.commit()

def load_scheme_rows_into_db(con: sqlite3.Connection, rows: list[dict], source_label: str = ""):
    required = {"question_id", "question_title", "group", "col_key", "col_text", "col_max", "col_order"}
    wipe_rubric(con)

    questions = {}
    parsed_rows = []
    for r in rows:
        qid = (r.get("question_id") or "").strip()
        qtitle = (r.get("question_title") or "").strip()
        sub_id = (r.get("sub_id") or "").strip()
        group = (r.get("group") or "").strip()
        col_key = (r.get("col_key") or "").strip()
        col_text = (r.get("col_text") or "").strip()
        if not qid or not col_key or not col_text:
            continue
        try:
            col_max = float(r.get("col_max") or 0)
        except Exception:
            col_max = 0.0
        try:
            col_order = int(float(r.get("col_order") or 0))
        except Exception:
            col_order = 0
        questions.setdefault(qid, {"title": qtitle or qid, "sub_id": sub_id})
        parsed_rows.append((qid, col_key, group, col_text, col_max, col_order))

    for qid, qd in questions.items():
        con.execute(
            "INSERT OR REPLACE INTO rubric_questions(question_id, question_title, sub_id) VALUES(?,?,?)",
            (qid, qd["title"], qd.get("sub_id", "")),
        )

    for qid, col_key, group, col_text, col_max, col_order in parsed_rows:
        con.execute(
            """
            INSERT OR REPLACE INTO rubric_columns(question_id, col_key, col_group, col_text, col_max, col_order)
            VALUES(?,?,?,?,?,?)
            """,
            (qid, col_key, group, col_text, col_max, col_order),
        )

    if source_label:
        meta_set(con, "scheme_csv_path", source_label)
    con.commit()

def load_scheme_csv_into_db(con: sqlite3.Connection, csv_path: Path):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"question_id", "question_title", "group", "col_key", "col_text", "col_max", "col_order"}
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")
        missing = required - set([h.strip() for h in reader.fieldnames])
        if missing:
            raise ValueError(f"CSV missing columns: {', '.join(sorted(missing))}")

        wipe_rubric(con)

        has_sub_id = any((h or "").strip() == "sub_id" for h in reader.fieldnames)

        questions = {}
        rows = []
        for r in reader:
            qid = (r.get("question_id") or "").strip()
            qtitle = (r.get("question_title") or "").strip()
            sub_id = (r.get("sub_id") or "").strip() if has_sub_id else ""
            group = (r.get("group") or "").strip()
            col_key = (r.get("col_key") or "").strip()
            col_text = (r.get("col_text") or "").strip()

            if not qid or not col_key or not col_text:
                continue

            try:
                col_max = float((r.get("col_max") or "").strip())
            except Exception:
                col_max = 0.0

            try:
                col_order = int(float((r.get("col_order") or "").strip()))
            except Exception:
                col_order = 0

            questions[qid] = (qtitle or qid, sub_id)
            rows.append((qid, col_key, group, col_text, col_max, col_order))

        for qid, payload in questions.items():
            title, sub_id = payload
            con.execute("""
              INSERT OR REPLACE INTO rubric_questions(question_id, question_title, sub_id)
              VALUES(?, ?, ?)
            """, (qid, title, sub_id))

        for qid, col_key, group, col_text, col_max, col_order in rows:
            con.execute("""
              INSERT OR REPLACE INTO rubric_columns(question_id, col_key, col_group, col_text, col_max, col_order)
              VALUES(?,?,?,?,?,?)
            """, (qid, col_key, group, col_text, col_max, col_order))

        con.commit()
        meta_set(con, "scheme_csv_path", str(csv_path))

def fetch_questions(con: sqlite3.Connection):
    return con.execute("""
      SELECT question_id, question_title, COALESCE(sub_id, '')
      FROM rubric_questions
      ORDER BY question_id
    """).fetchall()

def fetch_columns_for_question(con: sqlite3.Connection, question_id: str):
    return con.execute("""
      SELECT col_key, col_group, col_text, col_max
      FROM rubric_columns
      WHERE question_id=?
      ORDER BY col_order, col_key
    """, (question_id,)).fetchall()

def load_student_scores(con: sqlite3.Connection, student_id: str, question_id: str):
    rows = con.execute("""
      SELECT col_key, points, note
      FROM rubric_scores
      WHERE student_id=? AND question_id=?
    """, (student_id, question_id)).fetchall()
    score_map, note_map = {}, {}
    for k, p, n in rows:
        score_map[k] = p
        note_map[k] = n
    return score_map, note_map

def upsert_score(con: sqlite3.Connection, student_id: str, question_id: str, col_key: str, points: float | None, note: str, commit: bool = True):
    con.execute("""
      INSERT INTO rubric_scores(student_id, question_id, col_key, points, note, updated_at)
      VALUES(?,?,?,?,?,?)
      ON CONFLICT(student_id, question_id, col_key) DO UPDATE SET
        points=excluded.points,
        note=excluded.note,
        updated_at=excluded.updated_at
    """, (student_id, question_id, col_key, points, note, now_ts()))
    if commit:
        con.commit()

def compute_total(con: sqlite3.Connection, student_id: str, question_id: str | None) -> float:
    if not question_id:
        return 0.0
    row = con.execute("""
      SELECT COALESCE(SUM(COALESCE(points,0)),0)
      FROM rubric_scores
      WHERE student_id=? AND question_id=?
    """, (student_id, question_id)).fetchone()
    return float(row[0] if row else 0.0)

def build_question_display_map(con: sqlite3.Connection) -> dict[str, str]:
    rows = con.execute("SELECT question_id, COALESCE(sub_id, '') FROM rubric_questions ORDER BY question_id").fetchall()
    out = {}
    for qid, sub_id in rows:
        qid_clean = (qid or "").strip()
        sub_clean = (sub_id or "").strip()
        if not qid_clean:
            continue
        out[qid_clean] = sub_clean or qid_clean
    return out

def compute_total_by_display_id(con: sqlite3.Connection, student_id: str, display_question_id: str) -> float:
    if not display_question_id:
        return 0.0
    row = con.execute("""
      SELECT COALESCE(SUM(COALESCE(rs.points,0)),0)
      FROM rubric_scores rs
      JOIN rubric_questions rq ON rq.question_id = rs.question_id
      WHERE rs.student_id=? AND COALESCE(NULLIF(TRIM(rq.sub_id),''), rq.question_id)=?
    """, (student_id, display_question_id)).fetchone()
    return float(row[0] if row else 0.0)

def fetch_display_question_ids(con: sqlite3.Connection):
    rows = con.execute("""
      SELECT DISTINCT COALESCE(NULLIF(TRIM(sub_id),''), question_id) AS display_id
      FROM rubric_questions
      ORDER BY display_id
    """).fetchall()
    return [r[0] for r in rows if r and r[0]]

def compute_overall_total(con: sqlite3.Connection, student_id: str) -> float:
    row = con.execute("""
      SELECT COALESCE(SUM(COALESCE(points,0)),0)
      FROM rubric_scores
      WHERE student_id=?
    """, (student_id,)).fetchone()
    return float(row[0] if row else 0.0)
def upsert_student_note(con: sqlite3.Connection, student_id: str, question_id: str, rationale: str, overall_grade: float | None, commit: bool = True):
    con.execute("""
      INSERT INTO student_notes(student_id, question_id, rationale, overall_grade, updated_at)
      VALUES(?,?,?,?,?)
      ON CONFLICT(student_id, question_id) DO UPDATE SET
        rationale=excluded.rationale,
        overall_grade=excluded.overall_grade,
        updated_at=excluded.updated_at
    """, (student_id, question_id, rationale, overall_grade, now_ts()))
    if commit:
        con.commit()

def load_student_note(con: sqlite3.Connection, student_id: str, question_id: str):
    return con.execute("""
      SELECT rationale, overall_grade
      FROM student_notes
      WHERE student_id=? AND question_id=?
    """, (student_id, question_id)).fetchone()

def upsert_student_assignment(con: sqlite3.Connection, student_id: str, q1: str | None, q2: str | None):
    con.execute("""
      INSERT INTO student_assignments(student_id, q1, q2, updated_at)
      VALUES(?,?,?,?)
      ON CONFLICT(student_id) DO UPDATE SET
        q1=excluded.q1, q2=excluded.q2, updated_at=excluded.updated_at
    """, (student_id, q1, q2, now_ts()))
    con.commit()

def load_student_assignment(con: sqlite3.Connection, student_id: str) -> tuple[str | None, str | None]:
    row = con.execute("SELECT q1, q2 FROM student_assignments WHERE student_id=?", (student_id,)).fetchone()
    if not row:
        return None, None
    return (row[0] or None), (row[1] or None)


def upsert_grading_progress(con: sqlite3.Connection, student_id: str, question_id: str | None = None,
                           mark_graded: bool = True, reviewed: bool | None = None, commit: bool = True):
    now = now_ts()
    reviewed_val = None if reviewed is None else (1 if reviewed else 0)
    con.execute("""
      INSERT INTO grading_progress(student_id, graded, reviewed, first_graded_at, last_updated_at, last_question_id)
      VALUES(?,?,?,?,?,?)
      ON CONFLICT(student_id) DO UPDATE SET
        graded=CASE WHEN excluded.graded=1 THEN 1 ELSE grading_progress.graded END,
        reviewed=COALESCE(excluded.reviewed, grading_progress.reviewed),
        first_graded_at=CASE
          WHEN grading_progress.first_graded_at IS NULL OR grading_progress.first_graded_at='' THEN
            CASE WHEN excluded.graded=1 THEN excluded.first_graded_at ELSE grading_progress.first_graded_at END
          ELSE grading_progress.first_graded_at
        END,
        last_updated_at=excluded.last_updated_at,
        last_question_id=COALESCE(excluded.last_question_id, grading_progress.last_question_id)
    """, (student_id, 1 if mark_graded else 0, reviewed_val, now if mark_graded else None, now, question_id))
    if commit:
        con.commit()


def set_student_graded_flag(con: sqlite3.Connection, student_id: str, graded: bool, commit: bool = True):
    now = now_ts()
    con.execute("""
      INSERT INTO grading_progress(student_id, graded, first_graded_at, last_updated_at)
      VALUES(?,?,?,?)
      ON CONFLICT(student_id) DO UPDATE SET
        graded=excluded.graded,
        first_graded_at=CASE
          WHEN excluded.graded=1 AND (grading_progress.first_graded_at IS NULL OR grading_progress.first_graded_at='') THEN excluded.first_graded_at
          WHEN excluded.graded=0 THEN NULL
          ELSE grading_progress.first_graded_at
        END,
        last_updated_at=excluded.last_updated_at
    """, (student_id, 1 if graded else 0, now if graded else None, now))
    if commit:
        con.commit()


def load_grading_progress(con: sqlite3.Connection, student_id: str):
    row = con.execute("""
      SELECT graded, COALESCE(reviewed,1), COALESCE(first_graded_at,''), COALESCE(last_updated_at,''), COALESCE(last_question_id,'')
      FROM grading_progress WHERE student_id=?
    """, (student_id,)).fetchone()
    if not row:
        return (0, 1, "", "", "")
    return row

def set_student_reviewed_flag(con: sqlite3.Connection, student_id: str, reviewed: bool, commit: bool = True):
    now = now_ts()
    con.execute("""
      INSERT INTO grading_progress(student_id, reviewed, last_updated_at)
      VALUES(?,?,?)
      ON CONFLICT(student_id) DO UPDATE SET
        reviewed=excluded.reviewed,
        last_updated_at=excluded.last_updated_at
    """, (student_id, 1 if reviewed else 0, now))
    if commit:
        con.commit()

def _parse_tk_index_value(idx: str) -> tuple[int, int]:
    try:
        line, col = str(idx).split(".", 1)
        return max(1, int(line)), max(0, int(col))
    except Exception:
        return 1, 0


def _line_highlight_range(start_index: str, end_index: str) -> tuple[str, str]:
    """
    Convert any Tk text index range into whole-line Tk index bounds.
    This keeps review highlights stable even when character columns are noisy.
    """
    start, end = _normalize_index_range(start_index, end_index)
    s_line = max(1, int(start[0]))
    e_line = max(s_line, int(end[0]))
    return f"{s_line}.0", f"{e_line}.end"


def _format_comment_range_label(start_index: str, end_index: str) -> str:
    """
    Human-readable range label that clarifies line bounds and end-of-line selections.
    Tk Text end indices are exclusive, so line+1.0 is presented as previous line end.
    """
    start, end = _normalize_index_range(start_index, end_index)
    s_line, s_col = start
    e_line, e_col = end

    # Tk selections are end-exclusive. A boundary at col 0 usually means
    # the previous line's end should be highlighted.
    if e_col == 0 and e_line > s_line:
        return f"Line {s_line}:{s_col} to line {e_line - 1}:end"
    return f"Line {s_line}:{s_col} to line {e_line}:{e_col}"

def _normalize_index_range(start_index: str, end_index: str) -> tuple[tuple[int, int], tuple[int, int]]:
    start = _parse_tk_index_value(start_index)
    end = _parse_tk_index_value(end_index)
    if end < start:
        start, end = end, start
    return start, end

def _ranges_overlap(a_start: tuple[int, int], a_end: tuple[int, int], b_start: tuple[int, int], b_end: tuple[int, int]) -> bool:
    return not (a_end <= b_start or a_start >= b_end)

def add_code_comment(con: sqlite3.Connection, student_id: str, file_path: str, start_index: str, end_index: str,
                     comment_text: str, color: str = "#FFF2B2"):
    start, end = _normalize_index_range(start_index, end_index)
    sidx = f"{start[0]}.{start[1]}"
    eidx = f"{end[0]}.{end[1]}"
    con.execute("""
      INSERT INTO code_comments(student_id, file_path, start_index, end_index, comment_text, color, created_at)
      VALUES(?,?,?,?,?,?,?)
    """, (student_id, file_path, sidx, eidx, comment_text, color, now_ts()))
    con.commit()

def delete_code_comments_in_range(con: sqlite3.Connection, student_id: str, file_path: str, start_index: str, end_index: str):
    sel_start, sel_end = _normalize_index_range(start_index, end_index)
    rows = con.execute("""
      SELECT comment_id, start_index, end_index
      FROM code_comments
      WHERE student_id=? AND file_path=?
    """, (student_id, file_path)).fetchall()

    delete_ids = []
    for comment_id, sidx, eidx in rows:
        c_start, c_end = _normalize_index_range(sidx, eidx)
        if _ranges_overlap(c_start, c_end, sel_start, sel_end):
            delete_ids.append(comment_id)

    if delete_ids:
        con.executemany("DELETE FROM code_comments WHERE comment_id=?", [(cid,) for cid in delete_ids])
    con.commit()

def fetch_code_comments_for_file(con: sqlite3.Connection, student_id: str, file_path: str):
    return con.execute("""
      SELECT comment_id, start_index, end_index, comment_text, color, created_at
      FROM code_comments
      WHERE student_id=? AND file_path=?
      ORDER BY comment_id
    """, (student_id, file_path)).fetchall()

def fetch_code_comments_for_student(con: sqlite3.Connection, student_id: str):
    return con.execute("""
      SELECT file_path, start_index, end_index, comment_text, color, created_at
      FROM code_comments
      WHERE student_id=?
      ORDER BY file_path, comment_id
    """, (student_id,)).fetchall()


# =============================================================================
# 5) Assignment detection heuristic (kept)
# =============================================================================

def detect_assigned_questions(merged_code: str) -> list[str]:
    s = (merged_code or "").lower()

    tier2_signals = [
        "toofar(", "too far", "pixel", "getpixel", "setpixel", "getred", "getgreen", "getblue",
        "distance", "color distance", "colour distance", "swap", "one pixel"
    ]
    tier1_signals = [
        "drawstar(", "drawshape(", "random", "getsrandom", "getrandomints", "getrandomint",
        "color", "colour", "polygon", "star"
    ]

    tier2_score = sum(1 for k in tier2_signals if k in s)
    tier1_score = sum(1 for k in tier1_signals if k in s)

    if tier2_score >= 2 and tier2_score >= tier1_score:
        return ["P21", "P22"]
    return ["P11", "P12"]


# =============================================================================
# 6) Export: Excel
# =============================================================================

def fetch_all_question_ids(con: sqlite3.Connection):
    rows = con.execute("SELECT question_id FROM rubric_questions ORDER BY question_id").fetchall()
    return [r[0] for r in rows]


def fetch_rubric_parts(con: sqlite3.Connection):
    rows = con.execute("""
      SELECT question_id, col_key, COALESCE(col_group,''), col_text, col_max
      FROM rubric_columns
      ORDER BY question_id, col_order, col_key
    """).fetchall()
    return rows

def compute_question_max(con: sqlite3.Connection, question_id: str) -> float:
    row = con.execute("""
      SELECT COALESCE(SUM(rc.col_max),0)
      FROM rubric_columns rc
      JOIN rubric_questions rq ON rq.question_id = rc.question_id
      WHERE COALESCE(NULLIF(TRIM(rq.sub_id),''), rq.question_id)=?
    """, (question_id,)).fetchone()
    return float(row[0] if row else 0.0)

def export_all_to_excel(
    sub_con: sqlite3.Connection,
    grade_con: sqlite3.Connection,
    out_path: Path,
    student_filter: set[str] | None = None,
):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Brightspace_Summary"

    question_ids = fetch_display_question_ids(grade_con)
    qmax = {qid: compute_question_max(grade_con, qid) for qid in question_ids}

    parts = fetch_rubric_parts(grade_con)
    q_display_map = build_question_display_map(grade_con)
    part_headers = [f"{q_display_map.get(qid, qid)}:{ck}" for qid, ck, _g, _t, _m in parts]

    ws_sum.append(["Student ID", "Student Name", "LabID", *part_headers, *[f"{qid}_total" for qid in question_ids], "Overall raw"])

    students = sub_con.execute("""
      SELECT student_id, student_name, COALESCE(lab_id,'')
      FROM students
      WHERE included=1 OR LOWER(student_id)='full'
      ORDER BY student_id
    """).fetchall()

    assessed_students = [r for r in students if not is_full_student(r[0])]
    if student_filter is not None:
        assessed_students = [r for r in assessed_students if r[0] in student_filter]

    for sid, sname, lab in assessed_students:
        row = [sid, sname, lab]
        score_cache = {}
        for qid, ck, _g, _t, _m in parts:
            if qid not in score_cache:
                score_cache[qid] = load_student_scores(grade_con, sid, qid)[0]
            val = score_cache[qid].get(ck)
            row.append("" if val is None else val)
        for qid in question_ids:
            row.append(compute_total_by_display_id(grade_con, sid, qid))
        row.append(sum(compute_total_by_display_id(grade_con, sid, qid) for qid in question_ids))
        ws_sum.append(row)

    # Per-question sheets (grouped by display/sub id)
    raw_qids = fetch_all_question_ids(grade_con)
    for qid in question_ids:
        member_qids = [rq for rq in raw_qids if q_display_map.get(rq, rq) == qid] or [qid]
        title_row = grade_con.execute(
            "SELECT question_title FROM rubric_questions WHERE question_id=?",
            (member_qids[0],)
        ).fetchone()
        qtitle = title_row[0] if title_row else qid

        ws = wb.create_sheet(title=f"{qid}")
        ws.append([f"{qid} — {qtitle}"])
        ws.append([])

        cols = []
        for mqid in member_qids:
            cols.extend([(mqid, *c) for c in fetch_columns_for_question(grade_con, mqid)])

        header = ["Student ID", "Student Name", "LabID", "Total", "Rationale"]
        for mqid, col_key, group, text, mx in cols:
            header.append(f"{mqid} | {(group or '').strip()} | {text} (/ {mx:g})".strip(" |"))
            header.append("Note")
        ws.append(header)

        for sid, sname, lab in assessed_students:
            total = compute_total_by_display_id(grade_con, sid, qid)
            note_row = load_student_note(grade_con, sid, member_qids[0])
            rationale = note_row[0] if note_row and note_row[0] else ""

            row = [sid, sname, lab, total, rationale]
            for mqid, col_key, _group, _text, _mx in cols:
                score_map, note_map = load_student_scores(grade_con, sid, mqid)
                row.append("" if score_map.get(col_key) is None else score_map.get(col_key))
                row.append(note_map.get(col_key, "") or "")
            ws.append(row)

    # Code comment sheet (highlighted ranges + comment text)
    ws_comments = wb.create_sheet(title="Code_Comments")
    ws_comments.append(["Student ID", "Student Name", "Code (file)", "Highlighted part", "Comment"])
    for sid, sname, _lab in assessed_students:
        for fp, sidx, eidx, txt, _color, _ts in fetch_code_comments_for_student(grade_con, sid):
            ws_comments.append([sid, sname, Path(fp).name, _format_comment_range_label(sidx, eidx), txt or ""])

    # Reference model row if FULL exists
    full_row = next((r for r in students if is_full_student(r[0])), None)
    if full_row and student_filter is None:
        sid, sname, lab = full_row
        ws_full = wb.create_sheet(title="FULL_Model")
        ws_full.append(["Student ID", "Student Name", "LabID", "Overall raw"])
        ws_full.append([sid, sname, lab, sum(compute_total_by_display_id(grade_con, sid, qid) for qid in question_ids)])

        ws_full.append([])
        ws_full.append(["Question", "Total", "Rationale"])
        for qid in question_ids:
            qtotal = compute_total_by_display_id(grade_con, sid, qid)
            nrow = load_student_note(grade_con, sid, qid)
            ws_full.append([qid, qtotal, nrow[0] if nrow and nrow[0] else ""])

    # light formatting
    for ws in wb.worksheets:
        try:
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(48, max(12, len(str(col[0].value or "")) + 2))
        except Exception:
            pass

    wb.save(out_path)


# =============================================================================
# 7) PDF Export: student + summary + batch
# =============================================================================

class PDFExporter:
    def __init__(
        self,
        sub_con: sqlite3.Connection,
        grade_con: sqlite3.Connection,
        question_map: dict[str, str],
        student_pdf_options: dict | None = None,
    ):
        self.sub_con = sub_con
        self.grade_con = grade_con
        self.question_map = question_map
        opts = student_pdf_options or {}
        self.student_pdf_options = {
            "include_grade_tables": bool(opts.get("include_grade_tables", True)),
            "include_rationale": bool(opts.get("include_rationale", True)),
            "include_comment_highlights": bool(opts.get("include_comment_highlights", True)),
            "include_full_code_listing": bool(opts.get("include_full_code_listing", True)),
        }

    def _build_annotated_code_injected(self, sid: str) -> str:
        """
        Annotated code dump: injects comment blocks right before the line where comment starts.
        Uses start_index's line number from Tk indices like '12.0'.
        """
        out = []
        files = get_student_files(self.sub_con, sid)

        for fp in files:
            code = get_file_content(self.sub_con, fp)
            if code is None:
                try:
                    code = Path(fp).read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    code = ""

            comments = self.grade_con.execute("""
              SELECT comment_id, start_index, end_index, comment_text, created_at
              FROM code_comments
              WHERE student_id=? AND file_path=?
              ORDER BY comment_id
            """, (sid, fp)).fetchall()

            by_line = {}
            for cid, sidx, eidx, txt, ts in comments:
                try:
                    line = int(str(sidx).split(".", 1)[0])
                except Exception:
                    line = 1
                by_line.setdefault(line, []).append((cid, sidx, eidx, txt, ts))

            lines = code.splitlines()
            out.append(f"\n\n// ===== FILE: {Path(fp).name} =====\n")

            for i, line_text in enumerate(lines, start=1):
                if i in by_line:
                    for (cid, sidx, eidx, txt, ts) in by_line[i]:
                        out.append(f"// >>> COMMENT #{cid} [{sidx}–{eidx}] ({ts}):")
                        for c_line in (txt or "").splitlines():
                            out.append(f"//     {c_line}")
                        out.append("// >>> END COMMENT\n")
                out.append(line_text)

        return "\n".join(out)

    def _escape_pdf_text(self, text: str) -> str:
        return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _parse_tk_index(self, idx: str) -> tuple[int, int]:
        try:
            line, col = str(idx).split(".", 1)
            return max(1, int(line)), max(0, int(col))
        except Exception:
            return 1, 0

    def _tk_index_to_offset(self, line_starts: list[int], code_len: int, idx: str) -> int:
        line_no, col = self._parse_tk_index(idx)
        if not line_starts:
            return 0
        line_no = min(max(1, line_no), len(line_starts))
        base = line_starts[line_no - 1]
        return min(code_len, max(0, base + col))

    def _render_line_with_highlights(self, line_text: str, line_start: int, line_end: int,
                                     ranges: list[tuple[int, int]], line_no: int) -> str:
        if line_start >= line_end:
            return f"{line_no:04d} |"

        overlaps: list[tuple[int, int]] = []
        for r0, r1 in ranges:
            if r1 <= line_start or r0 >= line_end:
                continue
            overlaps.append((max(r0, line_start), min(r1, line_end)))

        if not overlaps:
            return f"{line_no:04d} | {self._escape_pdf_text(line_text)}"

        overlaps.sort()
        merged: list[tuple[int, int]] = []
        for s, e in overlaps:
            if not merged or s > merged[-1][1]:
                merged.append((s, e))
            else:
                merged[-1] = (merged[-1][0], max(merged[-1][1], e))

        parts = [f"{line_no:04d} | "]
        cursor = line_start
        for s, e in merged:
            if cursor < s:
                parts.append(self._escape_pdf_text(line_text[cursor - line_start:s - line_start]))
            parts.append(f"<font backColor=\"#FFF9A6\">{self._escape_pdf_text(line_text[s - line_start:e - line_start])}</font>")
            cursor = e
        if cursor < line_end:
            parts.append(self._escape_pdf_text(line_text[cursor - line_start:line_end - line_start]))
        return "".join(parts)

    def _build_highlighted_code_blocks(self, sid: str):
        """
        Returns per-file rows rendered with exact character-level highlights,
        matching the Tk text selection ranges used by graders.
        """
        blocks: list[tuple[str, list[str]]] = []
        files = get_student_files(self.sub_con, sid)

        rows = fetch_code_comments_for_student(self.grade_con, sid)
        file_ranges: dict[str, list[tuple[str, str]]] = {}
        for fp, sidx, eidx, _txt, _color, _ts in rows:
            file_ranges.setdefault(fp, []).append((sidx, eidx))

        for fp in files:
            code = get_file_content(self.sub_con, fp)
            if code is None:
                try:
                    code = Path(fp).read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    code = ""
            lines = code.splitlines(keepends=True)

            line_starts: list[int] = []
            pos = 0
            for line in lines:
                line_starts.append(pos)
                pos += len(line)
            code_len = len(code)

            normalized_ranges: list[tuple[int, int]] = []
            for sidx, eidx in file_ranges.get(fp, []):
                start_off = self._tk_index_to_offset(line_starts, code_len, sidx)
                end_off = self._tk_index_to_offset(line_starts, code_len, eidx)
                if end_off < start_off:
                    start_off, end_off = end_off, start_off

                while start_off < end_off and code[start_off].isspace():
                    start_off += 1
                while end_off > start_off and code[end_off - 1].isspace():
                    end_off -= 1

                if start_off == end_off:
                    continue
                normalized_ranges.append((start_off, end_off))

            rendered_lines: list[str] = []
            for i, raw_line in enumerate(lines, start=1):
                clean_line = raw_line.rstrip("\r\n")
                line_start = line_starts[i - 1]
                line_end = line_start + len(clean_line)
                rendered_lines.append(self._render_line_with_highlights(clean_line, line_start, line_end, normalized_ranges, i))
            blocks.append((Path(fp).name, rendered_lines))

        return blocks

    def _extract_code_snippet(self, file_path: str, start_index: str, end_index: str, max_chars: int = 420) -> str:
        code = get_file_content(self.sub_con, file_path)
        if code is None:
            try:
                code = Path(file_path).read_text(encoding="utf-8", errors="ignore")
            except Exception:
                code = ""

        lines = code.splitlines(keepends=True)
        if not lines:
            return ""

        def parse_idx(idx: str):
            try:
                a, b = str(idx).split(".", 1)
                return max(1, int(a)), max(0, int(b))
            except Exception:
                return 1, 0

        s_line, s_col = parse_idx(start_index)
        e_line, e_col = parse_idx(end_index)
        s_line = min(max(1, s_line), len(lines))
        e_line = min(max(1, e_line), len(lines))
        if (e_line, e_col) < (s_line, s_col):
            s_line, e_line = e_line, s_line
            s_col, e_col = e_col, s_col

        if s_line == e_line:
            snippet = lines[s_line - 1][s_col:e_col]
        else:
            parts = [lines[s_line - 1][s_col:]]
            for ln in range(s_line, e_line - 1):
                parts.append(lines[ln])
            parts.append(lines[e_line - 1][:e_col])
            snippet = "".join(parts)

        snippet = snippet.strip("\n\r ")
        if len(snippet) > max_chars:
            snippet = snippet[:max_chars].rstrip() + " …"
        return snippet

    def _format_code_block_for_pdf(self, code_text: str, max_width: int = 92, max_lines: int = 12) -> str:
        code_text = (code_text or "").replace("	", "    ")
        wrapped: list[str] = []
        for raw in code_text.splitlines() or [""]:
            parts = textwrap.wrap(raw, width=max_width, break_long_words=True, break_on_hyphens=False)
            wrapped.extend(parts or [""])
        if len(wrapped) > max_lines:
            wrapped = wrapped[:max_lines] + ["… [truncated]"]
        return "\n".join(wrapped)

    def export_student_pdf(self, sid: str, out_path: Path):
        if SimpleDocTemplate is None:
            raise RuntimeError("reportlab not installed. Install: pip install reportlab")

        srow = self.sub_con.execute(
            "SELECT student_name, COALESCE(lab_id,'') FROM students WHERE student_id=?",
            (sid,)
        ).fetchone()
        sname = srow[0] if srow else ""
        lab = srow[1] if srow else ""

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(
            str(out_path),
            pagesize=letter,
            title=f"{sid} grading report",
            leftMargin=28,
            rightMargin=28,
            topMargin=28,
            bottomMargin=28,
        )

        story = []
        story.append(Paragraph("<b>Grading Report</b>", styles["Title"]))
        story.append(Paragraph(f"<b>Student:</b> {sid} — {sname}", styles["Normal"]))
        if lab:
            story.append(Paragraph(f"<b>LabID:</b> {lab}", styles["Normal"]))
        overall = compute_overall_total(self.grade_con, sid)
        story.append(Paragraph(f"<b>Overall total:</b> {overall:g}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Per-question grading details (grouped by display/sub question id)
        q_display_map = build_question_display_map(self.grade_con)
        display_to_qids: dict[str, list[str]] = {}
        for qid in fetch_all_question_ids(self.grade_con):
            did = q_display_map.get(qid, qid)
            display_to_qids.setdefault(did, []).append(qid)

        for qid in fetch_display_question_ids(self.grade_con):
            members = display_to_qids.get(qid, [qid])
            qtitle = self.question_map.get(members[0], self.question_map.get(qid, qid))
            total = compute_total_by_display_id(self.grade_con, sid, qid)
            all_cols = []
            for mqid in members:
                all_cols.extend([(mqid, *c) for c in fetch_columns_for_question(self.grade_con, mqid)])
            if total <= 0 and not all_cols:
                continue
            story.append(Paragraph(f"<b>{qid}</b> — {qtitle} (Total: {total:g})", styles["Heading2"]))

            if self.student_pdf_options["include_grade_tables"]:

                cell_style = styles["BodyText"].clone("rubric_cell_style")
                cell_style.fontSize = 8
                cell_style.leading = 10

                table_data = [["Question", "Group", "Criterion", "Max", "Pts", "Note"]]
                for mqid, col_key, group, text, mx in all_cols:
                    score_map, note_map = load_student_scores(self.grade_con, sid, mqid)
                    pts = score_map.get(col_key, 0.0) or 0.0
                    note = (note_map.get(col_key, "") or "")
                    esc_group = (group or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    esc_text = (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    esc_note = note[:240].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    table_data.append([
                        mqid,
                        Paragraph(esc_group, cell_style),
                        Paragraph(esc_text, cell_style),
                        f"{mx:g}",
                        f"{pts:g}",
                        Paragraph(esc_note, cell_style),
                    ])

                tbl = Table(table_data, colWidths=[54, 70, 152, 36, 36, 204], repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EFE5FF")),
                    ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#3B2D5C")),
                    ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#B7B7B7")),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE", (0,0), (-1,-1), 8),
                    ("LEFTPADDING", (0,0), (-1,-1), 4),
                    ("RIGHTPADDING", (0,0), (-1,-1), 4),
                    ("TOPPADDING", (0,0), (-1,-1), 3),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 10))

            note_row = load_student_note(self.grade_con, sid, members[0])
            rationale = note_row[0] if note_row and note_row[0] else ""
            if self.student_pdf_options["include_rationale"] and rationale:
                story.append(Paragraph("<b>Rationale</b>", styles["Heading3"]))
                story.append(Paragraph(rationale.replace("\n", "<br/>"), styles["Normal"]))
                story.append(Spacer(1, 10))

        if self.student_pdf_options["include_comment_highlights"]:
            story.append(PageBreak())
            story.append(Paragraph("<b>Highlighted Code Review</b>", styles["Heading2"]))
            rows = fetch_code_comments_for_student(self.grade_con, sid)
            if not rows:
                story.append(Paragraph("(No code comments.)", styles["Normal"]))
            else:
                cell_style = styles["BodyText"].clone("comment_cell_style")
                cell_style.fontSize = 8
                cell_style.leading = 10
                code_cell_style = styles["Code"].clone("comment_code_cell_style")
                code_cell_style.fontSize = 7
                code_cell_style.leading = 8

                td = [["File", "Range", "Code (highlighted part)", "Comment"]]
                for fp, sidx, eidx, txt, _color, _ts in rows:
                    snippet = self._extract_code_snippet(fp, sidx, eidx)
                    esc_file = Path(fp).name.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    code_block = self._format_code_block_for_pdf(snippet or "(empty selection)", max_width=60, max_lines=10)
                    esc_comment = (txt or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    td.append([
                        Paragraph(esc_file, cell_style),
                        Paragraph(_format_comment_range_label(sidx, eidx), cell_style),
                        Preformatted(code_block, code_cell_style),
                        Paragraph(esc_comment[:280], cell_style),
                    ])

                tbl2 = Table(td, colWidths=[85, 90, 200, 181], repeatRows=1)
                tbl2.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EFE5FF")),
                    ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#FF4FA3")),
                    ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#B7B7B7")),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                    ("BACKGROUND", (2,1), (2,-1), colors.HexColor("#FFF9A6")),
                    ("BACKGROUND", (3,1), (3,-1), colors.HexColor("#FFFDF7")),
                    ("FONTSIZE", (0,0), (-1,-1), 8),
                    ("LEFTPADDING", (0,0), (-1,-1), 4),
                    ("RIGHTPADDING", (0,0), (-1,-1), 4),
                    ("TOPPADDING", (0,0), (-1,-1), 3),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ]))
                story.append(tbl2)

        if self.student_pdf_options["include_full_code_listing"]:
            story.append(PageBreak())
            story.append(Paragraph("<b>Full Code Listing</b>", styles["Heading2"]))
            blocks = self._build_highlighted_code_blocks(sid)
            line_style = styles["Code"].clone("code_line_style")
            line_style.fontSize = 7
            line_style.leading = 8
            for bi, (fname, rendered_lines) in enumerate(blocks):
                story.append(Paragraph(f"<b>{fname}</b>", styles["Heading3"]))
                if not rendered_lines:
                    story.append(Paragraph("(empty file)", styles["Normal"]))
                else:
                    compact_lines = [self._format_code_block_for_pdf(line, max_width=112, max_lines=2) for line in rendered_lines]
                    table_data = [[Paragraph(line or " ", line_style)] for line in compact_lines]
                    tbl_code = Table(table_data, colWidths=[516], repeatRows=0)
                    tbl_code.setStyle(TableStyle([
                        ("GRID", (0,0), (-1,-1), 0.2, colors.HexColor("#E0E0E0")),
                        ("LEFTPADDING", (0,0), (-1,-1), 4),
                        ("RIGHTPADDING", (0,0), (-1,-1), 4),
                        ("TOPPADDING", (0,0), (-1,-1), 1),
                        ("BOTTOMPADDING", (0,0), (-1,-1), 1),
                    ]))
                    story.append(tbl_code)
                if bi < len(blocks) - 1:
                    story.append(PageBreak())

        doc.build(story)

    def export_summary_pdf(self, out_path: Path, class_stats_text: str):
        if SimpleDocTemplate is None:
            raise RuntimeError("reportlab not installed. Install: pip install reportlab")

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(out_path), pagesize=letter, title="Summary")
        story = []
        story.append(Paragraph("<b>Grading Summary</b>", styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(class_stats_text, styles["Normal"]))
        story.append(Spacer(1, 12))

        students = self.sub_con.execute("""
          SELECT student_id, student_name, COALESCE(lab_id,'')
          FROM students
          WHERE included=1 OR LOWER(student_id)='full'
        """).fetchall()

        def key(r):
            sid = (r[0] or "")
            return (0 if sid.lower() == "full" else 1, sid)
        students = sorted(students, key=key)

        parts = fetch_rubric_parts(self.grade_con)
        q_display_map = build_question_display_map(self.grade_con)
        question_ids = fetch_display_question_ids(self.grade_con)
        headers = ["Student ID", "Name", "LabID"] + [f"{q_display_map.get(qid, qid)}:{ck}" for qid, ck, _g, _t, _m in parts] + [f"{qid}_total" for qid in question_ids] + ["Overall"]
        td = [headers]
        for sid, sname, lab in students:
            score_cache = {}
            row = [sid, sname, lab]
            for qid, ck, _g, _t, _m in parts:
                if qid not in score_cache:
                    score_cache[qid] = load_student_scores(self.grade_con, sid, qid)[0]
                val = score_cache[qid].get(ck)
                row.append("" if val is None else f"{val:g}")
            for qid in question_ids:
                row.append(f"{compute_total_by_display_id(self.grade_con, sid, qid):g}")
            overall = sum(compute_total_by_display_id(self.grade_con, sid, qid) for qid in question_ids)
            row.append(f"{overall:g}")
            td.append(row)

        tbl = Table(td, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EFE5FF")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(tbl)

        doc.build(story)

    def export_all_students_pdfs(self, out_dir: Path, report_tag: str = "Midterm", progress_cb=None):
        out_dir.mkdir(parents=True, exist_ok=True)
        students = self.sub_con.execute("""
          SELECT student_id, COALESCE(lab_id,'')
          FROM students
          WHERE LOWER(student_id) <> 'full' AND COALESCE(included,1)=1
          ORDER BY student_id
        """).fetchall()

        safe_tag = re.sub(r"[^A-Za-z0-9_-]+", "", (report_tag or "Midterm").strip()) or "Midterm"

        for i, (sid, lab) in enumerate(students, start=1):
            safe_lab = re.sub(r"[^A-Za-z0-9_-]+", "", (lab or "").strip()) or "LabX"
            out_name = f"{sid}_Report_{safe_lab}_{safe_tag}.pdf"
            out_path = out_dir / out_name
            try:
                self.export_student_pdf(sid, out_path)
            except Exception as e:
                # keep going, but record failures
                if progress_cb:
                    progress_cb(i, len(students), sid, False, str(e))
                continue
            if progress_cb:
                progress_cb(i, len(students), sid, True, "")



# =============================================================================
# 8) Auto Grader + GPT test helper (external modules)
# =============================================================================
from gpt_test import GPT_test
from auto_grader import AutoGrader

# =============================================================================
# 9) UI widgets
# =============================================================================

class ScrollableRubricGrid(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.hsb = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)

        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.hsb.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.inner.bind("<Configure>", lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda _e: self.canvas.itemconfigure(self.inner_id))

        self.columns = []
        self.score_vars = {}
        self.note_vars = {}
        self._change_callback = None

    def set_change_callback(self, callback):
        self._change_callback = callback

    def build(self, columns):
        for w in self.inner.winfo_children():
            w.destroy()

        self.columns = columns
        self.score_vars.clear()
        self.note_vars.clear()

        ttk.Label(self.inner, text="Criterion", style="Pastel.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Label(self.inner, text="Score", style="Pastel.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(self.inner, text="Note", style="Pastel.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky="w", padx=6, pady=4)

        r = 1
        current_group = None
        for col_key, group, text, maxp in columns:
            group = group or ""

            if group != current_group:
                current_group = group
                ttk.Separator(self.inner, orient="horizontal").grid(row=r, column=0, columnspan=3, sticky="ew", pady=(10, 6))
                r += 1
                if current_group.strip():
                    ttk.Label(self.inner, text=current_group, style="Pastel.TLabel", font=("Segoe UI", 10, "bold")).grid(row=r, column=0, columnspan=3, sticky="w", padx=6, pady=(0, 4))
                    r += 1

            header = f"{text}\n/{maxp:g}"
            ttk.Label(self.inner, text=header, style="Pastel.TLabel", justify="left", wraplength=560).grid(row=r, column=0, sticky="w", padx=6, pady=4)

            sv = tk.StringVar(value="")
            nv = tk.StringVar(value="")
            if self._change_callback is not None:
                sv.trace_add("write", lambda *_args: self._change_callback())
                nv.trace_add("write", lambda *_args: self._change_callback())
            ttk.Entry(self.inner, textvariable=sv, width=8).grid(row=r, column=1, sticky="w", padx=6, pady=4)
            ttk.Entry(self.inner, textvariable=nv, width=60).grid(row=r, column=2, sticky="ew", padx=6, pady=4)

            self.score_vars[col_key] = sv
            self.note_vars[col_key] = nv
            r += 1

        self.inner.columnconfigure(2, weight=1)

    def set_values(self, score_map, note_map):
        for k, sv in self.score_vars.items():
            v = score_map.get(k, None)
            sv.set("" if v is None else str(v))
        for k, nv in self.note_vars.items():
            nv.set(note_map.get(k, "") or "")

    def get_values(self):
        scores = {k: v.get().strip() for k, v in self.score_vars.items()}
        notes = {k: v.get().strip() for k, v in self.note_vars.items()}
        return scores, notes


# =============================================================================
# 10) Pastel theme
# =============================================================================

def pastel_style(root: tk.Tk):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    bg = "#F7F2FF"        # very light purple
    panel = "#FFFDE8"     # lemon-cream yellow
    accent = "#DCCBFF"    # light purple
    accent2 = "#FF4FA3"   # bright pink (used sparingly: active states)
    text = "#2A2440"
    select = "#EFE5FF"

    root.configure(bg=bg)

    style.configure(".", font=("Segoe UI", 10))
    style.configure("Pastel.TFrame", background=bg)
    style.configure("PastelCard.TFrame", background=panel, relief="flat")
    style.configure("Pastel.TLabel", background=bg, foreground=text)
    style.configure("PastelCard.TLabel", background=panel, foreground=text)

    style.configure("TButton", padding=6)
    style.map("TButton", background=[("active", accent2)], foreground=[("disabled", "#888")])

    style.configure("TNotebook", background=bg, borderwidth=0)
    style.configure("TNotebook.Tab", padding=(12, 6))
    style.map("TNotebook.Tab",
              background=[("selected", panel), ("!selected", bg)],
              foreground=[("selected", text), ("!selected", text)])

    style.configure("Treeview",
                    background=panel,
                    fieldbackground=panel,
                    foreground=text,
                    rowheight=24,
                    borderwidth=0)
    style.map("Treeview", background=[("selected", select)], foreground=[("selected", text)])

    style.configure("Treeview.Heading",
                    font=("Segoe UI", 10, "bold"),
                    background=accent,
                    foreground=text,
                    relief="flat")
    style.map("Treeview.Heading", background=[("active", accent2)])

    return {"bg": bg, "panel": panel, "accent": accent, "accent2": accent2, "text": text, "select": select}


# =============================================================================
# 11) Scan Window (folder-by-folder review + LabID + scan flexibility)
# =============================================================================

class ScanWindow(tk.Toplevel):
    def __init__(self, parent_app):
        super().__init__(parent_app.root)
        self.app = parent_app
        self.con = parent_app.sub_con

        self.title("Scan/Edit Submissions")
        self.geometry("1650x920")

        self.root_folder: Path | None = None
        # Always store file content so previews are consistently available after reload.
        self.store_file_content = tk.BooleanVar(value=True)

        # Flex settings
        self.file_globs_var = tk.StringVar(value="*.java")
        self.filename_regex_var = tk.StringVar(value="")  # optional
        self.exclude_filename_regex_var = tk.StringVar(value="")
        self.filename_regex_ignore_case_var = tk.BooleanVar(value=True)
        self.folder_id_regex_var = tk.StringVar(value="")
        self.folder_name_regex_var = tk.StringVar(value="")
        self.only_new_files_var = tk.BooleanVar(value=False)
        self.global_lab_id_var = tk.StringVar(value="")
        self.find_var = tk.StringVar(value="")
        self.find_regex_var = tk.BooleanVar(value=False)
        self.find_case_sensitive_var = tk.BooleanVar(value=False)
        self._find_from = "1.0"

        self.rows: dict[str, dict] = {}
        self.folder_order: list[str] = []

        # Skimming mode (quick review through files/students)
        self.skim_running = False
        self.skim_delay_ms_var = tk.IntVar(value=300)
        self._skim_folder_idx = 0
        self._skim_file_idx = 0
        self._skimmable_folder_keys: list[str] = []
        self.selected_folder_key: str | None = None

        self._last_folder_selection: tuple[str, ...] = ()
        self._last_scan_file_selection: tuple[str, ...] = ()

        self._build()
        self.apply_profile_settings(self.app.get_active_regex_payload())
        self.load_existing_rows_from_db()
        self.after(120, self._poll_scan_selections)

    def _build(self):
        top = ttk.Frame(self, padding=10, style="Pastel.TFrame")
        top.pack(fill=tk.X)

        actions = ttk.Frame(top, style="Pastel.TFrame")
        actions.pack(fill=tk.X)

        ttk.Button(actions, text="Choose ROOT Folder", command=self.choose_root).pack(side=tk.LEFT)
        ttk.Button(actions, text="Scan / Rescan All", command=self.scan).pack(side=tk.LEFT, padx=8)
        ttk.Button(actions, text="Rescan Selected Folder", command=self.rescan_selected_folder).pack(side=tk.LEFT)

        ttk.Button(actions, text="Commit", command=self.commit_current_scan).pack(side=tk.RIGHT)
        ttk.Button(actions, text="Save Scan to DB", command=self.save_to_db).pack(side=tk.RIGHT, padx=(0, 8))

        self.status_lbl = ttk.Label(actions, text="No folder selected.", style="Pastel.TLabel")
        self.status_lbl.pack(side=tk.RIGHT, padx=10)

        opts_nb = ttk.Notebook(top)
        opts_nb.pack(fill=tk.X, pady=(8, 0))

        filters_tab = ttk.Frame(opts_nb, style="Pastel.TFrame", padding=8)
        skim_tab = ttk.Frame(opts_nb, style="Pastel.TFrame", padding=8)
        opts_nb.add(filters_tab, text="Scan Filters")
        opts_nb.add(skim_tab, text="Skimming")

        ttk.Label(filters_tab, text="File globs (comma/space/newline)", style="Pastel.TLabel").pack(side=tk.LEFT)
        ttk.Entry(filters_tab, textvariable=self.file_globs_var, width=30).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(filters_tab, text="Filename include-regex", style="Pastel.TLabel").pack(side=tk.LEFT)
        ttk.Entry(filters_tab, textvariable=self.filename_regex_var, width=24).pack(side=tk.LEFT, padx=(6, 6))
        ttk.Label(filters_tab, text="Exclude-regex", style="Pastel.TLabel").pack(side=tk.LEFT)
        ttk.Entry(filters_tab, textvariable=self.exclude_filename_regex_var, width=20).pack(side=tk.LEFT, padx=(6, 6))
        ttk.Checkbutton(filters_tab, text="Ignore case", variable=self.filename_regex_ignore_case_var).pack(side=tk.LEFT)
        ttk.Checkbutton(filters_tab, text="Only show files not already registered", variable=self.only_new_files_var).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Label(filters_tab, text="Regex profile:", style="Pastel.TLabel").pack(side=tk.LEFT, padx=(12, 4))
        ttk.Label(filters_tab, textvariable=self.app.active_regex_profile_var, style="Pastel.TLabel").pack(side=tk.LEFT)

        ttk.Label(skim_tab, text="Skim delay (ms)", style="Pastel.TLabel").pack(side=tk.LEFT)
        ttk.Entry(skim_tab, textvariable=self.skim_delay_ms_var, width=7).pack(side=tk.LEFT, padx=(6, 10))
        ttk.Button(skim_tab, text="Start Skimming (from selected)", command=self.start_skimming).pack(side=tk.LEFT)
        ttk.Button(skim_tab, text="Skim Unassigned", command=self.start_skimming_unassigned).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(skim_tab, text="Stop Skimming", command=self.stop_skimming).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(skim_tab, text="Stop keys: Q / Esc", style="Pastel.TLabel").pack(side=tk.LEFT, padx=(12, 0))

        main = ttk.Frame(self, padding=10, style="Pastel.TFrame")
        main.pack(fill=tk.BOTH, expand=True)

        main.columnconfigure(0, weight=2)
        main.columnconfigure(1, weight=2)
        main.columnconfigure(2, weight=3)
        main.rowconfigure(0, weight=1)

        left = ttk.Frame(main, style="Pastel.TFrame")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(1, weight=3)
        left.rowconfigure(3, weight=2)
        ttk.Label(left, text="Full list (all scanned folders)", style="Pastel.TLabel").grid(row=0, column=0, sticky="w")

        cols = ("is_student", "folder", "det_id", "det_name", "final_id", "final_name", "lab_id", "nfiles")
        self.tree = ttk.Treeview(left, columns=cols, show="headings", selectmode="browse")
        for c, w in [("is_student", 80), ("folder", 330), ("det_id", 120), ("det_name", 160),
                     ("final_id", 160), ("final_name", 170), ("lab_id", 90), ("nfiles", 70)]:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="w")
        self.tree.grid(row=1, column=0, sticky="nsew")

        ttk.Label(left, text="Student list (valid ID + name)", style="Pastel.TLabel").grid(row=2, column=0, sticky="w", pady=(10, 0))
        student_cols = ("student_id", "student_name", "lab_id", "folder", "nfiles")
        self.student_tree = ttk.Treeview(left, columns=student_cols, show="headings", selectmode="none", height=8)
        for c, w in [("student_id", 130), ("student_name", 170), ("lab_id", 90), ("folder", 220), ("nfiles", 60)]:
            self.student_tree.heading(c, text=c)
            self.student_tree.column(c, width=w, anchor="w")
        self.student_tree.grid(row=3, column=0, sticky="nsew")

        mid = ttk.Frame(main, style="Pastel.TFrame")
        mid.grid(row=0, column=1, sticky="nsew", padx=(0, 10))
        mid.columnconfigure(0, weight=1)

        ttk.Label(mid, text="Selected Folder Edit", style="Pastel.TLabel").grid(row=0, column=0, sticky="w")
        self.sel_folder_var = tk.StringVar()
        ttk.Entry(mid, textvariable=self.sel_folder_var, state="readonly").grid(row=1, column=0, sticky="ew", pady=(2, 8))

        self.include_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(mid, text="IsStudent (include in grading)", variable=self.include_var,
                        command=self.apply_include_toggle).grid(row=2, column=0, sticky="w")

        ttk.Label(mid, text="Final Student ID", style="Pastel.TLabel").grid(row=3, column=0, sticky="w")
        self.final_id_var = tk.StringVar()
        ttk.Entry(mid, textvariable=self.final_id_var).grid(row=4, column=0, sticky="ew")

        ttk.Label(mid, text="Final Student Name", style="Pastel.TLabel").grid(row=5, column=0, sticky="w", pady=(8, 0))
        self.final_name_var = tk.StringVar()
        ttk.Entry(mid, textvariable=self.final_name_var).grid(row=6, column=0, sticky="ew")

        ttk.Label(mid, text="LabID (section/group)", style="Pastel.TLabel").grid(row=7, column=0, sticky="w", pady=(8, 0))
        self.lab_id_var = tk.StringVar()
        ttk.Entry(mid, textvariable=self.lab_id_var).grid(row=8, column=0, sticky="ew")

        ttk.Label(mid, text="Apply LabID to IsStudent folders", style="Pastel.TLabel").grid(row=9, column=0, sticky="w", pady=(10, 0))
        ttk.Entry(mid, textvariable=self.global_lab_id_var).grid(row=10, column=0, sticky="ew")
        ttk.Button(mid, text="Apply LabID to All IsStudent", command=self.apply_global_lab_id).grid(row=11, column=0, sticky="ew", pady=(4, 0))

        ttk.Label(mid, text="Selection shortcuts", style="Pastel.TLabel").grid(row=12, column=0, sticky="w", pady=(10, 0))
        quick = ttk.Frame(mid, style="Pastel.TFrame")
        quick.grid(row=13, column=0, sticky="ew")
        ttk.Button(quick, text="Use selected text as ID (I)", command=self.use_selection_as_id).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(quick, text="Use selected text as Name (N)", command=self.use_selection_as_name).pack(side=tk.LEFT)

        self._suspend_auto_apply = False
        self._bind_auto_apply()

        right = ttk.Frame(main, style="Pastel.TFrame")
        right.grid(row=0, column=2, sticky="nsew")
        right.rowconfigure(3, weight=1)
        right.columnconfigure(0, weight=1)

        ttk.Label(right, text="Files in Selected Folder", style="Pastel.TLabel").grid(row=0, column=0, sticky="w")
        self.files_tree = ttk.Treeview(right, columns=("file",), show="headings", height=10, selectmode="browse")
        self.files_tree.heading("file", text="file path")
        self.files_tree.column("file", width=520, anchor="w")
        self.files_tree.grid(row=1, column=0, sticky="nsew")
        self.files_tree.bind("<<TreeviewSelect>>", self.on_scan_file_select)

        preview_frame = ttk.Frame(right, style="Pastel.TFrame")
        preview_frame.grid(row=3, column=0, sticky="nsew")
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        self.preview = tk.Text(preview_frame, wrap="none")
        self.preview_font_normal = tkfont.Font(font=self.preview["font"])
        self.preview_font_bold = tkfont.Font(font=self.preview["font"])
        self.preview_font_bold.configure(weight="bold")
        self.preview.configure(font=self.preview_font_normal)
        self.preview.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview.yview)
        sb.grid(row=0, column=1, sticky="ns")
        self.preview.configure(yscrollcommand=sb.set)
        self.preview.tag_configure("find_hit", background="#ffe082")

        find_row = ttk.Frame(right, style="Pastel.TFrame")
        find_row.grid(row=4, column=0, sticky="ew", pady=(8, 0))
        ttk.Label(find_row, text="Find in file", style="Pastel.TLabel").pack(side=tk.LEFT)
        self.find_entry = ttk.Entry(find_row, textvariable=self.find_var, width=26)
        self.find_entry.pack(side=tk.LEFT, padx=(6, 6))
        ttk.Checkbutton(find_row, text="Regex", variable=self.find_regex_var).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Checkbutton(find_row, text="Case sensitive", variable=self.find_case_sensitive_var).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(find_row, text="Find Next", command=self.find_next).pack(side=tk.LEFT)

        self.preview.bind("<KeyPress-i>", self._hotkey_use_id)
        self.preview.bind("<KeyPress-I>", self._hotkey_use_id)
        self.preview.bind("<KeyPress-n>", self._hotkey_use_name)
        self.preview.bind("<KeyPress-N>", self._hotkey_use_name)
        self.preview.bind("<Control-f>", lambda _e: self.focus_find_entry())
        self.bind("<KeyPress-q>", lambda _e: self.stop_skimming())
        self.bind("<Escape>", lambda _e: self.stop_skimming())

    def _poll_scan_selections(self):
        folder_sel = tuple(self.tree.selection())
        if folder_sel != self._last_folder_selection:
            self._last_folder_selection = folder_sel
            if folder_sel:
                self.on_folder_select()

        file_sel = tuple(self.files_tree.selection())
        if file_sel != self._last_scan_file_selection:
            self._last_scan_file_selection = file_sel
            if file_sel:
                self.on_scan_file_select()

        self.after(120, self._poll_scan_selections)

    def choose_root(self):
        folder = filedialog.askdirectory(title="Select ROOT submissions folder")
        if not folder:
            return
        self.root_folder = Path(folder)
        self._set_scan_status(prefix=f"Root: {self.root_folder}")
        self.rows.clear()
        self.folder_order.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        self.preview.configure(font=self.preview_font_normal)
        self.preview.delete("1.0", tk.END)

    def _parse_globs(self) -> list[str]:
        raw = (self.file_globs_var.get() or "").strip()
        if not raw:
            return ["*.java"]
        parts = [p.strip() for p in re.split(r"[\s,;]+", raw) if p.strip()]
        return parts if parts else ["*.java"]

    def _scan_counts(self) -> dict:
        total_folders = len(self.folder_order)
        total_files = 0
        blank_folders = 0
        total_students = 0
        included_rows = 0
        included_students = 0

        for folder_key in self.folder_order:
            r = self.rows.get(folder_key, {})
            files = r.get("files") or []
            nfiles = len(files)
            total_files += nfiles
            if nfiles == 0:
                blank_folders += 1

            is_student = has_required_student_fields(r.get("final_id", ""), r.get("final_name", ""))
            if is_student:
                total_students += 1

            if r.get("include"):
                included_rows += 1
                if is_student:
                    included_students += 1

        return {
            "total_computers": total_folders,
            "folders": total_folders,
            "files": total_files,
            "blank_folders": blank_folders,
            "total_students": total_students,
            "included_rows": included_rows,
            "included_students": included_students,
        }

    def _set_scan_status(self, prefix: str = ""):
        c = self._scan_counts()
        parts = [
            f"Total Computers: {c['total_computers']}",
            f"Files: {c['files']}",
            f"Total students: {c['total_students']}",
            f"IsStudent rows: {c['included_rows']}",
            f"IsStudent students: {c['included_students']}",
            f"Blank folders: {c['blank_folders']}",
        ]
        msg = " | ".join(parts)
        if prefix:
            msg = f"{prefix} | {msg}"
        self.status_lbl.config(text=msg)

    def reset_regex_defaults(self):
        self.file_globs_var.set("*.java")
        self.filename_regex_var.set("")
        self.exclude_filename_regex_var.set("")
        self.filename_regex_ignore_case_var.set(True)
        self.folder_id_regex_var.set("")
        self.folder_name_regex_var.set("")


    def _bind_auto_apply(self):
        self.final_id_var.trace_add("write", self._on_manual_edit)
        self.final_name_var.trace_add("write", self._on_manual_edit)
        self.lab_id_var.trace_add("write", self._on_manual_edit)

    def _on_manual_edit(self, *_args):
        if self._suspend_auto_apply:
            return
        self.apply_edits()

    @staticmethod
    def _normalized_name(name: str, folder_label: str, detected_name: str = "") -> str:
        """
        Keep blank names blank.
        Do not auto-store folder names as student names when nothing was detected.
        """
        cleaned = re.sub(r"\s+", " ", (name or "").strip())
        if not cleaned:
            return ""

        folder_base = (Path(folder_label).name if folder_label else "").strip()
        detected = re.sub(r"\s+", " ", (detected_name or "").strip())
        if folder_base and cleaned == folder_base and not detected:
            return ""
        return cleaned

    def _scan_settings_payload(self) -> dict:
        return {
            "file_globs": (self.file_globs_var.get() or "").strip() or "*.java",
            "filename_regex": (self.filename_regex_var.get() or "").strip(),
            "exclude_filename_regex": (self.exclude_filename_regex_var.get() or "").strip(),
            "filename_regex_ignore_case": bool(self.filename_regex_ignore_case_var.get()),
            "folder_id_regex": (self.folder_id_regex_var.get() or "").strip(),
            "folder_name_regex": (self.folder_name_regex_var.get() or "").strip(),
            "only_new_files": bool(self.only_new_files_var.get()),
        }

    def apply_profile_settings(self, payload: dict):
        self.file_globs_var.set((payload.get("file_globs") or "*.java").strip() or "*.java")
        self.filename_regex_var.set((payload.get("filename_regex") or "").strip())
        self.exclude_filename_regex_var.set((payload.get("exclude_filename_regex") or "").strip())
        self.filename_regex_ignore_case_var.set(bool(payload.get("filename_regex_ignore_case", True)))
        self.folder_id_regex_var.set((payload.get("folder_id_regex") or "").strip())
        self.folder_name_regex_var.set((payload.get("folder_name_regex") or "").strip())
        self.only_new_files_var.set(bool(payload.get("only_new_files", False)))

    def commit_current_scan(self):
        self.save_to_db(show_message=False)
        payload_rows = [self.rows[k] for k in self.folder_order if k in self.rows]
        self.app.commit_scan_session_from_window(
            root_folder=str(self.root_folder or ""),
            global_lab_id=(self.global_lab_id_var.get() or "").strip(),
            session_rows=payload_rows,
        )
        messagebox.showinfo("Committed", "Current scan + profile committed to DB.")

    def save_regex_copy(self):
        path = filedialog.asksaveasfilename(
            title="Save regex copy",
            defaultextension=".json",
            filetypes=[("JSON", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        payload = self._scan_settings_payload()
        try:
            Path(path).write_text(json.dumps(payload, indent=2), encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Save failed", str(e))
            return
        messagebox.showinfo("Saved", f"Regex copy saved:\n{path}")

    def load_regex_copy(self):
        path = filedialog.askopenfilename(
            title="Load regex copy",
            filetypes=[("JSON", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as e:
            messagebox.showerror("Load failed", str(e))
            return

        self.file_globs_var.set((payload.get("file_globs") or "*.java").strip() or "*.java")
        self.filename_regex_var.set((payload.get("filename_regex") or "").strip())
        self.exclude_filename_regex_var.set((payload.get("exclude_filename_regex") or "").strip())
        self.filename_regex_ignore_case_var.set(bool(payload.get("filename_regex_ignore_case", True)))
        self.folder_id_regex_var.set((payload.get("folder_id_regex") or "").strip())
        self.folder_name_regex_var.set((payload.get("folder_name_regex") or "").strip())
        self.only_new_files_var.set(bool(payload.get("only_new_files", False)))
        messagebox.showinfo("Loaded", f"Regex copy loaded:\n{path}")

    def load_existing_rows_from_db(self):
        """
        Populate Scan/Edit with current submissions DB rows.
        Includes both saved students and any file groups that are not currently
        attached to a valid student record so they can be edited/included later.
        """
        self.rows.clear()
        self.folder_order.clear()

        students = self.con.execute("""
            SELECT student_id, student_name, COALESCE(lab_id,''), COALESCE(folder_path,'')
            FROM students
            ORDER BY student_id
        """).fetchall()
        student_map = {sid: {"name": sname or "", "lab": lab or "", "folder_path": folder_path or ""}
                       for sid, sname, lab, folder_path in students}

        used_keys = set()
        known_students = {sid for sid, *_rest in students}

        def _unique_key(base_key: str) -> str:
            folder_key = base_key
            n = 2
            while folder_key in used_keys:
                folder_key = f"{base_key}#{n}"
                n += 1
            used_keys.add(folder_key)
            return folder_key

        folder_groups: dict[str, dict] = {}

        for sid, sname, lab, folder_path in students:
            base_key = (folder_path or "").strip() or f"DB:{sid}"
            grp = folder_groups.setdefault(base_key, {
                "files": [],
                "det_id": "",
                "det_name": "",
                "student_id": "",
                "lab_id": "",
            })
            grp["student_id"] = sid or grp.get("student_id", "")
            grp["det_id"] = sid or grp.get("det_id", "")
            grp["det_name"] = sname or grp.get("det_name", "")
            grp["lab_id"] = lab or grp.get("lab_id", "")

        file_rows = self.con.execute("""
            SELECT
              file_path,
              COALESCE(student_id, ''),
              COALESCE(source_folder, ''),
              COALESCE(detected_id, ''),
              COALESCE(detected_name, '')
            FROM files
            ORDER BY source_folder, file_path
        """).fetchall()

        for fp, sid, source_folder, det_id, det_name in file_rows:
            sid = (sid or "").strip()
            folder_path = (source_folder or "").strip() or str(Path(fp).parent)
            if not folder_path:
                folder_path = f"UNASSIGNED:{fp}"

            grp = folder_groups.setdefault(folder_path, {
                "files": [],
                "det_id": "",
                "det_name": "",
                "student_id": "",
                "lab_id": "",
            })
            grp["files"].append(fp)
            if sid and not grp["student_id"]:
                grp["student_id"] = sid
            if sid and not grp["lab_id"]:
                grp["lab_id"] = (student_map.get(sid) or {}).get("lab", "")
            if det_id and not grp["det_id"]:
                grp["det_id"] = det_id
            if det_name and not grp["det_name"]:
                grp["det_name"] = det_name

        represented_students = set()

        for folder_path in sorted(folder_groups.keys()):
            grp = folder_groups[folder_path]
            files = grp.get("files") or []
            sid = (grp.get("student_id") or "").strip()
            student_row = student_map.get(sid)
            if student_row:
                represented_students.add(sid)

            det_id = (grp.get("det_id") or "").strip()
            det_name = (grp.get("det_name") or "").strip()
            final_id = sid if student_row else det_id
            final_name = self._normalized_name(
                student_row["name"] if student_row else det_name,
                folder_path,
                det_name,
            )
            lab = student_row["lab"] if student_row else (grp.get("lab_id") or "")

            include_row = bool(files) and has_required_student_fields(final_id, final_name)
            folder_key = _unique_key(folder_path)
            self.folder_order.append(folder_key)
            self.rows[folder_key] = {
                "include": include_row,
                "manual_include_override": None,
                "folder": folder_path,
                "det_id": final_id or det_id,
                "det_name": final_name or det_name,
                "final_id": final_id,
                "final_name": final_name,
                "lab_id": lab,
                "files": files,
            }

        # Keep students without files editable too (legacy or pre-setup rows).
        for sid, sname, lab, folder_path in students:
            if sid in represented_students:
                continue
            base_key = (folder_path or "").strip() or f"DB:{sid}"
            folder_key = _unique_key(base_key)
            self.folder_order.append(folder_key)
            self.rows[folder_key] = {
                "include": False,
                "manual_include_override": None,
                "folder": base_key,
                "det_id": sid or "",
                "det_name": sname or "",
                "final_id": sid or "",
                "final_name": self._normalized_name(sname or "", base_key, sname or ""),
                "lab_id": lab or "",
                "files": [],
            }

        self._reload_tree_rows()

        self._set_scan_status(prefix="Loaded from DB")

    def scan(self):
        if not self.root_folder:
            messagebox.showinfo("Pick folder", "Choose ROOT folder first.")
            return

        existing_files = set()
        if self.only_new_files_var.get():
            existing_files = {r[0] for r in self.con.execute("SELECT file_path FROM files").fetchall()}

        file_globs = self._parse_globs()
        filename_regex = (self.filename_regex_var.get() or "").strip()
        exclude_filename_regex = (self.exclude_filename_regex_var.get() or "").strip()
        filename_regex_ignore_case = bool(self.filename_regex_ignore_case_var.get())
        folder_id_regex = (self.folder_id_regex_var.get() or "").strip()
        folder_name_regex = (self.folder_name_regex_var.get() or "").strip()

        scanner = DefaultFolderScanner(
            self.root_folder,
            file_globs=file_globs,
            include_filename_regex=filename_regex,
            exclude_filename_regex=exclude_filename_regex,
            filename_regex_ignore_case=filename_regex_ignore_case,
            folder_id_regex=folder_id_regex,
            folder_name_regex=folder_name_regex,
        )
        folders = scanner.collect_folders()

        existing_rows = {
            k: {
                "include": bool(v.get("include")),
                "manual_include_override": v.get("manual_include_override"),
                "folder": v.get("folder", k),
                "det_id": v.get("det_id", ""),
                "det_name": v.get("det_name", ""),
                "final_id": v.get("final_id", ""),
                "final_name": v.get("final_name", ""),
                "lab_id": v.get("lab_id", ""),
                "files": list(v.get("files") or []),
            }
            for k, v in self.rows.items()
        }
        existing_order = list(self.folder_order)

        self.rows.clear()
        self.folder_order.clear()

        for sub in folders:
            final_id, final_name, det_id, det_name, files = scanner.detect_folder(sub)
            if existing_files:
                files = [fp for fp in files if fp not in existing_files]
            folder_key = str(sub)
            self.folder_order.append(folder_key)
            is_student = has_required_student_fields(final_id or "", final_name or "")
            include_row = bool(files) and is_student
            self.rows[folder_key] = {
                "include": include_row,
                "manual_include_override": None,
                "folder": folder_key,
                "det_id": det_id or "",
                "det_name": det_name or "",
                "final_id": final_id or "",
                "final_name": self._normalized_name(final_name or "", str(sub), det_name or ""),
                "lab_id": "",
                "files": files,
            }

            prior = existing_rows.get(folder_key)
            if prior:
                self.rows[folder_key]["final_id"] = prior.get("final_id", self.rows[folder_key]["final_id"])
                self.rows[folder_key]["final_name"] = self._normalized_name(
                    prior.get("final_name", self.rows[folder_key]["final_name"]),
                    folder_key,
                    self.rows[folder_key].get("det_name", ""),
                )
                self.rows[folder_key]["lab_id"] = prior.get("lab_id", "")
                self.rows[folder_key]["manual_include_override"] = prior.get("manual_include_override")

                is_student = has_required_student_fields(
                    self.rows[folder_key]["final_id"],
                    self.rows[folder_key]["final_name"],
                )
                has_files = bool(self.rows[folder_key].get("files"))
                manual = self.rows[folder_key]["manual_include_override"]
                if not (is_student and has_files):
                    self.rows[folder_key]["include"] = False
                elif manual is False:
                    self.rows[folder_key]["include"] = False
                elif manual is True:
                    self.rows[folder_key]["include"] = True
                else:
                    self.rows[folder_key]["include"] = True

        for folder_key in existing_order:
            if folder_key in self.rows:
                continue
            self.folder_order.append(folder_key)
            self.rows[folder_key] = existing_rows[folder_key]

        self._reload_tree_rows()

        self._set_scan_status(prefix="Scan done")

    def rescan_selected_folder(self):
        folder_key = self._selected_folder_from_tree()
        if not folder_key:
            messagebox.showinfo("Rescan", "Select a folder row first.")
            return
        row = self.rows.get(folder_key)
        if not row:
            return

        file_globs = self._parse_globs()
        filename_regex = (self.filename_regex_var.get() or "").strip()
        exclude_filename_regex = (self.exclude_filename_regex_var.get() or "").strip()
        filename_regex_ignore_case = bool(self.filename_regex_ignore_case_var.get())
        folder_id_regex = (self.folder_id_regex_var.get() or "").strip()
        folder_name_regex = (self.folder_name_regex_var.get() or "").strip()

        scanner = DefaultFolderScanner(
            Path(folder_key),
            file_globs=file_globs,
            include_filename_regex=filename_regex,
            exclude_filename_regex=exclude_filename_regex,
            filename_regex_ignore_case=filename_regex_ignore_case,
            folder_id_regex=folder_id_regex,
            folder_name_regex=folder_name_regex,
        )
        final_id, final_name, det_id, det_name, files = scanner.detect_folder(Path(folder_key))
        if self.only_new_files_var.get():
            existing_files = {r[0] for r in self.con.execute("SELECT file_path FROM files").fetchall()}
            files = [fp for fp in files if fp not in existing_files]

        row["files"] = files
        row["det_id"] = det_id or ""
        row["det_name"] = det_name or ""
        if not (row.get("final_id") or "").strip():
            row["final_id"] = final_id or ""
        if not (row.get("final_name") or "").strip():
            row["final_name"] = self._normalized_name(final_name or "", folder_key, det_name or "")

        is_student = has_required_student_fields(row.get("final_id", ""), row.get("final_name", ""))
        manual = row.get("manual_include_override")
        if not (is_student and files):
            row["include"] = False
        elif manual is False:
            row["include"] = False
        else:
            row["include"] = True

        self._refresh_tree_row(folder_key)
        self.on_folder_select()
        self._reload_student_rows()
        self._set_scan_status(prefix="Rescanned selected folder")

    def on_folder_select(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        folder_key = sel[0]
        self.selected_folder_key = folder_key
        r = self.rows.get(folder_key)
        if not r:
            return

        self.sel_folder_var.set(folder_key)
        self.include_var.set(bool(r["include"]))
        self._suspend_auto_apply = True
        self.final_id_var.set(r["final_id"])
        self.final_name_var.set(r["final_name"])
        self.lab_id_var.set(r.get("lab_id",""))
        self._suspend_auto_apply = False

        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        for idx, fp in enumerate(r["files"]):
            self.files_tree.insert("", "end", iid=f"file-{idx}", values=(fp,))

        self.preview.configure(font=self.preview_font_normal if bool(r.get("include")) else self.preview_font_bold)
        self.preview.delete("1.0", tk.END)

    def on_scan_file_select(self, _evt=None, file_iid: str | None = None):
        if file_iid is not None:
            sel = (file_iid,)
        else:
            sel = self.files_tree.selection()
        if not sel:
            return
        fp = self.files_tree.item(sel[0], "values")[0]
        try:
            content = Path(fp).read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            content = f"Error reading file:\n{e}"
        self.preview.delete("1.0", tk.END)
        self.preview.insert("1.0", content)
        self.preview.tag_remove("find_hit", "1.0", tk.END)
        self._find_from = "1.0"

    def apply_include_toggle(self):
        folder_key = self.selected_folder_key or self.sel_folder_var.get().strip()
        if not folder_key or folder_key not in self.rows:
            return
        row = self.rows[folder_key]
        files = row.get("files") or []
        if not files:
            row["include"] = False
            row["manual_include_override"] = False
            self.include_var.set(False)
            self._refresh_tree_row(folder_key)
            self._reload_student_rows()
            self._set_scan_status(prefix="Scan info")
            return

        row["include"] = bool(self.include_var.get())
        row["manual_include_override"] = row["include"]
        self._refresh_tree_row(folder_key)
        self.preview.configure(font=self.preview_font_normal if bool(self.rows[folder_key].get("include")) else self.preview_font_bold)
        self._reload_student_rows()
        self._set_scan_status(prefix="Scan info")

    def apply_edits(self):
        folder_key = self.selected_folder_key or self.sel_folder_var.get().strip()
        if not folder_key or folder_key not in self.rows:
            return

        raw_fid = self.final_id_var.get().strip()
        fid = "FULL" if raw_fid.lower() == "full" else extract_numeric_id(raw_fid)
        fname = self._normalized_name(self.final_name_var.get(), folder_key, self.rows[folder_key].get("det_name", ""))
        lab = self.lab_id_var.get().strip()

        self.rows[folder_key]["final_id"] = fid
        self.rows[folder_key]["final_name"] = fname
        self.rows[folder_key]["lab_id"] = lab

        has_files = bool(self.rows[folder_key].get("files"))
        manual_override = self.rows[folder_key].get("manual_include_override")

        if not has_files:
            self.rows[folder_key]["include"] = False
        elif manual_override is False:
            self.rows[folder_key]["include"] = False
        else:
            self.rows[folder_key]["include"] = True

        self._suspend_auto_apply = True
        self.include_var.set(bool(self.rows[folder_key]["include"]))
        self.final_name_var.set(fname)
        self._suspend_auto_apply = False

        self._refresh_tree_row(folder_key)
        self.preview.configure(font=self.preview_font_normal if bool(self.rows[folder_key].get("include")) else self.preview_font_bold)
        self._reload_student_rows()
        self._set_scan_status(prefix="Scan info")

    def _refresh_tree_row(self, folder_key: str):
        r = self.rows[folder_key]
        self.tree.item(folder_key, values=(
            "YES" if r["include"] else "NO",
            Path(r["folder"]).name,
            r["det_id"],
            r["det_name"],
            r["final_id"],
            r["final_name"],
            r.get("lab_id",""),
            str(len(r["files"]))
        ))

    def apply_global_lab_id(self):
        lab = (self.global_lab_id_var.get() or "").strip()
        for folder_key in self.folder_order:
            r = self.rows.get(folder_key)
            if not r or not r.get("include"):
                continue
            r["lab_id"] = lab
            self._refresh_tree_row(folder_key)

        current = self.selected_folder_key or self.sel_folder_var.get().strip()
        if current and current in self.rows:
            self._suspend_auto_apply = True
            self.lab_id_var.set(self.rows[current].get("lab_id", ""))
            self._suspend_auto_apply = False

        self._reload_student_rows()
        self._set_scan_status(prefix="Scan info")

    def _selected_preview_text(self) -> str:
        try:
            return self.preview.get("sel.first", "sel.last").strip()
        except tk.TclError:
            return ""

    def _hotkey_use_id(self, _evt=None):
        self.use_selection_as_id()
        return "break"

    def _hotkey_use_name(self, _evt=None):
        self.use_selection_as_name()
        return "break"

    def use_selection_as_id(self):
        selected = self._selected_preview_text()
        if not selected:
            return
        # Prefer numeric extraction, but still allow full raw selection so
        # quick-paste IDs such as NAME:... or custom keys can be used.
        only_id = extract_numeric_id(selected)
        self.final_id_var.set(only_id if only_id else selected)

    def _selected_folder_from_tree(self) -> str | None:
        sel = self.tree.selection()
        if not sel:
            return None
        folder_key = sel[0]
        return folder_key if folder_key in self.rows else None

    def _build_skimmable_sequence(self, start_folder_key: str | None, only_unassigned: bool = False) -> list[str]:
        folders = self._get_skimmable_folders(only_unassigned=only_unassigned)
        if not folders:
            return []
        if not start_folder_key or start_folder_key not in folders:
            return folders

        start_idx = folders.index(start_folder_key)
        return folders[start_idx:] + folders[:start_idx]

    def use_selection_as_name(self):
        selected = self._selected_preview_text()
        if not selected:
            return
        self.final_name_var.set(selected)

    def focus_find_entry(self):
        self.find_entry.focus_set()
        return "break"

    def find_next(self):
        query = (self.find_var.get() or "").strip()
        self.preview.tag_remove("find_hit", "1.0", tk.END)
        if not query:
            return

        content = self.preview.get("1.0", tk.END)
        start_offset = 0
        if self._find_from and "." in self._find_from:
            try:
                start_offset = int(self.preview.count("1.0", self._find_from, "chars")[0])
            except Exception:
                start_offset = 0

        flags = 0 if self.find_case_sensitive_var.get() else re.IGNORECASE

        if self.find_regex_var.get():
            try:
                pattern = re.compile(query, flags)
            except re.error as e:
                messagebox.showerror("Invalid regex", str(e))
                return
            match = pattern.search(content, pos=start_offset) or pattern.search(content, pos=0)
            if not match:
                return
            sidx, eidx = match.span()
            if eidx <= sidx:
                eidx = sidx + 1
            start = f"1.0+{sidx}c"
            end = f"1.0+{eidx}c"
        else:
            start = self.preview.search(query, self._find_from, stopindex=tk.END, nocase=not self.find_case_sensitive_var.get())
            if not start:
                start = self.preview.search(query, "1.0", stopindex=tk.END, nocase=not self.find_case_sensitive_var.get())
                if not start:
                    return
            end = f"{start}+{len(query)}c"

        self.preview.tag_add("find_hit", start, end)
        self.preview.mark_set(tk.INSERT, end)
        self.preview.see(start)
        self.preview.focus_set()
        self._find_from = end

    def _get_skimmable_folders(self, only_unassigned: bool = False) -> list[str]:
        """
        Skim every folder that has at least one file so names/IDs can be verified quickly.
        Optional: keep only unassigned/non-student folders.
        """
        out = []
        for folder_key in self.folder_order:
            row = self.rows.get(folder_key) or {}
            if not (row.get("files") or []):
                continue
            if only_unassigned and has_required_student_fields(row.get("final_id", ""), row.get("final_name", "")):
                continue
            out.append(folder_key)
        return out

    def _select_folder_by_index(self, idx: int):
        if idx < 0 or idx >= len(self.folder_order):
            return None
        folder_key = self.folder_order[idx]
        self.tree.selection_set(folder_key)
        self.tree.focus(folder_key)
        self.tree.see(folder_key)
        self.on_folder_select()
        return folder_key

    def _reload_tree_rows(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for folder_key in self.folder_order:
            r = self.rows[folder_key]
            self.tree.insert("", "end", iid=folder_key, values=(
                "YES" if r["include"] else "NO",
                Path(r["folder"]).name,
                r["det_id"],
                r["det_name"],
                r["final_id"],
                r["final_name"],
                r.get("lab_id", ""),
                str(len(r["files"]))
            ))

        self._reload_student_rows()

    def _reload_student_rows(self):
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)

        idx = 0
        for folder_key in self.folder_order:
            r = self.rows.get(folder_key) or {}
            if not has_required_student_fields(r.get("final_id", ""), r.get("final_name", "")):
                continue
            idx += 1
            self.student_tree.insert("", "end", iid=f"student-{idx}", values=(
                r.get("final_id", ""),
                r.get("final_name", ""),
                r.get("lab_id", ""),
                Path(r.get("folder", folder_key)).name,
                str(len(r.get("files") or [])),
            ))

    def _select_file_in_current_folder(self, file_idx: int) -> bool:
        self.update_idletasks()
        children = self.files_tree.get_children()
        size = len(children)
        if file_idx < 0 or file_idx >= size:
            return False
        iid = children[file_idx]
        self.files_tree.selection_set(iid)
        self.files_tree.focus(iid)
        self.files_tree.see(iid)
        self.files_tree.focus_set()
        self.update_idletasks()
        self.files_tree.event_generate("<<TreeviewSelect>>")
        self.on_scan_file_select(file_iid=iid)
        return True

    def start_skimming(self, only_unassigned: bool = False):
        if not self.folder_order:
            if self.root_folder:
                self.scan()
            else:
                self.load_existing_rows_from_db()
        if not self.folder_order:
            messagebox.showinfo("Skimming", "No rows loaded. Scan folders or load from DB first.")
            return

        start_folder_key = self._selected_folder_from_tree()
        self._skimmable_folder_keys = self._build_skimmable_sequence(start_folder_key, only_unassigned=only_unassigned)
        if not self._skimmable_folder_keys:
            if only_unassigned:
                messagebox.showinfo("Skimming", "No unassigned/non-student folders with files to skim.")
            else:
                messagebox.showinfo("Skimming", "No folders with files to skim.")
            return

        try:
            delay = int(self.skim_delay_ms_var.get())
        except Exception:
            delay = 300
        if delay < 100:
            delay = 100
        self.skim_delay_ms_var.set(delay)

        self._skim_folder_idx = 0
        self._skim_file_idx = 0
        self.skim_running = True

        first_folder_key = self._skimmable_folder_keys[0]
        try:
            first_folder_idx = self.folder_order.index(first_folder_key)
            self._select_folder_by_index(first_folder_idx)
        except Exception:
            pass

        self._skim_total_files = sum(len(self.rows.get(k, {}).get("files") or []) for k in self._skimmable_folder_keys)
        self._skim_seen_files = 0
        self._set_scan_status(prefix="Skimming started")
        self.after(int(self.skim_delay_ms_var.get()), self._skim_step)

    def start_skimming_unassigned(self):
        self.start_skimming(only_unassigned=True)

    def stop_skimming(self):
        if self.skim_running:
            self.skim_running = False
            self._set_scan_status(prefix="Skimming stopped")
        return "break"

    def _skim_step(self):
        if not self.skim_running:
            return

        if self._skim_folder_idx >= len(self._skimmable_folder_keys):
            self.skim_running = False
            self._set_scan_status(prefix="Skimming done")
            return

        folder_key = self._skimmable_folder_keys[self._skim_folder_idx]
        files = self.rows.get(folder_key, {}).get("files") or []
        if not files:
            self._skim_folder_idx += 1
            self._skim_file_idx = 0
            self.after(int(self.skim_delay_ms_var.get()), self._skim_step)
            return

        try:
            folder_idx = self.folder_order.index(folder_key)
        except ValueError:
            self._skim_folder_idx += 1
            self._skim_file_idx = 0
            self.after(int(self.skim_delay_ms_var.get()), self._skim_step)
            return

        current_selected = self._selected_folder_from_tree()
        if current_selected != folder_key:
            if not self._select_folder_by_index(folder_idx):
                self._skim_folder_idx += 1
                self._skim_file_idx = 0
                self.after(int(self.skim_delay_ms_var.get()), self._skim_step)
                return

        if self._skim_file_idx >= len(files):
            self._skim_folder_idx += 1
            self._skim_file_idx = 0
            self.after(int(self.skim_delay_ms_var.get()), self._skim_step)
            return

        selected = self._select_file_in_current_folder(self._skim_file_idx)
        if selected:
            self._skim_seen_files += 1
            self._set_scan_status(prefix=f"Skimming: student {self._skim_folder_idx + 1}/{len(self._skimmable_folder_keys)}, file {self._skim_file_idx + 1}/{len(files)} (overall {self._skim_seen_files}/{max(1, self._skim_total_files)})")

        self._skim_file_idx += 1
        self.after(int(self.skim_delay_ms_var.get()), self._skim_step)

    def save_to_db(self, show_message: bool = True):
        if not self.rows:
            self.load_existing_rows_from_db()
        if not self.rows:
            messagebox.showinfo("Nothing", "No rows loaded. Scan folders or load from DB first.")
            return

        store_content = self.store_file_content.get()

        committed_folders = 0
        committed_files = 0
        created_students = 0
        skipped_folders = 0

        with self.con:
            for folder_key in self.folder_order:
                r = self.rows[folder_key]
                if not r.get("files"):
                    skipped_folders += 1
                    continue

                include_row = bool(r.get("include"))
                fid = (r["final_id"] or "").strip()
                fname = (r["final_name"] or "").strip()
                lab = (r.get("lab_id") or "").strip() or None

                student_ok = has_required_student_fields(fid, fname)
                normalized_sid = (fid or "").strip()
                if normalized_sid.lower() != "full":
                    normalized_sid = extract_numeric_id(normalized_sid)

                if include_row and student_ok and normalized_sid:
                    upsert_student(self.con, normalized_sid, fname, lab, r.get("folder") or folder_key, included=True, commit=False)
                    created_students += 1
                elif include_row and student_ok and (fid or "").strip().lower() == "full":
                    upsert_student(self.con, "FULL", fname or "FULL", lab, r.get("folder") or folder_key, included=True, commit=False)
                    normalized_sid = "FULL"
                    created_students += 1
                elif not include_row:
                    skipped_folders += 1
                    if student_ok and normalized_sid:
                        set_student_included(self.con, normalized_sid, False, commit=False)

                for fp in r["files"]:
                    pth = Path(fp)
                    file_text = None
                    file_hash = None
                    try:
                        if store_content:
                            file_text = read_file_text(pth)
                            file_hash = sha256_text(file_text)
                    except Exception:
                        file_text = None
                        file_hash = None

                    upsert_file(
                        self.con,
                        file_path=fp,
                        student_id=normalized_sid if include_row and student_ok and normalized_sid else None,
                        source_folder=r.get("folder") or folder_key,
                        detected_id=fid or r["det_id"] or None,
                        detected_name=fname or r["det_name"] or None,
                        file_hash=file_hash,
                        file_content=file_text,
                        commit=False,
                    )
                    committed_files += 1

                committed_folders += 1

        self.app.refresh_students(keep_selected=False)
        if show_message:
            messagebox.showinfo(
                "Saved",
                f"Folders committed: {committed_folders}\nFolders skipped: {skipped_folders}\nFiles saved: {committed_files}\nStudents created/updated: {created_students}"
            )


# =============================================================================
# 12) Main App
# =============================================================================

def clamp_points(val: float | None, mx: float) -> float | None:
    if val is None:
        return None
    if val < 0:
        return 0.0
    if val > mx:
        return mx
    return val

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Java Grading Gallery — v2.2 (Assignments + Comments + PDF + Pastel + LabID + Histogram)")
        self.root.geometry("1920x1040")

        self.palette = pastel_style(root)

        self.sub_db_path = Path(SUBMISSIONS_DB)
        self.sub_con = db_connect(self.sub_db_path)
        submissions_db_init(self.sub_con)
        grading_db_init(self.sub_con)

        # Grading data is stored in the same submissions DB.
        self.grade_db_path: Path | None = self.sub_db_path
        self.grade_con: sqlite3.Connection | None = self.sub_con

        self.student_ids: list[str] = []
        self.selected_student_id: str | None = None
        self.selected_file_path: str | None = None

        self.question_map: dict[str, str] = {}
        self.question_display_to_id: dict[str, str] = {}
        self.selected_question_id: str | None = None
        self.question_pick_var = tk.StringVar(value="")
        self._rubric_ui_map: dict[str, tuple[str, str]] = {}
        self._rubric_max_map: dict[str, float] = {}
        self._auto_save_job = None
        self._suspend_auto_save = False
        self.session_started_at = None
        self.session_elapsed_seconds = 0
        self.session_timer_running = False
        self.grade_meta_var = tk.StringVar(value="Graded: NO | Reviewed: YES | First graded: - | Last updated: -")
        self.progress_var = tk.StringVar(value="Progress: assessed 0/0 | left 0")
        self.main_menu_progress_var = tk.StringVar(value="Students assessed: 0/0 | Left: 0 | Unreviewed: 0 | Curve preview ×1.00")
        self._clock_job = None

        # curve preview factor (histogram overlay)
        self.curve_preview_var = tk.DoubleVar(value=1.0)
        self.leniency_level_var = tk.DoubleVar(value=0.0)
        self.leniency_label_var = tk.StringVar(value="Leniency level: 0.00 (strict ↔ lenient)")

        # Regex profile state
        self.active_regex_profile_var = tk.StringVar(value="Default")

        self.gpt_api_key_var = tk.StringVar(value="")
        self.gpt_model_var = tk.StringVar(value="gpt-4.1-mini")
        self.gpt_remote_enabled_var = tk.BooleanVar(value=False)
        self.auto_prompt_text_widget = None
        self.theme_settings_text = None
        self.prompt_process_text_widget = None
        self.prompt_result_text_widget = None
        self.chat_message_widget = None
        self.chat_preview_widget = None
        self.chat_transcript_widget = None
        self.chat_bundle_widget = None
        self.chat_auto_bundle_var = tk.BooleanVar(value=True)
        self.chat_include_code_var = tk.BooleanVar(value=True)
        self.chat_include_scheme_var = tk.BooleanVar(value=True)
        self.chat_include_prompt_var = tk.BooleanVar(value=True)
        self.chat_compact_code_var = tk.BooleanVar(value=True)
        self.chat_code_char_limit_var = tk.IntVar(value=7000)
        self.chat_student_var = tk.StringVar(value="Chat student: none selected")
        self.pdf_menu_include_student_var = tk.BooleanVar(value=True)
        self.pdf_menu_include_summary_var = tk.BooleanVar(value=True)
        self.pdf_menu_include_batch_var = tk.BooleanVar(value=False)
        self.pdf_menu_batch_tag_var = tk.StringVar(value="Midterm")
        self.pdf_content_grades_var = tk.BooleanVar(value=True)
        self.pdf_content_rationale_var = tk.BooleanVar(value=True)
        self.pdf_content_highlights_var = tk.BooleanVar(value=True)
        self.pdf_content_full_code_var = tk.BooleanVar(value=True)
        self.last_auto_grade_trace = {"prompt": "", "result": "", "source": ""}

        self.db_table_pick_var = tk.StringVar(value="")
        self.db_table_listbox = None
        self.db_schema_text = None
        self.db_rows_tree = None
        self.db_rows_status_var = tk.StringVar(value="No table selected.")

        # Auto-grader (heuristic + optional GPT)
        self.gpt_tester = GPT_test()
        self.auto_grader = AutoGrader(self.gpt_tester)

        self._grade_last_student_selection: tuple[int, ...] = ()
        self._grade_last_file_selection: tuple[int, ...] = ()

        self._build_ui()
        self._load_theme_and_leniency_from_db()
        self.refresh_students(keep_selected=False)
        self.root.after(120, self._poll_grade_selections)

    def _poll_grade_selections(self):
        student_sel = tuple(self.student_list.curselection())
        if student_sel != self._grade_last_student_selection:
            self._grade_last_student_selection = student_sel
            if student_sel:
                self.on_student_select()

        file_sel = tuple(self.file_list.curselection())
        if file_sel != self._grade_last_file_selection:
            self._grade_last_file_selection = file_sel
            if file_sel:
                self.on_grade_file_select()

        self.root.after(120, self._poll_grade_selections)

    def require_grading_db(self) -> bool:
        if self.grade_con is None:
            self.grade_con = self.sub_con
            self.grade_db_path = self.sub_db_path
            grading_db_init(self.grade_con)
        return True

    def _load_theme_and_leniency_from_db(self):
        if not self.require_grading_db():
            return
        theme = meta_get(self.grade_con, "theme", DEFAULT_THEME)
        self.theme_text.delete("1.0", tk.END)
        self.theme_text.insert("1.0", theme)
        if self.theme_settings_text is not None:
            self.theme_settings_text.delete("1.0", tk.END)
            self.theme_settings_text.insert("1.0", theme)
        try:
            self.leniency_level_var.set(max(-1.0, min(1.0, float(meta_get(self.grade_con, "leniency_level", "0") or 0.0))))
        except Exception:
            self.leniency_level_var.set(0.0)
        self._on_leniency_change()

    def _build_ui(self):
        top = ttk.Frame(self.root, padding=10, style="Pastel.TFrame")
        top.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(top, text="Open Scan / Edit", command=self.open_scan_window).pack(side=tk.LEFT)

        ttk.Separator(top, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(top, text="New Submissions DB...", command=self.new_submissions_db).pack(side=tk.LEFT)
        ttk.Button(top, text="Open Submissions DB...", command=self.open_submissions_db).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Save Submissions DB As...", command=self.save_submissions_db_as).pack(side=tk.LEFT, padx=6)

        ttk.Separator(top, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(top, text="Load Scheme CSV...", command=self.load_scheme_csv).pack(side=tk.LEFT)
        ttk.Button(top, text="Open Scheme Generator / Editor", command=self.open_scheme_editor).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Reload Last Scheme", command=self.reload_last_scheme).pack(side=tk.LEFT, padx=6)

        ttk.Separator(top, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        self.sub_db_status_lbl = ttk.Label(top, text=f"Submissions DB: {self.sub_db_path.name}", style="Pastel.TLabel")
        self.sub_db_status_lbl.pack(side=tk.LEFT, padx=10)

        self.db_status_lbl = ttk.Label(top, text=f"Grading: integrated with {self.sub_db_path.name}", style="Pastel.TLabel")
        self.db_status_lbl.pack(side=tk.LEFT, padx=10)

        self.student_count_lbl = ttk.Label(top, text="Students: 0", style="Pastel.TLabel")
        self.student_count_lbl.pack(side=tk.RIGHT)

        # Histogram + stats area on top right
        stats_frame = ttk.Frame(top, style="Pastel.TFrame")
        stats_frame.pack(side=tk.RIGHT, padx=14)

        self.class_stats_lbl = ttk.Label(stats_frame, text="Class: avg - | min - | max - | curve -", style="Pastel.TLabel")
        self.class_stats_lbl.pack(side=tk.TOP, anchor="e")

        self.main_menu_progress_lbl = ttk.Label(stats_frame, textvariable=self.main_menu_progress_var, style="Pastel.TLabel")
        self.main_menu_progress_lbl.pack(side=tk.TOP, anchor="e", pady=(2, 0))

        curve_row = ttk.Frame(stats_frame, style="Pastel.TFrame")
        curve_row.pack(side=tk.TOP, anchor="e", pady=(2, 0))
        ttk.Label(curve_row, text="Curve preview ×", style="Pastel.TLabel").pack(side=tk.LEFT)
        self.curve_scale = ttk.Scale(curve_row, from_=0.8, to=1.3, variable=self.curve_preview_var, command=lambda _v: self.refresh_summary())
        self.curve_scale.pack(side=tk.LEFT, padx=(6, 0))

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill=tk.BOTH, expand=True)

        self.tab_grade = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_summary = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_stats = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_progress = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_regex = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_ai_trace = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_settings = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_pdf_menu = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)
        self.tab_db = ttk.Frame(self.nb, style="Pastel.TFrame", padding=10)

        self.nb.add(self.tab_grade, text="Grade")
        self.nb.add(self.tab_summary, text="Summary")
        self.nb.add(self.tab_stats, text="Stats")
        self.nb.add(self.tab_progress, text="Progress")
        self.nb.add(self.tab_regex, text="Regex / Patterns")
        self.nb.add(self.tab_ai_trace, text="Prompt + Chat")
        self.nb.add(self.tab_settings, text="Settings")
        self.nb.add(self.tab_pdf_menu, text="PDF Menu")
        self.nb.add(self.tab_db, text="DB Browser")

        self._build_grade_tab()
        self._build_summary_tab()
        self._build_stats_tab()
        self._build_progress_tab()
        self._build_regex_tab()
        self._build_ai_trace_tab()
        self._build_settings_hub_tab()
        self._build_pdf_menu_tab()
        self._build_db_tab()
        self._ensure_default_regex_profile()
        self.load_gpt_settings()
        self.load_ui_preferences()
        self._on_leniency_change()
        self.reset_session_timer()

    def _build_grade_tab(self):
        main = ttk.Frame(self.tab_grade, style="Pastel.TFrame")
        main.pack(fill=tk.BOTH, expand=True)

        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=2)
        main.columnconfigure(2, weight=2)
        main.rowconfigure(0, weight=1)

        # LEFT
        left = ttk.Frame(main, style="PastelCard.TFrame", padding=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(3, weight=1)
        left.columnconfigure(0, weight=1)

        ttk.Label(left, text="Students", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.student_list = tk.Listbox(left, bg=self.palette["panel"], fg=self.palette["text"], highlightthickness=0, selectbackground=self.palette["select"])
        self.student_list.grid(row=1, column=0, sticky="nsew")

        ttk.Label(left, text="Files", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="w", pady=(10, 0))
        self.file_list = tk.Listbox(left, height=8, bg=self.palette["panel"], fg=self.palette["text"], highlightthickness=0, selectbackground=self.palette["select"])
        self.file_list.grid(row=3, column=0, sticky="nsew")
        self.file_list.bind("<<ListboxSelect>>", self.on_grade_file_select)

        # MIDDLE preview + comments
        mid = ttk.Frame(main, style="PastelCard.TFrame", padding=10)
        mid.grid(row=0, column=1, sticky="nsew", padx=(0, 10))
        mid.rowconfigure(3, weight=1)
        mid.columnconfigure(0, weight=1)

        self.student_header = ttk.Label(mid, text="No student selected", style="PastelCard.TLabel", font=("Segoe UI", 12, "bold"))
        self.student_header.grid(row=0, column=0, sticky="w")

        self.grade_meta_lbl = ttk.Label(mid, textvariable=self.grade_meta_var, style="PastelCard.TLabel")
        self.grade_meta_lbl.grid(row=1, column=0, sticky="w", pady=(4, 4))

        codebar = ttk.Frame(mid, style="PastelCard.TFrame")
        codebar.grid(row=2, column=0, sticky="ew", pady=(6, 6))
        ttk.Button(codebar, text="Add comment to selection", command=self.add_comment_to_selection).pack(side=tk.LEFT)
        ttk.Button(codebar, text="Clear comments in selection", command=self.clear_comments_in_selection).pack(side=tk.LEFT, padx=6)
        ttk.Button(codebar, text="Export PDF (this student)", command=self.export_student_pdf).pack(side=tk.RIGHT)

        preview_frame = ttk.Frame(mid, style="PastelCard.TFrame")
        preview_frame.grid(row=3, column=0, sticky="nsew")
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        self.preview = tk.Text(preview_frame, wrap="none",
                               bg="#FFFDF7", fg=self.palette["text"],
                               insertbackground=self.palette["text"],
                               selectbackground=self.palette["select"],
                               highlightthickness=1, highlightbackground="#E8E1FF")
        self.preview.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview.yview)
        sb.grid(row=0, column=1, sticky="ns")
        self.preview.configure(yscrollcommand=sb.set)

        self.preview.tag_configure("comment_highlight", background="#FFF9A6")

        # RIGHT grading
        right = ttk.Frame(main, style="PastelCard.TFrame", padding=10)
        right.grid(row=0, column=2, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(15, weight=1)

        ttk.Label(right, text="Questions: all loaded rubric questions", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.question_mode_lbl = ttk.Label(right, text="Rubric below shows all questions (question picker removed)", style="PastelCard.TLabel")
        self.question_mode_lbl.grid(row=1, column=0, sticky="w", pady=(4,8))

        ttk.Separator(right, orient="horizontal").grid(row=4, column=0, sticky="ew", pady=(8,10))

        ttk.Label(right, text="Theme / Instructions (saved in DB)", style="PastelCard.TLabel").grid(row=5, column=0, sticky="w", pady=(10, 0))
        self.theme_text = tk.Text(right, height=4,
                                  bg="#FFFDF7", fg=self.palette["text"],
                                  insertbackground=self.palette["text"],
                                  highlightthickness=1, highlightbackground="#E8E1FF")
        self.theme_text.grid(row=6, column=0, sticky="nsew")
        self.theme_text.insert("1.0", DEFAULT_THEME)

        leniency_row = ttk.Frame(right, style="PastelCard.TFrame")
        leniency_row.grid(row=7, column=0, sticky="ew", pady=(8, 0))
        ttk.Label(leniency_row, textvariable=self.leniency_label_var, style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Scale(leniency_row, from_=-1.0, to=1.0, variable=self.leniency_level_var, command=lambda _v: self._on_leniency_change()).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))

        btns = ttk.Frame(right, style="PastelCard.TFrame")
        btns.grid(row=8, column=0, sticky="ew", pady=(8, 8))

        ttk.Button(btns, text="Save Theme", command=lambda: self.save_theme(source="grade")).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(btns, text="AutoFill", command=self.auto_fill_student).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Grade Files", command=self.auto_grade_files_for_student).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Grade All", command=self.auto_grade_all_students).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Clear Grade", command=self.clear_selected_student_grade).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Save (all questions)", command=self.save_scores_and_rationale).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Mark Reviewed", command=self.mark_selected_student_reviewed).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(btns, text="Export Grade (selected student)", command=self.export_selected_excel).pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)

        ttk.Label(right, text="Rationale (applied to all questions for this student)", style="PastelCard.TLabel").grid(row=9, column=0, sticky="w")
        self.rationale_text = tk.Text(right, height=6,
                                      bg="#FFFDF7", fg=self.palette["text"],
                                      insertbackground=self.palette["text"],
                                      highlightthickness=1, highlightbackground="#E8E1FF")
        self.rationale_text.grid(row=10, column=0, sticky="nsew")
        self.rationale_text.bind("<KeyRelease>", lambda _e: self.schedule_auto_save())
        self.rationale_text.bind("<FocusOut>", lambda _e: self.schedule_auto_save())

        self.total_lbl = ttk.Label(right, text="Overall Total: -", style="PastelCard.TLabel", font=("Segoe UI", 11, "bold"))
        self.total_lbl.grid(row=11, column=0, sticky="w", pady=(6, 6))

        ttk.Label(right, text="Rubric Table (all questions)", style="PastelCard.TLabel").grid(row=12, column=0, sticky="w")
        self.rubric_grid = ScrollableRubricGrid(right)
        self.rubric_grid.set_change_callback(self.schedule_auto_save)
        self.rubric_grid.grid(row=13, column=0, sticky="nsew")

        ttk.Label(right, text="Code comments (this file)", style="PastelCard.TLabel").grid(row=14, column=0, sticky="w", pady=(10,0))
        self.comment_list = tk.Listbox(right, height=7, bg=self.palette["panel"], fg=self.palette["text"],
                                       highlightthickness=0, selectbackground=self.palette["select"])
        self.comment_list.grid(row=15, column=0, sticky="nsew")

    def _current_theme_instructions(self, source: str = "auto") -> str:
        grade_theme = self.theme_text.get("1.0", tk.END).strip() if self.theme_text is not None else ""
        settings_theme = self.theme_settings_text.get("1.0", tk.END).strip() if self.theme_settings_text is not None else ""

        theme = ""
        if source == "grade":
            theme = grade_theme or settings_theme
        elif source == "settings":
            theme = settings_theme or grade_theme
        else:
            focus_widget = self.root.focus_get()
            if focus_widget is self.theme_settings_text:
                theme = settings_theme or grade_theme
            elif focus_widget is self.theme_text:
                theme = grade_theme or settings_theme
            elif grade_theme and settings_theme and grade_theme != settings_theme:
                theme = grade_theme
            else:
                theme = grade_theme or settings_theme

        theme = theme or DEFAULT_THEME

        if self.theme_settings_text is not None:
            self.theme_settings_text.delete("1.0", tk.END)
            self.theme_settings_text.insert("1.0", theme)
        if self.theme_text is not None:
            self.theme_text.delete("1.0", tk.END)
            self.theme_text.insert("1.0", theme)
        return theme

    def _on_leniency_change(self):
        self.leniency_label_var.set(f"Leniency level: {float(self.leniency_level_var.get()):.2f} (strict ↔ lenient)")

    def _build_ai_trace_tab(self):
        self.tab_ai_trace.columnconfigure(0, weight=1)
        self.tab_ai_trace.rowconfigure(1, weight=1)

        top = ttk.Frame(self.tab_ai_trace, style="Pastel.TFrame")
        top.grid(row=0, column=0, sticky="ew")
        ttk.Label(top, text="Prompt process + model output + chat", style="Pastel.TLabel", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        ttk.Label(top, textvariable=self.chat_student_var, style="Pastel.TLabel").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(top, text="Refresh latest grading trace", command=self.refresh_prompt_trace_tab).pack(side=tk.LEFT, padx=8)

        panes = ttk.Panedwindow(self.tab_ai_trace, orient=tk.HORIZONTAL)
        panes.grid(row=1, column=0, sticky="nsew", pady=(8, 0))

        left = ttk.Frame(panes, style="Pastel.TFrame", padding=(0, 0, 8, 0))
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)
        left.rowconfigure(3, weight=1)
        left.rowconfigure(5, weight=2)

        process_box = ttk.Frame(left, style="PastelCard.TFrame", padding=8)
        process_box.grid(row=0, column=0, sticky="nsew", pady=(0, 6))
        process_box.columnconfigure(0, weight=1)
        process_box.rowconfigure(1, weight=1)
        ttk.Label(process_box, text="Prompt process", style="PastelCard.TLabel").grid(row=0, column=0, sticky="w")
        self.prompt_process_text_widget = tk.Text(process_box, height=8, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.prompt_process_text_widget.grid(row=1, column=0, sticky="nsew", pady=(4, 0))

        result_box = ttk.Frame(left, style="PastelCard.TFrame", padding=8)
        result_box.grid(row=1, column=0, sticky="nsew", pady=(0, 6))
        result_box.columnconfigure(0, weight=1)
        result_box.rowconfigure(1, weight=1)
        ttk.Label(result_box, text="What the grader returned", style="PastelCard.TLabel").grid(row=0, column=0, sticky="w")
        self.prompt_result_text_widget = tk.Text(result_box, height=6, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.prompt_result_text_widget.grid(row=1, column=0, sticky="nsew", pady=(4, 0))

        bundle_box = ttk.Frame(left, style="PastelCard.TFrame", padding=8)
        bundle_box.grid(row=2, column=0, sticky="nsew", pady=(0, 6))
        bundle_box.columnconfigure(0, weight=1)
        bundle_box.rowconfigure(3, weight=1)
        ttk.Label(bundle_box, text="Bundle builder", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        checks = ttk.Frame(bundle_box, style="PastelCard.TFrame")
        checks.grid(row=1, column=0, sticky="w", pady=(4, 6))
        ttk.Checkbutton(checks, text="Include student code", variable=self.chat_include_code_var).pack(side=tk.LEFT)
        ttk.Checkbutton(checks, text="Include rubric scheme", variable=self.chat_include_scheme_var).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Checkbutton(checks, text="Include prompt process + output", variable=self.chat_include_prompt_var).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Checkbutton(checks, text="Compact code", variable=self.chat_compact_code_var).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(checks, text="Code chars", style="PastelCard.TLabel").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Entry(checks, textvariable=self.chat_code_char_limit_var, width=8).pack(side=tk.LEFT)

        leniency_row = ttk.Frame(bundle_box, style="PastelCard.TFrame")
        leniency_row.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        ttk.Label(leniency_row, textvariable=self.leniency_label_var, style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Scale(
            leniency_row,
            from_=-1.0,
            to=1.0,
            variable=self.leniency_level_var,
            command=lambda _v: self._on_leniency_change(),
        ).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))

        self.chat_bundle_widget = tk.Text(bundle_box, height=10, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.chat_bundle_widget.grid(row=3, column=0, sticky="nsew", pady=(0, 6))

        bundle_btns = ttk.Frame(bundle_box, style="PastelCard.TFrame")
        bundle_btns.grid(row=4, column=0, sticky="w")
        ttk.Button(bundle_btns, text="Build bundle", command=self.refresh_chat_preview).pack(side=tk.LEFT)
        ttk.Button(bundle_btns, text="Copy bundle", command=self.copy_chat_bundle).pack(side=tk.LEFT, padx=6)
        ttk.Checkbutton(bundle_btns, text="Auto-refresh bundle", variable=self.chat_auto_bundle_var).pack(side=tk.LEFT, padx=(6, 0))

        settings_box = ttk.Frame(left, style="PastelCard.TFrame", padding=8)
        settings_box.grid(row=3, column=0, sticky="nsew")
        settings_box.columnconfigure(0, weight=1)
        self._build_settings_tab(parent=settings_box)

        right = ttk.Frame(panes, style="Pastel.TFrame", padding=(8, 0, 0, 0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        chat_box = ttk.Frame(right, style="PastelCard.TFrame", padding=10)
        chat_box.grid(row=0, column=0, sticky="nsew")
        chat_box.columnconfigure(0, weight=1)
        chat_box.rowconfigure(1, weight=1)
        ttk.Label(chat_box, text="Chat", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")

        self.chat_transcript_widget = tk.Text(chat_box, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.chat_transcript_widget.grid(row=1, column=0, sticky="nsew", pady=(4, 8))

        ttk.Label(chat_box, text="Your message", style="PastelCard.TLabel").grid(row=2, column=0, sticky="w")
        self.chat_message_widget = tk.Text(chat_box, height=4, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.chat_message_widget.grid(row=3, column=0, sticky="ew", pady=(4, 6))

        chat_btns = ttk.Frame(chat_box, style="PastelCard.TFrame")
        chat_btns.grid(row=4, column=0, sticky="w")
        ttk.Button(chat_btns, text="Send message", command=self.send_chat_message).pack(side=tk.LEFT)
        ttk.Button(chat_btns, text="Send bundle", command=self.send_bundle_message).pack(side=tk.LEFT, padx=6)
        ttk.Button(chat_btns, text="Clear chat", command=self.clear_chat_transcript).pack(side=tk.LEFT, padx=6)

        panes.add(left, weight=3)
        panes.add(right, weight=2)

        self.chat_preview_widget = self.chat_bundle_widget
        self.refresh_prompt_trace_tab()
        self.refresh_chat_preview()

    def _format_last_prompt_process(self):
        data = self.last_auto_grade_trace.get("prompt")
        if not data:
            return "No prompt process captured yet. Run Grade Files first."
        return data

    def _format_last_result(self):
        data = self.last_auto_grade_trace.get("result")
        if not data:
            return "No grading output captured yet."
        return data

    def refresh_prompt_trace_tab(self):
        if self.prompt_process_text_widget is None or self.prompt_result_text_widget is None:
            return
        self.prompt_process_text_widget.delete("1.0", tk.END)
        self.prompt_process_text_widget.insert("1.0", self._format_last_prompt_process())
        self.prompt_result_text_widget.delete("1.0", tk.END)
        self.prompt_result_text_widget.insert("1.0", self._format_last_result())

    def _build_scheme_text(self):
        if not self.grade_con:
            return ""
        rows = fetch_rubric_parts(self.grade_con)
        if not rows:
            return ""
        lines = ["# Rubric scheme"]
        for qid, col_key, group, col_text, col_max in rows:
            lines.append(f"- {qid} | {col_key} | {group or '-'} | max {col_max:g} | {col_text}")
        return "\n".join(lines)

    def _selected_student_label(self):
        if not self.selected_student_id:
            return "none selected"
        row = self.sub_con.execute("SELECT COALESCE(student_name,'') FROM students WHERE student_id=?", (self.selected_student_id,)).fetchone()
        name = (row[0] or "").strip() if row else ""
        return f"{self.selected_student_id} — {name}" if name else self.selected_student_id

    def _build_code_bundle_text(self):
        if not self.selected_student_id:
            return ""
        merged_code, _file_map, _line_ranges = self._merged_code_and_file_map(self.selected_student_id)
        max_chars = max(1000, int(self.chat_code_char_limit_var.get() or 7000))
        compact = bool(self.chat_compact_code_var.get())
        if compact and len(merged_code) > max_chars:
            merged_code = merged_code[:max_chars] + "\n\n// [truncated for token efficiency]"
        header = f"# Student: {self._selected_student_label()}\n"
        return header + merged_code

    def _compose_chat_bundle(self):
        parts = []
        if self.chat_include_code_var.get():
            code = self._build_code_bundle_text()
            if code:
                parts.append(code)
        if self.chat_include_scheme_var.get():
            scheme = self._build_scheme_text()
            if scheme:
                parts.append(scheme)
        if self.chat_include_prompt_var.get():
            parts.append("# Prompt process\n" + self._format_last_prompt_process())
            parts.append("# Model output\n" + self._format_last_result())

        msg = self.chat_message_widget.get("1.0", tk.END).strip() if self.chat_message_widget else ""
        if msg:
            parts.append("# My message\n" + msg)
        return "\n\n".join(p for p in parts if p.strip())

    def refresh_chat_preview(self):
        if self.chat_preview_widget is None:
            return
        payload = self._compose_chat_bundle()
        self.chat_preview_widget.delete("1.0", tk.END)
        self.chat_preview_widget.insert("1.0", payload or "Nothing selected yet. Tick at least one source and add a message.")

    def copy_chat_bundle(self):
        payload = self._compose_chat_bundle()
        if not payload.strip():
            messagebox.showinfo("Nothing to copy", "Pick at least one item and/or type your chat message first.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(payload)
        self.root.update_idletasks()
        self.refresh_chat_preview()
        messagebox.showinfo("Copied", "Chat bundle copied to clipboard.")

    def _append_chat_transcript(self, role: str, text: str):
        if self.chat_transcript_widget is None:
            return
        stamp = datetime.now().strftime("%H:%M:%S")
        self.chat_transcript_widget.insert(tk.END, f"[{stamp}] {role}\n{text.strip()}\n\n")
        self.chat_transcript_widget.see(tk.END)

    def _send_to_chat(self, message: str, context_bundle: str, show_label: str):
        self._append_chat_transcript(show_label, message)
        try:
            # Keep runtime settings in sync so chat immediately uses any API/model edits
            # without requiring an explicit "Save GPT Settings" click first.
            self._refresh_gpt_client()
            reply = self.gpt_tester.chat(message=message, context_bundle=context_bundle)
        except Exception as exc:
            reply = f"Chat failed: {exc}"
        self._append_chat_transcript("Assistant", reply)

    def send_chat_message(self):
        msg = self.chat_message_widget.get("1.0", tk.END).strip() if self.chat_message_widget else ""
        if not msg:
            messagebox.showinfo("Missing", "Type a chat message first.")
            return
        if self.chat_auto_bundle_var.get():
            self.refresh_chat_preview()
        bundle = self.chat_bundle_widget.get("1.0", tk.END).strip() if self.chat_bundle_widget else ""
        self._send_to_chat(msg, bundle, "You")
        if self.chat_message_widget is not None:
            self.chat_message_widget.delete("1.0", tk.END)

    def send_bundle_message(self):
        if self.chat_auto_bundle_var.get():
            self.refresh_chat_preview()
        bundle = self.chat_bundle_widget.get("1.0", tk.END).strip() if self.chat_bundle_widget else ""
        if not bundle:
            messagebox.showinfo("Missing", "Build a bundle first.")
            return
        self._send_to_chat(
            "Please review the attached grading bundle and suggest fixes/improvements.",
            bundle,
            "You (bundle)",
        )

    def _build_settings_hub_tab(self):
        self.tab_settings.columnconfigure(0, weight=1)
        self.tab_settings.rowconfigure(0, weight=1)

        wrap = ttk.Frame(self.tab_settings, style="Pastel.TFrame")
        wrap.grid(row=0, column=0, sticky="nsew")
        wrap.columnconfigure(0, weight=1)
        wrap.rowconfigure(0, weight=1)
        wrap.rowconfigure(1, weight=1)
        wrap.rowconfigure(2, weight=1)
        wrap.rowconfigure(3, weight=0)

        include_box = ttk.Frame(wrap, style="PastelCard.TFrame", padding=10)
        include_box.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        ttk.Label(include_box, text="Chat bundle include settings", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Checkbutton(include_box, text="Include student code", variable=self.chat_include_code_var).pack(anchor="w", pady=(8, 0))
        ttk.Checkbutton(include_box, text="Include rubric scheme", variable=self.chat_include_scheme_var).pack(anchor="w")
        ttk.Checkbutton(include_box, text="Include prompt process + output", variable=self.chat_include_prompt_var).pack(anchor="w")
        ttk.Checkbutton(include_box, text="Compact code before bundling", variable=self.chat_compact_code_var).pack(anchor="w")

        chars_row = ttk.Frame(include_box, style="PastelCard.TFrame")
        chars_row.pack(anchor="w", pady=(8, 0))
        ttk.Label(chars_row, text="Code char limit", style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Entry(chars_row, textvariable=self.chat_code_char_limit_var, width=10).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Checkbutton(include_box, text="Auto-refresh bundle", variable=self.chat_auto_bundle_var).pack(anchor="w", pady=(8, 0))

        pdf_box = ttk.Frame(wrap, style="PastelCard.TFrame", padding=10)
        pdf_box.grid(row=1, column=0, sticky="nsew")
        ttk.Label(pdf_box, text="PDF output settings", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Checkbutton(pdf_box, text="Include student PDF", variable=self.pdf_menu_include_student_var).pack(anchor="w", pady=(8, 0))
        ttk.Checkbutton(pdf_box, text="Include summary PDF", variable=self.pdf_menu_include_summary_var).pack(anchor="w")
        ttk.Checkbutton(pdf_box, text="Include all student PDFs", variable=self.pdf_menu_include_batch_var).pack(anchor="w")
        ttk.Checkbutton(pdf_box, text="Student PDF: include scheme/grades table", variable=self.pdf_content_grades_var).pack(anchor="w", pady=(8, 0))
        ttk.Checkbutton(pdf_box, text="Student PDF: write rationale", variable=self.pdf_content_rationale_var).pack(anchor="w")
        ttk.Checkbutton(pdf_box, text="Student PDF: write code highlights", variable=self.pdf_content_highlights_var).pack(anchor="w")
        ttk.Checkbutton(pdf_box, text="Student PDF: write full code listing", variable=self.pdf_content_full_code_var).pack(anchor="w")
        tag_row = ttk.Frame(pdf_box, style="PastelCard.TFrame")
        tag_row.pack(anchor="w", pady=(8, 0))
        ttk.Label(tag_row, text="Batch tag", style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Entry(tag_row, textvariable=self.pdf_menu_batch_tag_var, width=24).pack(side=tk.LEFT, padx=(8, 0))

        auto_box = ttk.Frame(wrap, style="PastelCard.TFrame", padding=10)
        auto_box.grid(row=2, column=0, sticky="nsew", pady=(8, 0))
        auto_box.columnconfigure(0, weight=1)
        auto_box.rowconfigure(1, weight=1)
        ttk.Label(auto_box, text="Grading preflight settings", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.theme_settings_text = tk.Text(auto_box, height=5, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.theme_settings_text.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        self.theme_settings_text.insert("1.0", DEFAULT_THEME)

        leniency_row = ttk.Frame(auto_box, style="PastelCard.TFrame")
        leniency_row.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        ttk.Label(leniency_row, text="Leniency", style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Scale(leniency_row, from_=-1.0, to=1.0, variable=self.leniency_level_var, command=lambda _v: self._on_leniency_change()).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))
        ttk.Label(auto_box, textvariable=self.leniency_label_var, style="PastelCard.TLabel").grid(row=3, column=0, sticky="w", pady=(6, 0))

        preflight_btns = ttk.Frame(auto_box, style="PastelCard.TFrame")
        preflight_btns.grid(row=4, column=0, sticky="w", pady=(8, 0))
        ttk.Button(preflight_btns, text="Apply to Grade tab", command=lambda: self._current_theme_instructions(source="settings")).pack(side=tk.LEFT)
        ttk.Button(preflight_btns, text="Save theme + leniency", command=lambda: self.save_theme(source="settings")).pack(side=tk.LEFT, padx=6)

        btns = ttk.Frame(wrap, style="Pastel.TFrame")
        btns.grid(row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Button(btns, text="Save UI Settings", command=self.save_ui_preferences).pack(side=tk.LEFT)
        ttk.Button(btns, text="Reload UI Settings", command=self.load_ui_preferences).pack(side=tk.LEFT, padx=6)

    def save_ui_preferences(self):
        if not self.require_grading_db():
            return
        prefs = {
            "chat_auto_bundle": bool(self.chat_auto_bundle_var.get()),
            "chat_include_code": bool(self.chat_include_code_var.get()),
            "chat_include_scheme": bool(self.chat_include_scheme_var.get()),
            "chat_include_prompt": bool(self.chat_include_prompt_var.get()),
            "chat_compact_code": bool(self.chat_compact_code_var.get()),
            "chat_code_char_limit": int(self.chat_code_char_limit_var.get() or 7000),
            "pdf_include_student": bool(self.pdf_menu_include_student_var.get()),
            "pdf_include_summary": bool(self.pdf_menu_include_summary_var.get()),
            "pdf_include_batch": bool(self.pdf_menu_include_batch_var.get()),
            "pdf_batch_tag": self.pdf_menu_batch_tag_var.get().strip() or "Midterm",
            "pdf_content_grades": bool(self.pdf_content_grades_var.get()),
            "pdf_content_rationale": bool(self.pdf_content_rationale_var.get()),
            "pdf_content_highlights": bool(self.pdf_content_highlights_var.get()),
            "pdf_content_full_code": bool(self.pdf_content_full_code_var.get()),
            "leniency_level": float(self.leniency_level_var.get()),
        }
        meta_set(self.grade_con, "ui_preferences", json.dumps(prefs))
        self.refresh_chat_preview()
        messagebox.showinfo("Saved", "UI settings saved.")

    def load_ui_preferences(self):
        if not self.require_grading_db():
            return
        raw = meta_get(self.grade_con, "ui_preferences", "")
        if not raw:
            return
        try:
            prefs = json.loads(raw)
        except Exception:
            return

        self.chat_auto_bundle_var.set(bool(prefs.get("chat_auto_bundle", True)))
        self.chat_include_code_var.set(bool(prefs.get("chat_include_code", True)))
        self.chat_include_scheme_var.set(bool(prefs.get("chat_include_scheme", True)))
        self.chat_include_prompt_var.set(bool(prefs.get("chat_include_prompt", True)))
        self.chat_compact_code_var.set(bool(prefs.get("chat_compact_code", True)))
        try:
            self.chat_code_char_limit_var.set(max(1000, int(prefs.get("chat_code_char_limit", 7000))))
        except Exception:
            self.chat_code_char_limit_var.set(7000)
        self.pdf_menu_include_student_var.set(bool(prefs.get("pdf_include_student", True)))
        self.pdf_menu_include_summary_var.set(bool(prefs.get("pdf_include_summary", True)))
        self.pdf_menu_include_batch_var.set(bool(prefs.get("pdf_include_batch", False)))
        self.pdf_menu_batch_tag_var.set((prefs.get("pdf_batch_tag") or "Midterm").strip() or "Midterm")
        self.pdf_content_grades_var.set(bool(prefs.get("pdf_content_grades", True)))
        self.pdf_content_rationale_var.set(bool(prefs.get("pdf_content_rationale", True)))
        self.pdf_content_highlights_var.set(bool(prefs.get("pdf_content_highlights", True)))
        self.pdf_content_full_code_var.set(bool(prefs.get("pdf_content_full_code", True)))
        try:
            self.leniency_level_var.set(max(-1.0, min(1.0, float(prefs.get("leniency_level", 0.0)))))
        except Exception:
            self.leniency_level_var.set(0.0)
        self._on_leniency_change()
        self.refresh_chat_preview()

    def clear_chat_transcript(self):
        if self.chat_transcript_widget is not None:
            self.chat_transcript_widget.delete("1.0", tk.END)

    def clear_selected_student_grade(self):
        if not self.require_grading_db():
            return
        sid = self.selected_student_id
        if not sid:
            messagebox.showinfo("Missing", "Select a student first.")
            return
        if not messagebox.askyesno("Clear grade", f"Clear all saved grades, rationale, comments, and progress for {sid}?"):
            return

        with self.grade_con:
            self.grade_con.execute("DELETE FROM rubric_scores WHERE student_id=?", (sid,))
            self.grade_con.execute("DELETE FROM student_notes WHERE student_id=?", (sid,))
            self.grade_con.execute("DELETE FROM code_comments WHERE student_id=?", (sid,))
            self.grade_con.execute("DELETE FROM grading_progress WHERE student_id=?", (sid,))

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_chat_preview()
        self.refresh_progress_tab()
        self.refresh_prompt_trace_tab()
        self.refresh_chat_preview()
        messagebox.showinfo("Cleared", f"Grade data cleared for {sid}.")

    def _capture_auto_grade_trace(self, qid: str):
        prompt_payload = getattr(self.gpt_tester, "last_prompt_payload", None)
        request_body = getattr(self.gpt_tester, "last_request_body", None)
        result_payload = getattr(self.gpt_tester, "last_result", None)
        source = "remote" if request_body else "heuristic"

        prompt_lines = [
            f"Question: {qid}",
            f"Source: {source}",
            "",
            "Prompt payload:",
            json.dumps(prompt_payload, indent=2, ensure_ascii=False) if prompt_payload else "(none)",
        ]
        if request_body:
            prompt_lines.extend(["", "Raw request body:", request_body])

        self.last_auto_grade_trace = {
            "source": source,
            "prompt": "\n".join(prompt_lines),
            "result": json.dumps(result_payload, indent=2, ensure_ascii=False) if result_payload else "(none)",
        }
        self.refresh_prompt_trace_tab()
        self.refresh_chat_preview()

    def _build_summary_tab(self):
        top = ttk.Frame(self.tab_summary, style="Pastel.TFrame")
        top.pack(fill=tk.X)

        ttk.Button(top, text="Refresh Summary", command=self.refresh_summary).pack(side=tk.LEFT)
        ttk.Button(top, text="Export Grade (selected student)", command=self.export_selected_excel).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Export This Student PDF", command=self.export_student_pdf).pack(side=tk.LEFT, padx=6)

        self.summary_stats_lbl = ttk.Label(top, text="", style="Pastel.TLabel")
        self.summary_stats_lbl.pack(side=tk.RIGHT)

        cols = ("student_id", "student_name", "lab_id", "graded_questions", "overall", "overall_curved")
        self.sum_tree = ttk.Treeview(self.tab_summary, columns=cols, show="headings", height=25)
        for c, w in [("student_id", 140), ("student_name", 220), ("lab_id", 80),
                     ("graded_questions", 130), ("overall", 90), ("overall_curved", 120)]:
            self.sum_tree.heading(c, text=c)
            self.sum_tree.column(c, width=w, anchor="w")
        self.sum_tree.pack(fill=tk.BOTH, expand=True, pady=(10,0))

        note = ttk.Label(
            self.tab_summary,
            text="Summary includes only students marked as included. The table shows rubric points, per-question totals, and overall totals.",
            style="Pastel.TLabel"
        )
        note.pack(anchor="w", pady=(8,0))


    def _build_pdf_menu_tab(self):
        top = ttk.Frame(self.tab_pdf_menu, style="Pastel.TFrame")
        top.pack(fill=tk.X)
        ttk.Label(top, text="PDF export menu", style="Pastel.TLabel", font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)

        opts = ttk.Frame(self.tab_pdf_menu, style="PastelCard.TFrame", padding=10)
        opts.pack(fill=tk.X, pady=(10, 8))
        ttk.Checkbutton(opts, text="Include student PDF", variable=self.pdf_menu_include_student_var).pack(anchor="w")
        ttk.Checkbutton(opts, text="Include summary PDF", variable=self.pdf_menu_include_summary_var).pack(anchor="w")
        ttk.Checkbutton(opts, text="Include batch PDFs", variable=self.pdf_menu_include_batch_var).pack(anchor="w")

        content = ttk.LabelFrame(self.tab_pdf_menu, text="Student PDF content", padding=10)
        content.pack(fill=tk.X, pady=(0, 8))
        ttk.Checkbutton(content, text="Scheme table (rubric rows + points + notes)", variable=self.pdf_content_grades_var).pack(anchor="w")
        ttk.Checkbutton(content, text="Rationale text", variable=self.pdf_content_rationale_var).pack(anchor="w")
        ttk.Checkbutton(content, text="Code comment highlights", variable=self.pdf_content_highlights_var).pack(anchor="w")
        ttk.Checkbutton(content, text="Full code listing", variable=self.pdf_content_full_code_var).pack(anchor="w")

        tag_row = ttk.Frame(opts, style="PastelCard.TFrame")
        tag_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(tag_row, text="Batch report tag", style="PastelCard.TLabel").pack(side=tk.LEFT)
        ttk.Entry(tag_row, textvariable=self.pdf_menu_batch_tag_var, width=20).pack(side=tk.LEFT, padx=(8, 0))

        btns = ttk.Frame(self.tab_pdf_menu, style="Pastel.TFrame")
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="Run selected exports", command=self.run_pdf_menu_exports).pack(side=tk.LEFT)
        ttk.Button(btns, text="Export Student PDF", command=self.export_student_pdf).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="Export Summary PDF", command=self.export_summary_pdf).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="Export All Student PDFs", command=self.export_all_students_pdfs).pack(side=tk.LEFT, padx=6)

        ttk.Label(
            self.tab_pdf_menu,
            text="Select exactly what will be written in Student PDFs. Unticked items will not be written.",
            style="Pastel.TLabel"
        ).pack(anchor="w", pady=(8, 0))

    def _student_pdf_options(self) -> dict:
        return {
            "include_grade_tables": bool(self.pdf_content_grades_var.get()),
            "include_rationale": bool(self.pdf_content_rationale_var.get()),
            "include_comment_highlights": bool(self.pdf_content_highlights_var.get()),
            "include_full_code_listing": bool(self.pdf_content_full_code_var.get()),
        }

    def run_pdf_menu_exports(self):
        ran_any = False
        if self.pdf_menu_include_student_var.get():
            ran_any = True
            self.export_student_pdf()
        if self.pdf_menu_include_summary_var.get():
            ran_any = True
            self.export_summary_pdf()
        if self.pdf_menu_include_batch_var.get():
            ran_any = True
            self.export_all_students_pdfs(report_tag_override=self.pdf_menu_batch_tag_var.get().strip())
        if not ran_any:
            messagebox.showinfo("PDF Menu", "Tick at least one PDF export option.")

    def _build_progress_tab(self):
        top = ttk.Frame(self.tab_progress, style="Pastel.TFrame")
        top.pack(fill=tk.X)
        ttk.Button(top, text="Refresh Progress", command=self.refresh_progress_tab).pack(side=tk.LEFT)
        ttk.Button(top, text="Mark selected graded", command=lambda: self.set_selected_student_graded(True)).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Mark selected ungraded", command=lambda: self.set_selected_student_graded(False)).pack(side=tk.LEFT)
        ttk.Button(top, text="Mark selected reviewed", command=lambda: self.mark_selected_student_reviewed(True)).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Mark selected unreviewed", command=lambda: self.mark_selected_student_reviewed(False)).pack(side=tk.LEFT)
        ttk.Button(top, text="Clear all reviewed", command=self.clear_all_reviewed_flags).pack(side=tk.LEFT, padx=(10, 0))

        self.progress_header_lbl = ttk.Label(top, text="", style="Pastel.TLabel")
        self.progress_header_lbl.pack(side=tk.RIGHT)

        controls = ttk.Frame(self.tab_progress, style="Pastel.TFrame")
        controls.pack(fill=tk.X, pady=(8, 0))
        self.session_clock_lbl = ttk.Label(controls, text="Session: 00:00", style="Pastel.TLabel")
        self.session_clock_lbl.pack(side=tk.LEFT)
        ttk.Button(controls, text="Start Timer", command=self.start_session_timer).pack(side=tk.LEFT, padx=(8, 4))
        ttk.Button(controls, text="Pause Timer", command=self.pause_session_timer).pack(side=tk.LEFT, padx=4)
        ttk.Button(controls, text="Reset Timer", command=self.reset_session_timer).pack(side=tk.LEFT, padx=4)
        ttk.Button(controls, text="Mark Assessed", command=self.mark_current_student_assessed).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(controls, textvariable=self.progress_var, style="Pastel.TLabel").pack(side=tk.RIGHT)

        cols = ("student_id", "student_name", "graded", "reviewed", "scored_q", "rationale_q", "first_graded", "last_updated", "last_question")
        self.progress_tree = ttk.Treeview(self.tab_progress, columns=cols, show="headings", height=25)
        for c, w in [("student_id",120),("student_name",180),("graded",80),("reviewed",90),("scored_q",90),("rationale_q",100),("first_graded",160),("last_updated",160),("last_question",120)]:
            self.progress_tree.heading(c, text=c)
            self.progress_tree.column(c, width=w, anchor="w")
        self.progress_tree.pack(fill=tk.BOTH, expand=True, pady=(10,0))

    def _build_stats_tab(self):
        ttk.Label(self.tab_stats, text="Class Histogram (raw + curved preview)", style="Pastel.TLabel",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w")

        self.hist_frame = ttk.Frame(self.tab_stats, style="Pastel.TFrame")
        self.hist_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        if FigureCanvasTkAgg is None or Figure is None:
            ttk.Label(self.hist_frame, text="Matplotlib not available. Install: pip install matplotlib",
                      style="Pastel.TLabel").pack(anchor="w")
            self.hist_canvas = None
            self.hist_fig = None
            self.hist_ax = None
            return

        self.hist_fig = Figure(figsize=(8, 4.2), dpi=110)
        self.hist_ax = self.hist_fig.add_subplot(111)
        self.hist_canvas = FigureCanvasTkAgg(self.hist_fig, master=self.hist_frame)
        self.hist_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        ttk.Label(self.tab_stats,
                  text="Note: Curve preview uses multiplication factor (clipped at 0). It does not change stored grades unless you export/transform them externally.",
                  style="Pastel.TLabel").pack(anchor="w", pady=(8, 0))

    def _build_db_tab(self):
        self.tab_db.columnconfigure(1, weight=1)
        self.tab_db.rowconfigure(1, weight=1)

        top = ttk.Frame(self.tab_db, style="Pastel.TFrame")
        top.grid(row=0, column=0, columnspan=2, sticky="ew")
        ttk.Button(top, text="Refresh Tables", command=self.refresh_db_tables).pack(side=tk.LEFT)
        ttk.Button(top, text="Preview 200 rows", command=self.preview_selected_table).pack(side=tk.LEFT, padx=6)
        ttk.Label(top, textvariable=self.db_rows_status_var, style="Pastel.TLabel").pack(side=tk.RIGHT)

        left = ttk.Frame(self.tab_db, style="PastelCard.TFrame", padding=8)
        left.grid(row=1, column=0, sticky="nsw", padx=(0, 8))
        ttk.Label(left, text="Tables", style="PastelCard.TLabel", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.db_table_listbox = tk.Listbox(left, width=36, bg=self.palette["panel"], fg=self.palette["text"], highlightthickness=0, selectbackground=self.palette["select"])
        self.db_table_listbox.pack(fill=tk.BOTH, expand=True, pady=(6, 0))
        self.db_table_listbox.bind("<<ListboxSelect>>", lambda _e: self.preview_selected_table())

        right = ttk.Frame(self.tab_db, style="PastelCard.TFrame", padding=8)
        right.grid(row=1, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        ttk.Label(right, text="Table schema", style="PastelCard.TLabel").grid(row=0, column=0, sticky="w")
        self.db_schema_text = tk.Text(right, height=7, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.db_schema_text.grid(row=0, column=0, sticky="ew", pady=(20, 8))

        rows_frame = ttk.Frame(right, style="PastelCard.TFrame")
        rows_frame.grid(row=1, column=0, sticky="nsew")
        rows_frame.rowconfigure(0, weight=1)
        rows_frame.columnconfigure(0, weight=1)

        self.db_rows_tree = ttk.Treeview(rows_frame, columns=(), show="headings")
        self.db_rows_tree.grid(row=0, column=0, sticky="nsew")
        ysb = ttk.Scrollbar(rows_frame, orient="vertical", command=self.db_rows_tree.yview)
        ysb.grid(row=0, column=1, sticky="ns")
        xsb = ttk.Scrollbar(rows_frame, orient="horizontal", command=self.db_rows_tree.xview)
        xsb.grid(row=1, column=0, sticky="ew")
        self.db_rows_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        self.refresh_db_tables()

    def refresh_db_tables(self):
        if self.db_table_listbox is None:
            return
        self.db_table_listbox.delete(0, tk.END)
        rows = self.sub_con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
        for (name,) in rows:
            self.db_table_listbox.insert(tk.END, name)
        if rows:
            self.db_table_listbox.selection_set(0)
            self.preview_selected_table()

    def _selected_db_table(self) -> str:
        if self.db_table_listbox is None:
            return ""
        sel = self.db_table_listbox.curselection()
        if not sel:
            return ""
        return self.db_table_listbox.get(sel[0])

    def preview_selected_table(self):
        table = self._selected_db_table()
        if not table:
            self.db_rows_status_var.set("No table selected.")
            return

        safe_table = '"' + table.replace('"', '""') + '"'
        cols = self.sub_con.execute(f"PRAGMA table_info({safe_table})").fetchall()
        col_names = [r[1] for r in cols]
        schema_lines = [f"{r[1]} {r[2]}{' PRIMARY KEY' if r[5] else ''}" for r in cols]
        if self.db_schema_text is not None:
            self.db_schema_text.delete("1.0", tk.END)
            self.db_schema_text.insert("1.0", "\n".join(schema_lines) or "(no schema)")

        rows = self.sub_con.execute(f"SELECT * FROM {safe_table} LIMIT 200").fetchall()

        if self.db_rows_tree is None:
            return
        self.db_rows_tree.delete(*self.db_rows_tree.get_children())
        self.db_rows_tree["columns"] = col_names
        for c in col_names:
            self.db_rows_tree.heading(c, text=c)
            self.db_rows_tree.column(c, width=140, anchor="w")
        for row in rows:
            vals = ["" if v is None else str(v) for v in row]
            self.db_rows_tree.insert("", tk.END, values=vals)

        total = self.sub_con.execute(f"SELECT COUNT(*) FROM {safe_table}").fetchone()[0]
        self.db_rows_status_var.set(f"Table: {table} | showing {len(rows)} of {total} row(s)")

    def _default_regex_payload(self) -> dict:
        return {
            "file_globs": "*.java",
            "filename_regex": "",
            "exclude_filename_regex": "",
            "filename_regex_ignore_case": True,
            "folder_id_regex": "",
            "folder_name_regex": "",
            "only_new_files": False,
            "student_id_regex": r"\b\d{5,12}\b",
            "student_name_regex": r"",
        }

    def _ensure_default_regex_profile(self):
        if not list_regex_profiles(self.sub_con):
            upsert_regex_profile(self.sub_con, "Default", self._default_regex_payload())
        names = list_regex_profiles(self.sub_con)
        self.regex_profile_combo["values"] = names
        active = sub_meta_get(self.sub_con, "active_regex_profile", "Default") or "Default"
        if active not in names:
            active = names[0]
        self.active_regex_profile_var.set(active)
        self.regex_profile_pick_var.set(active)
        self.load_regex_profile_into_editor(active)

    def get_active_regex_payload(self) -> dict:
        name = self.active_regex_profile_var.get().strip() or "Default"
        payload = load_regex_profile(self.sub_con, name)
        return payload or self._default_regex_payload()

    def _build_regex_tab(self):
        top = ttk.Frame(self.tab_regex, style="Pastel.TFrame")
        top.pack(fill=tk.X)
        self.regex_profile_pick_var = tk.StringVar(value="Default")
        ttk.Label(top, text="Load Profile", style="Pastel.TLabel").pack(side=tk.LEFT)
        self.regex_profile_combo = ttk.Combobox(top, textvariable=self.regex_profile_pick_var, state="readonly", width=36)
        self.regex_profile_combo.pack(side=tk.LEFT, padx=(6, 10))
        self.regex_profile_combo.bind("<<ComboboxSelected>>", lambda _e: self.load_regex_profile_into_editor(self.regex_profile_pick_var.get()))
        ttk.Button(top, text="Save Profile", command=self.save_regex_profile).pack(side=tk.LEFT)
        ttk.Button(top, text="Save Copy As...", command=self.save_regex_profile_copy_as).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Commit", command=self.commit_regex_profile).pack(side=tk.LEFT)
        ttk.Button(top, text="Open Scan / Rescan", command=self.open_scan_window).pack(side=tk.LEFT, padx=(10, 0))

        form = ttk.Frame(self.tab_regex, style="Pastel.TFrame")
        form.pack(fill=tk.X, pady=(10, 0))
        self.regex_vars = {
            "student_id_regex": tk.StringVar(value=r"\b\d{5,12}\b"),
            "student_name_regex": tk.StringVar(value=""),
            "folder_id_regex": tk.StringVar(value=""),
            "folder_name_regex": tk.StringVar(value=""),
            "filename_regex": tk.StringVar(value=""),
            "exclude_filename_regex": tk.StringVar(value=""),
            "filename_regex_ignore_case": tk.StringVar(value="1"),
            "file_globs": tk.StringVar(value="*.java"),
        }
        fields = [
            ("Student ID extraction regex", "student_id_regex"),
            ("Student name extraction regex", "student_name_regex"),
            ("Optional folder ID regex", "folder_id_regex"),
            ("Optional folder name regex", "folder_name_regex"),
            ("Filename include regex", "filename_regex"),
            ("Filename exclude regex", "exclude_filename_regex"),
            ("Filename regex ignore case (1/0)", "filename_regex_ignore_case"),
            ("Filename glob patterns (comma/space/newline)", "file_globs"),
        ]
        for i, (label, key) in enumerate(fields):
            ttk.Label(form, text=label, style="Pastel.TLabel").grid(row=i, column=0, sticky="w", pady=4)
            ttk.Entry(form, textvariable=self.regex_vars[key], width=80).grid(row=i, column=1, sticky="ew", padx=(8, 0), pady=4)
        form.columnconfigure(1, weight=1)

    def _current_regex_payload_from_editor(self) -> dict:
        return {
            "student_id_regex": self.regex_vars["student_id_regex"].get().strip(),
            "student_name_regex": self.regex_vars["student_name_regex"].get().strip(),
            "folder_id_regex": self.regex_vars["folder_id_regex"].get().strip(),
            "folder_name_regex": self.regex_vars["folder_name_regex"].get().strip(),
            "filename_regex": self.regex_vars["filename_regex"].get().strip(),
            "exclude_filename_regex": self.regex_vars["exclude_filename_regex"].get().strip(),
            "filename_regex_ignore_case": self.regex_vars["filename_regex_ignore_case"].get().strip() != "0",
            "file_globs": self.regex_vars["file_globs"].get().strip() or "*.java",
            "only_new_files": False,
        }

    def load_regex_profile_into_editor(self, profile_name: str):
        payload = load_regex_profile(self.sub_con, profile_name) or self._default_regex_payload()
        for k, v in self.regex_vars.items():
            if k == "filename_regex_ignore_case":
                v.set("1" if bool(payload.get(k, True)) else "0")
            else:
                v.set(str(payload.get(k, "")))
        self.regex_profile_pick_var.set(profile_name)

    def save_regex_profile(self):
        name = self.regex_profile_pick_var.get().strip()
        if not name:
            return
        upsert_regex_profile(self.sub_con, name, self._current_regex_payload_from_editor())
        self._ensure_default_regex_profile()

    def save_regex_profile_copy_as(self):
        name = simpledialog.askstring("Save Profile Copy", "New profile name:")
        if not name:
            return
        upsert_regex_profile(self.sub_con, name.strip(), self._current_regex_payload_from_editor())
        self._ensure_default_regex_profile()
        self.regex_profile_pick_var.set(name.strip())

    def commit_regex_profile(self):
        name = (self.regex_profile_pick_var.get() or "Default").strip() or "Default"
        upsert_regex_profile(self.sub_con, name, self._current_regex_payload_from_editor())
        self.active_regex_profile_var.set(name)
        sub_meta_set(self.sub_con, "active_regex_profile", name)
        sub_meta_set(self.sub_con, "active_regex_payload", json.dumps(self._current_regex_payload_from_editor()))
        messagebox.showinfo("Committed", f"Committed regex profile for scans: {name}")

    def _build_settings_tab(self, parent=None):
        container = parent if parent is not None else self.tab_ai_trace
        box = ttk.Frame(container, style="PastelCard.TFrame", padding=12)
        box.pack(fill=tk.BOTH, expand=True)
        box.columnconfigure(1, weight=1)

        ttk.Label(box, text="GPT Grading Settings", style="PastelCard.TLabel", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(box, text="Enable API key for Auto Grade + Chat", style="PastelCard.TLabel").grid(row=1, column=0, sticky="w", pady=(10, 4))
        ttk.Checkbutton(box, variable=self.gpt_remote_enabled_var, text="Use online model", style="PastelCard.TCheckbutton").grid(row=1, column=1, sticky="w", padx=(8, 0), pady=(10, 4))

        ttk.Label(box, text="API Key", style="PastelCard.TLabel").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(box, textvariable=self.gpt_api_key_var, show="*", width=70).grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=4)

        ttk.Label(box, text="Model", style="PastelCard.TLabel").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Entry(box, textvariable=self.gpt_model_var, width=40).grid(row=3, column=1, sticky="w", padx=(8, 0), pady=4)

        ttk.Label(box, text="Prompt override", style="PastelCard.TLabel").grid(row=4, column=0, sticky="nw", pady=(8, 4))
        self.auto_prompt_text_widget = tk.Text(box, height=10, bg="#FFFDF7", fg=self.palette["text"], highlightthickness=1, highlightbackground="#E8E1FF")
        self.auto_prompt_text_widget.grid(row=4, column=1, sticky="nsew", padx=(8, 0), pady=(8, 4))
        box.rowconfigure(4, weight=1)

        btn_row = ttk.Frame(box, style="PastelCard.TFrame")
        btn_row.grid(row=5, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Button(btn_row, text="Save GPT Settings", command=self.save_gpt_settings).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="Reload", command=self.load_gpt_settings).pack(side=tk.LEFT, padx=6)

    def save_gpt_settings(self):
        if not self.require_grading_db():
            return
        prompt = self.auto_prompt_text_widget.get("1.0", tk.END).strip() if self.auto_prompt_text_widget else ""
        meta_set(self.grade_con, "gpt_api_key", self.gpt_api_key_var.get().strip())
        meta_set(self.grade_con, "gpt_model", self.gpt_model_var.get().strip() or "gpt-4.1-mini")
        meta_set(self.grade_con, "gpt_prompt", prompt)
        if self.gpt_api_key_var.get().strip() and not self.gpt_remote_enabled_var.get():
            self.gpt_remote_enabled_var.set(True)
        meta_set(self.grade_con, "gpt_remote_enabled", "1" if self.gpt_remote_enabled_var.get() else "0")
        self._refresh_gpt_client()
        messagebox.showinfo("Saved", "GPT settings saved.")

    def load_gpt_settings(self):
        if not self.require_grading_db():
            return
        self.gpt_api_key_var.set(meta_get(self.grade_con, "gpt_api_key", ""))
        self.gpt_model_var.set(meta_get(self.grade_con, "gpt_model", "gpt-4.1-mini"))
        self.gpt_remote_enabled_var.set((meta_get(self.grade_con, "gpt_remote_enabled", "0") or "0") == "1")
        prompt = meta_get(self.grade_con, "gpt_prompt", "")
        if self.auto_prompt_text_widget is not None:
            self.auto_prompt_text_widget.delete("1.0", tk.END)
            self.auto_prompt_text_widget.insert("1.0", prompt)
        self._refresh_gpt_client()

    def _refresh_gpt_client(self):
        key = self.gpt_api_key_var.get().strip() if self.gpt_remote_enabled_var.get() else ""
        prompt = self.auto_prompt_text_widget.get("1.0", tk.END).strip() if self.auto_prompt_text_widget else ""
        self.gpt_tester = GPT_test(api_key=key, model=self.gpt_model_var.get().strip() or "gpt-4.1-mini", system_prompt=prompt)
        self.auto_grader = AutoGrader(self.gpt_tester)

    def commit_scan_session_from_window(self, root_folder: str, global_lab_id: str, session_rows: list[dict]):
        name = self.active_regex_profile_var.get().strip() or "Default"
        payload = load_regex_profile(self.sub_con, name) or self._default_regex_payload()
        commit_scan_session(
            self.sub_con,
            root_folder=root_folder,
            lab_id=global_lab_id,
            profile_name=name,
            profile_payload=payload,
            session_payload={"rows": session_rows},
        )
    # ---- scan window ----
    def open_scan_window(self):
        return ScanWindow(self)

    # ---- submissions DB open/create/save ----
    def new_submissions_db(self):
        path = filedialog.asksaveasfilename(
            title="Create submissions DB",
            defaultextension=".sqlite",
            filetypes=[("SQLite DB", "*.sqlite"), ("All files", "*.*")]
        )
        if not path:
            return
        self._open_submissions_db(Path(path))

    def open_submissions_db(self):
        path = filedialog.askopenfilename(
            title="Open submissions DB",
            filetypes=[("SQLite DB", "*.sqlite"), ("All files", "*.*")]
        )
        if not path:
            return
        self._open_submissions_db(Path(path))

    def save_submissions_db_as(self):
        path = filedialog.asksaveasfilename(
            title="Save submissions DB as",
            defaultextension=".sqlite",
            filetypes=[("SQLite DB", "*.sqlite"), ("All files", "*.*")]
        )
        if not path:
            return

        target = Path(path)
        try:
            out_con = db_connect(target)
            self.sub_con.backup(out_con)
            out_con.close()
        except Exception as e:
            messagebox.showerror("Save failed", str(e))
            return

        self._open_submissions_db(target)
        messagebox.showinfo("Saved", f"Submissions DB saved to:\n{target}")

    def _open_submissions_db(self, path: Path):
        if self.sub_con is not None:
            try:
                self.sub_con.close()
            except Exception:
                pass

        self.sub_db_path = path
        self.sub_con = db_connect(path)
        submissions_db_init(self.sub_con)
        grading_db_init(self.sub_con)
        self.grade_con = self.sub_con
        self.grade_db_path = self.sub_db_path

        self.sub_db_status_lbl.config(text=f"Submissions DB: {path.name}")
        self.db_status_lbl.config(text=f"Grading: integrated with {path.name}")

        self._load_theme_and_leniency_from_db()
        self.load_gpt_settings()
        self.refresh_question_lists()

        self.selected_student_id = None
        self.selected_file_path = None
        self.refresh_students(keep_selected=False)
        self.refresh_summary()
        self.preview.delete("1.0", tk.END)
        self.file_list.delete(0, tk.END)
        self.comment_list.delete(0, tk.END)

    # ---- grading DB open/create ----
    def new_grading_db(self):
        messagebox.showinfo("Integrated DB", "Grading data is stored in the current submissions DB.")

    def open_grading_db(self):
        messagebox.showinfo("Integrated DB", "Grading data is stored in the current submissions DB.")

    def _open_grade_db(self, path: Path):
        messagebox.showinfo("Integrated DB", "Grading data is stored in the current submissions DB.")

    # ---- Scheme CSV loading ----
    def load_scheme_csv(self):
        if not self.require_grading_db():
            return
        path = filedialog.askopenfilename(
            title="Load scheme CSV",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            load_scheme_csv_into_db(self.grade_con, Path(path))
        except Exception as e:
            messagebox.showerror("CSV load failed", str(e))
            return

        self.refresh_question_lists()
        messagebox.showinfo("Loaded", f"Scheme loaded from:\n{path}")

    def reload_last_scheme(self):
        if not self.require_grading_db():
            return
        p = meta_get(self.grade_con, "scheme_csv_path", "")
        if not p or not Path(p).exists():
            messagebox.showinfo("No saved scheme", "No last scheme path saved (or file missing). Load Scheme CSV first.")
            return
        try:
            load_scheme_csv_into_db(self.grade_con, Path(p))
        except Exception as e:
            messagebox.showerror("Reload failed", str(e))
            return
        self.refresh_question_lists()
        messagebox.showinfo("Reloaded", f"Reloaded:\n{p}")

    def open_scheme_editor(self):
        if not self.require_grading_db():
            return
        win = tk.Toplevel(self.root)
        win.title("Scheme Generator / Editor")
        win.geometry("1280x820")

        top = ttk.Frame(win, style="Pastel.TFrame", padding=8)
        top.pack(fill=tk.X)

        body = ttk.Panedwindow(win, orient=tk.VERTICAL)
        body.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        text_wrap = ttk.Frame(body, style="Pastel.TFrame")
        preview_wrap = ttk.Frame(body, style="Pastel.TFrame")
        body.add(text_wrap, weight=3)
        body.add(preview_wrap, weight=2)

        text = tk.Text(text_wrap, wrap="none")
        text.pack(fill=tk.BOTH, expand=True)

        header = "question_id,question_title,group,col_key,col_text,col_max,col_order,sub_id\n"
        text.insert("1.0", header)

        preview_cols = ("question_id", "question_title", "group", "col_key", "col_text", "col_max", "col_order", "sub_id")
        ttk.Label(preview_wrap, text="Scheme table preview", style="Pastel.TLabel").pack(anchor="w", pady=(0, 4))
        tree = ttk.Treeview(preview_wrap, columns=preview_cols, show="headings", height=12)
        col_widths = {
            "question_id": 100,
            "question_title": 220,
            "group": 120,
            "col_key": 90,
            "col_text": 360,
            "col_max": 70,
            "col_order": 80,
            "sub_id": 90,
        }
        for c in preview_cols:
            tree.heading(c, text=c)
            tree.column(c, width=col_widths.get(c, 120), anchor="w")
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh_preview():
            for item in tree.get_children():
                tree.delete(item)
            raw = text.get("1.0", tk.END).strip()
            if not raw:
                return
            try:
                reader = csv.DictReader(io.StringIO(raw))
                for idx, row in enumerate(reader, start=1):
                    tree.insert("", "end", iid=f"scheme-{idx}", values=tuple((row.get(c, "") or "").strip() for c in preview_cols))
            except Exception:
                return

        def do_new():
            text.delete("1.0", tk.END)
            text.insert("1.0", header)
            refresh_preview()

        def do_load():
            path = filedialog.askopenfilename(title="Load scheme CSV", filetypes=[("CSV", "*.csv"), ("All files", "*.*")])
            if not path:
                return
            try:
                raw = Path(path).read_text(encoding="utf-8-sig", errors="ignore")
            except Exception as e:
                messagebox.showerror("Load failed", str(e))
                return
            text.delete("1.0", tk.END)
            text.insert("1.0", raw)
            refresh_preview()

        def do_save():
            path = filedialog.asksaveasfilename(title="Save scheme CSV", defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            try:
                Path(path).write_text(text.get("1.0", tk.END).strip() + "\n", encoding="utf-8")
            except Exception as e:
                messagebox.showerror("Save failed", str(e))
                return
            messagebox.showinfo("Saved", f"Saved scheme CSV:\n{path}")

        def do_set_current():
            raw = text.get("1.0", tk.END).strip()
            if not raw:
                messagebox.showinfo("Empty", "Enter scheme rows first.")
                return
            try:
                reader = csv.DictReader(io.StringIO(raw))
                if not reader.fieldnames:
                    raise ValueError("CSV has no header row.")
                required = {"question_id", "question_title", "group", "col_key", "col_text", "col_max", "col_order"}
                missing = required - set([(h or "").strip() for h in reader.fieldnames])
                if missing:
                    raise ValueError(f"CSV missing columns: {', '.join(sorted(missing))}")
                rows = list(reader)
                load_scheme_rows_into_db(self.grade_con, rows, source_label="(scheme-editor)")
            except Exception as e:
                messagebox.showerror("Set current failed", str(e))
                return
            self.refresh_question_lists()
            messagebox.showinfo("Current scheme", "Scheme set as current and loaded.")

        ttk.Button(top, text="New", command=do_new).pack(side=tk.LEFT)
        ttk.Button(top, text="Load CSV", command=do_load).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Save CSV", command=do_save).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Preview Table", command=refresh_preview).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Set as Current Scheme", command=do_set_current).pack(side=tk.LEFT, padx=6)

        text.bind("<KeyRelease>", lambda _e: refresh_preview())
        refresh_preview()

    def auto_export_pack(self):
        if not self.require_grading_db():
            return
        out_dir = filedialog.askdirectory(title="Choose output folder for auto export pack")
        if not out_dir:
            return
        report_tag = simpledialog.askstring("Report tag", "Enter report tag:", initialvalue="Midterm")
        if report_tag is None:
            return

        out_base = Path(out_dir)
        excel_path = out_base / "All_Grades.xlsx"
        summary_pdf = out_base / "Summary.pdf"
        stats_txt = out_base / "Class_Stats.txt"
        pdf_dir = out_base / "Student_PDFs"

        try:
            export_all_to_excel(self.sub_con, self.grade_con, excel_path)
            exporter = PDFExporter(self.sub_con, self.grade_con, self.question_map, self._student_pdf_options())
            exporter.export_summary_pdf(summary_pdf, self.compute_class_stats_text())
            exporter.export_all_students_pdfs(pdf_dir, report_tag=report_tag)
            stats_txt.write_text(self.compute_class_stats_text() + "\n", encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Auto export failed", str(e))
            return

        messagebox.showinfo(
            "Auto export complete",
            f"Saved:\n- {excel_path}\n- {summary_pdf}\n- {pdf_dir}\n- {stats_txt}",
        )

    def refresh_question_lists(self):
        if self.grade_con is None:
            return
        qs = fetch_questions(self.grade_con)
        self.question_map = {qid: (f"{title} [{sub_id}]" if sub_id else title) for qid, title, sub_id in qs}
        self.question_display_to_id = {
            f"{qid} - {self.question_map[qid]}": qid
            for qid in self.question_map.keys()
        }
        self.refresh_question_picker_for_student()

    def refresh_question_picker_for_student(self):
        if self.grade_con is None:
            return
        allowed = list(self.question_map.keys())
        if not allowed:
            self.selected_question_id = None
            self.question_pick_var.set("")
            return
        if not self.selected_question_id or self.selected_question_id not in self.question_map:
            self.selected_question_id = allowed[0]
        for label, qid in self.question_display_to_id.items():
            if qid == self.selected_question_id:
                self.question_pick_var.set(label)
                break

    def on_question_select(self, _evt=None):
        self.load_student_question_view()

    def on_question_picker_change(self, _evt=None):
        # Question picker removed from UI; kept for backward compatibility.
        self.load_student_question_view()

    def _build_full_rubric_rows(self):
        rows = []
        ui_map = {}
        max_map = {}
        for qid, qtitle in self.question_map.items():
            cols = fetch_columns_for_question(self.grade_con, qid)
            group_header = f"{qid} — {qtitle}"
            for col_key, group, text, mx in cols:
                ui_key = f"{qid}::{col_key}"
                label = f"[{qid}] {text}"
                display_group = group_header if not group else f"{group_header} / {group}"
                rows.append((ui_key, display_group, label, float(mx or 0.0)))
                ui_map[ui_key] = (qid, col_key)
                max_map[ui_key] = float(mx or 0.0)
        return rows, ui_map, max_map

    def schedule_auto_save(self):
        if self._suspend_auto_save:
            return
        if self._auto_save_job is not None:
            try:
                self.root.after_cancel(self._auto_save_job)
            except Exception:
                pass
        self._auto_save_job = self.root.after(500, self.auto_save_scores_silent)

    def auto_save_scores_silent(self):
        self._auto_save_job = None
        if self._suspend_auto_save:
            return
        if not self.selected_student_id:
            return
        self.save_scores_and_rationale(show_message=False)

    def _all_question_ids(self):
        return list(self.question_map.keys())

    def _merged_code_and_file_map(self, sid: str):
        file_map = {}
        merged_parts = []
        line_ranges = []
        current_line = 1
        for fp in get_student_files(self.sub_con, sid):
            content = get_file_content(self.sub_con, fp)
            if content is None:
                try:
                    content = read_file_text(Path(fp))
                except Exception:
                    content = ""
            file_map[fp] = content or ""
            header = f"// FILE: {Path(fp).name}"
            code_text = content or ""
            merged_parts.append(f"{header}\n{code_text}")

            header_lines = 1
            code_lines = len(code_text.splitlines()) if code_text else 0
            start_line = current_line + header_lines
            end_line = start_line + max(0, code_lines - 1)
            if code_lines > 0:
                line_ranges.append((fp, start_line, end_line))

            current_line += header_lines + code_lines + 2  # two spacer lines from "\n\n" join
        return "\n\n".join(merged_parts), file_map, line_ranges

    def _resolve_comment_targets(self, line_no: int, line_ranges: list[tuple[str, int, int]], file_map: dict[str, str]):
        for fp, start_line, end_line in line_ranges:
            if start_line <= line_no <= end_line:
                local_line = (line_no - start_line) + 1
                return [(fp, local_line)]

        # Fallback when model line number does not match merged-code offsets:
        # highlight that line in all files so both files get review markers.
        targets = []
        for fp, content in file_map.items():
            lines = (content or "").splitlines()
            if not lines:
                continue
            local_line = max(1, min(line_no, len(lines)))
            targets.append((fp, local_line))
        return targets

    def _line_to_index(self, content: str, line_no: int):
        lines = (content or "").splitlines()
        ln = max(1, min(line_no, len(lines) if lines else 1))
        col_end = len(lines[ln - 1]) if lines else 0
        return f"{ln}.0", f"{ln}.{col_end}"

    def auto_fill_student(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id:
            messagebox.showinfo("Missing", "Select a student first.")
            return
        qids = self._all_question_ids()
        if not qids:
            messagebox.showinfo("Missing", "Load a rubric scheme first.")
            return

        rationale = "AutoFill draft: assigned full marks by default. Please review and adjust as needed."
        with self.grade_con:
            for qid in qids:
                cols = fetch_columns_for_question(self.grade_con, qid)
                for col_key, _group, text, mx in cols:
                    pts = float(mx or 0.0)
                    upsert_score(self.grade_con, self.selected_student_id, qid, col_key, pts, f"AutoFill full credit for: {text}", commit=False)
                total_q = compute_total(self.grade_con, self.selected_student_id, qid)
                upsert_student_note(self.grade_con, self.selected_student_id, qid, rationale, overall_grade=total_q, commit=False)
            upsert_grading_progress(self.grade_con, self.selected_student_id, self.selected_question_id, mark_graded=True, reviewed=False, commit=False)

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_progress_tab()
        messagebox.showinfo("AutoFill", "AutoFill applied across all rubric rows. Mark reviewed after checking.")

    def _auto_grade_one_student(self, sid: str, qids: list[str], theme: str):
        self._refresh_gpt_client()
        merged_code, file_map, line_ranges = self._merged_code_and_file_map(sid)
        for qid in qids:
            cols = fetch_columns_for_question(self.grade_con, qid)
            rubric_items = [{"col_key": col_key, "group": (group or ""), "criterion": text, "min_points": 0.0, "max_points": float(mx)}
                            for col_key, group, text, mx in cols]
            res = self.auto_grader.auto_grade(
                question_id=qid,
                question_title=(self.question_map.get(qid, qid) or qid),
                merged_code=merged_code,
                rubric_items=rubric_items,
                theme_text=theme,
                leniency_level=float(self.leniency_level_var.get()),
            )
            self._capture_auto_grade_trace(qid)

            score_map = {x.get("col_key"): x.get("points", 0.0) for x in res.get("scores", []) if x.get("col_key")}
            note_map = {x.get("col_key"): x.get("note", "") for x in res.get("scores", []) if x.get("col_key")}
            max_map = {k: float(mx) for (k, _g, _t, mx) in cols}
            for col_key in max_map.keys():
                pts = float(score_map.get(col_key, 0.0))
                pts = clamp_points(pts, max_map[col_key])
                upsert_score(self.grade_con, sid, qid, col_key, pts, note_map.get(col_key, ""), commit=False)

            rationale = (res.get("rationale") or "").strip() or f"Draft feedback for {qid}."
            total_q = compute_total(self.grade_con, sid, qid)
            upsert_student_note(self.grade_con, sid, qid, rationale, overall_grade=total_q, commit=False)

            for c in (res.get("comments") or []):
                try:
                    line_no = int(c.get("line", 1))
                except Exception:
                    line_no = 1
                txt = (c.get("comment") or "").strip()
                if not is_mistake_focused_comment(txt):
                    continue
                for fp, local_line in self._resolve_comment_targets(line_no, line_ranges, file_map):
                    content = file_map.get(fp, "")
                    if not content:
                        continue
                    sidx, eidx = self._line_to_index(content, local_line)
                    add_code_comment(self.grade_con, sid, fp, sidx, eidx, f"[{qid}] {txt}", color="#FFE8A3")

        upsert_grading_progress(self.grade_con, sid, self.selected_question_id, mark_graded=True, reviewed=False, commit=False)

    def auto_grade_files_for_student(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id:
            messagebox.showinfo("Missing", "Select a student first.")
            return

        qids = self._all_question_ids()
        if not qids:
            messagebox.showinfo("Missing", "Load a rubric scheme first.")
            return

        theme = self._current_theme_instructions()
        try:
            with self.grade_con:
                self._auto_grade_one_student(self.selected_student_id, qids, theme)
        except Exception as e:
            messagebox.showerror("Auto grade failed", str(e))
            return

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_progress_tab()
        messagebox.showinfo("Grade Files", "Draft grading complete for this student. Please review.")

    def auto_grade_all_students(self):
        if not self.require_grading_db():
            return
        qids = self._all_question_ids()
        if not qids:
            messagebox.showinfo("Missing", "Load a rubric scheme first.")
            return
        rows = self.sub_con.execute("SELECT student_id FROM students WHERE include_in_summary=1 ORDER BY student_id").fetchall()
        student_ids = [r[0] for r in rows if r and r[0]]
        if not student_ids:
            messagebox.showinfo("Missing", "No included students found.")
            return

        theme = self._current_theme_instructions()
        failed = []
        with self.grade_con:
            for sid in student_ids:
                try:
                    self._auto_grade_one_student(sid, qids, theme)
                except Exception as exc:
                    failed.append(f"{sid}: {exc}")

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_progress_tab()
        if failed:
            messagebox.showwarning("Grade All", "Completed with some errors\n" + "\n".join(failed[:8]))
        else:
            messagebox.showinfo("Grade All", f"Draft grading complete for {len(student_ids)} students. Please review.")

    # ---- theme ----
    def save_theme(self, source: str = "auto"):
        if not self.require_grading_db():
            return
        theme = self._current_theme_instructions(source=source)
        meta_set(self.grade_con, "theme", theme)
        meta_set(self.grade_con, "leniency_level", f"{float(self.leniency_level_var.get()):.3f}")
        messagebox.showinfo("Saved", "Theme + leniency saved.")


    def reset_session_timer(self):
        self.pause_session_timer()
        self.session_started_at = None
        self.session_elapsed_seconds = 0
        self.session_clock_lbl.config(text="Session: 00:00")

    def start_session_timer(self):
        if self.session_timer_running:
            return
        self.session_started_at = datetime.now()
        self.session_timer_running = True
        self._tick_session_clock()

    def pause_session_timer(self):
        if not self.session_timer_running:
            return
        elapsed = datetime.now() - self.session_started_at
        self.session_elapsed_seconds += int(elapsed.total_seconds())
        self.session_timer_running = False
        if self._clock_job is not None:
            try:
                self.root.after_cancel(self._clock_job)
            except Exception:
                pass
            self._clock_job = None

    def _tick_session_clock(self):
        if not self.session_timer_running or self.session_started_at is None:
            return
        elapsed = datetime.now() - self.session_started_at
        secs = self.session_elapsed_seconds + int(elapsed.total_seconds())
        hh = secs // 3600
        mm = (secs % 3600) // 60
        ss = secs % 60
        fmt = f"{hh:02d}:{mm:02d}:{ss:02d}" if hh > 0 else f"{mm:02d}:{ss:02d}"
        self.session_clock_lbl.config(text=f"Session: {fmt}")
        if self._clock_job is not None:
            try:
                self.root.after_cancel(self._clock_job)
            except Exception:
                pass
        self._clock_job = self.root.after(1000, self._tick_session_clock)

    def mark_current_student_assessed(self):
        if not self.selected_student_id:
            messagebox.showinfo("Missing", "Select a student first.")
            return
        set_student_graded_flag(self.grade_con, self.selected_student_id, True)
        self.refresh_progress_tab()
        self.refresh_summary()

    def _fetch_progress_rows(self):
        if self.grade_con is None:
            return []
        rows = self.sub_con.execute("""
          SELECT s.student_id, s.student_name,
                 COALESCE(gp.graded,0),
                 COALESCE(gp.reviewed,1),
                 (SELECT COUNT(DISTINCT rs.question_id) FROM rubric_scores rs WHERE rs.student_id=s.student_id AND rs.points IS NOT NULL),
                 (SELECT COUNT(DISTINCT sn.question_id) FROM student_notes sn WHERE sn.student_id=s.student_id AND TRIM(COALESCE(sn.rationale,''))<>''),
                 COALESCE(gp.first_graded_at,''),
                 COALESCE(gp.last_updated_at,''),
                 COALESCE(gp.last_question_id,'')
          FROM students s
          LEFT JOIN grading_progress gp ON gp.student_id=s.student_id
          WHERE LOWER(s.student_id) <> 'full' AND COALESCE(s.included,1)=1
          ORDER BY s.student_id
        """).fetchall()
        return [r for r in rows if has_required_student_fields(r[0], r[1])]

    def _compute_progress_counts(self):
        rows = self._fetch_progress_rows()
        total = len(rows)
        assessed = 0
        reviewed_count = 0
        for _sid, _name, graded, reviewed, scored_q, rationale_q, _fg, _lu, _lq in rows:
            if int(graded or 0) == 1:
                assessed += 1
            if int(reviewed or 0) == 1:
                reviewed_count += 1
        return total, assessed, reviewed_count, max(0, total - assessed)

    def _update_student_progress_labels(self, total: int, assessed: int, reviewed_count: int, left: int):
        self.progress_var.set(f"Progress: assessed {assessed}/{total} | left {left}")
        curve_k = float(self.curve_preview_var.get())
        self.main_menu_progress_var.set(
            f"Students assessed: {assessed}/{total} | Left: {left} | Unreviewed: {max(0, assessed-reviewed_count)} | Curve preview ×{curve_k:.2f}"
        )

    def refresh_progress_tab(self):
        if not hasattr(self, "progress_tree"):
            return
        for item in self.progress_tree.get_children():
            self.progress_tree.delete(item)

        rows = self._fetch_progress_rows()
        for sid, sname, graded, reviewed, scored_q, rationale_q, first_g, last_u, last_q in rows:
            self.progress_tree.insert("", "end", iid=sid, values=(
                sid, sname, "YES" if int(graded or 0) == 1 else "NO", "YES" if int(reviewed or 0) == 1 else "NO", scored_q, rationale_q,
                first_g or "-", last_u or "-", last_q or "-"
            ))
        total, assessed, reviewed_count, left = self._compute_progress_counts()
        self.progress_header_lbl.config(text=f"Assessed {assessed}/{total} | Reviewed {reviewed_count} | Left {left}")
        self._update_student_progress_labels(total, assessed, reviewed_count, left)

    def mark_selected_student_reviewed(self, reviewed: bool = True):
        sid = self.selected_student_id
        if sid is None and hasattr(self, "progress_tree"):
            sel = self.progress_tree.selection()
            if sel:
                sid = sel[0]
        if not sid:
            messagebox.showinfo("Select", "Select a student first.")
            return
        set_student_reviewed_flag(self.grade_con, sid, reviewed)
        self.refresh_progress_tab()
        self.refresh_summary()
        if self.selected_student_id == sid:
            g, rv, fg, lu, lq = load_grading_progress(self.grade_con, sid)
            self.grade_meta_var.set(
                f"Graded: {'YES' if int(g or 0)==1 else 'NO'} | Reviewed: {'YES' if int(rv or 0)==1 else 'NO'} | First graded: {fg or '-'} | Last updated: {lu or '-'} | Last question: {lq or '-'}"
            )

    def set_selected_student_graded(self, graded: bool):
        if not hasattr(self, "progress_tree"):
            return
        sel = self.progress_tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a student row in Progress tab.")
            return
        sid = sel[0]
        set_student_graded_flag(self.grade_con, sid, graded)
        self.refresh_progress_tab()
        self.refresh_summary()
        if self.selected_student_id == sid:
            g, rv, fg, lu, lq = load_grading_progress(self.grade_con, sid)
            self.grade_meta_var.set(
                f"Graded: {'YES' if int(g or 0)==1 else 'NO'} | Reviewed: {'YES' if int(rv or 0)==1 else 'NO'} | First graded: {fg or '-'} | Last updated: {lu or '-'} | Last question: {lq or '-'}"
            )

    def clear_all_reviewed_flags(self):
        if self.grade_con is None:
            return
        rows = self._fetch_progress_rows()
        if not rows:
            messagebox.showinfo("Progress", "No students found in progress table.")
            return
        if not messagebox.askyesno("Clear all reviewed", "Mark every student as unreviewed?"):
            return

        with self.grade_con:
            for sid, *_rest in rows:
                set_student_reviewed_flag(self.grade_con, sid, False, commit=False)

        self.refresh_progress_tab()
        self.refresh_summary()
        if self.selected_student_id:
            graded, reviewed, first_g, last_u, last_q = load_grading_progress(self.grade_con, self.selected_student_id)
            self.grade_meta_var.set(
                f"Graded: {'YES' if int(graded or 0)==1 else 'NO'} | Reviewed: {'YES' if int(reviewed or 0)==1 else 'NO'} | First graded: {first_g or '-'} | Last updated: {last_u or '-'} | Last question: {last_q or '-'}"
            )
        messagebox.showinfo("Progress", "All students are now marked unreviewed.")

    # ---- students ----
    def refresh_students(self, keep_selected: bool = True):
        selected_id = self.selected_student_id if keep_selected else None

        rows = get_students(self.sub_con)
        self.student_list.delete(0, tk.END)
        self.student_ids = []

        for sid, name, lab, file_count in rows:
            if (sid or "").strip().lower() == "full":
                continue
            if not has_required_student_fields(sid, name):
                continue
            lab_txt = f" | Lab:{lab}" if lab else ""
            self.student_list.insert(tk.END, f"{sid} — {name}{lab_txt} | files: {file_count}")
            self.student_ids.append(sid)

        self.student_count_lbl.config(text=f"Students: {len(self.student_ids)}")

        if keep_selected and selected_id and selected_id in self.student_ids:
            idx = self.student_ids.index(selected_id)
            self.student_list.selection_clear(0, tk.END)
            self.student_list.selection_set(idx)
            self.student_list.activate(idx)
            self.student_list.see(idx)
        else:
            self.selected_student_id = None
            self.chat_student_var.set("Chat student: none selected")
            self.grade_meta_var.set("Graded: NO | Reviewed: YES | First graded: - | Last updated: -")

        self.refresh_progress_tab()
        self.refresh_chat_preview()

    def on_student_select(self, _evt=None):
        sel = self.student_list.curselection()
        if not sel:
            return
        sid = self.student_ids[sel[0]]
        self.selected_student_id = sid

        row = self.sub_con.execute("SELECT student_name, COALESCE(lab_id,'') FROM students WHERE student_id=?", (sid,)).fetchone()
        name = row[0] if row else ""
        lab = row[1] if row else ""
        lab_txt = f" | Lab:{lab}" if lab else ""
        self.student_header.config(text=f"{sid} — {name}{lab_txt}")
        self.chat_student_var.set(f"Chat student: {sid} — {name}" if name else f"Chat student: {sid}")
        graded, reviewed, first_g, last_u, last_q = load_grading_progress(self.grade_con, sid)
        self.grade_meta_var.set(
            f"Graded: {'YES' if int(graded or 0)==1 else 'NO'} | Reviewed: {'YES' if int(reviewed or 0)==1 else 'NO'} | First graded: {first_g or '-'} | Last updated: {last_u or '-'} | Last question: {last_q or '-'}"
        )

        files = get_student_files(self.sub_con, sid)
        self.file_list.delete(0, tk.END)
        for f in files:
            self.file_list.insert(tk.END, f)

        self.preview.delete("1.0", tk.END)
        self.selected_file_path = None
        self.comment_list.delete(0, tk.END)

        if self.grade_con is not None:
            self.refresh_question_picker_for_student()

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_chat_preview()

    def on_grade_file_select(self, _evt=None):
        sel = self.file_list.curselection()
        if not sel:
            return
        fp = self.file_list.get(sel[0])
        self.selected_file_path = fp

        content = get_file_content(self.sub_con, fp)
        if content is None:
            try:
                content = Path(fp).read_text(encoding="utf-8", errors="ignore")
            except Exception as e:
                content = f"Error reading file:\n{e}"

        self.preview.delete("1.0", tk.END)
        self.preview.insert("1.0", content)
        self._apply_comments_highlights()

    # ---- code comments ----
    def _apply_comments_highlights(self):
        self.preview.tag_remove("comment_highlight", "1.0", tk.END)
        self.comment_list.delete(0, tk.END)

        if self.grade_con is None:
            return
        if not self.selected_student_id or not self.selected_file_path:
            return

        comments = fetch_code_comments_for_file(self.grade_con, self.selected_student_id, self.selected_file_path)
        for cid, sidx, eidx, text, color, _created_at in comments:
            trimmed = self._trim_preview_range(sidx, eidx)
            if not trimmed:
                continue
            sidx, eidx = trimmed
            tag = "comment_highlight"
            if color:
                try:
                    self.preview.tag_configure(tag, background=color)
                except Exception:
                    pass
            try:
                self.preview.tag_add(tag, sidx, eidx)
            except Exception:
                pass
            self.comment_list.insert(tk.END, f"#{cid} {_format_comment_range_label(sidx, eidx)}: {text}")

    def _trim_preview_range(self, sidx: str, eidx: str):
        try:
            selected_text = self.preview.get(sidx, eidx)
        except Exception:
            return None
        if not selected_text:
            return None

        left_trim = len(selected_text) - len(selected_text.lstrip())
        right_trim = len(selected_text) - len(selected_text.rstrip())
        if left_trim + right_trim >= len(selected_text):
            return None

        trimmed_start = self.preview.index(f"{sidx}+{left_trim}c") if left_trim else sidx
        trimmed_end = self.preview.index(f"{eidx}-{right_trim}c") if right_trim else eidx
        return trimmed_start, trimmed_end

    def add_comment_to_selection(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id or not self.selected_file_path:
            messagebox.showinfo("Select", "Select a student and a file first.")
            return
        try:
            sidx = self.preview.index("sel.first")
            eidx = self.preview.index("sel.last")
        except Exception:
            messagebox.showinfo("Select", "Highlight a region of code first.")
            return

        trimmed = self._trim_preview_range(sidx, eidx)
        if not trimmed:
            messagebox.showinfo("Select", "Selection only contains whitespace. Highlight code text first.")
            return
        sidx, eidx = trimmed

        txt = simpledialog.askstring("Add comment", "Comment for highlighted code:")
        if not txt:
            return
        add_code_comment(self.grade_con, self.selected_student_id, self.selected_file_path, sidx, eidx, txt, color="#FFF9A6")
        self._apply_comments_highlights()

    def clear_comments_in_selection(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id or not self.selected_file_path:
            return
        try:
            sidx = self.preview.index("sel.first")
            eidx = self.preview.index("sel.last")
        except Exception:
            messagebox.showinfo("Select", "Highlight a region first (selection).")
            return
        delete_code_comments_in_range(self.grade_con, self.selected_student_id, self.selected_file_path, sidx, eidx)
        self._apply_comments_highlights()

    # ---- load rubric table for selected student (all questions) ----
    def load_student_question_view(self):
        self._suspend_auto_save = True
        self.rationale_text.delete("1.0", tk.END)
        self.total_lbl.config(text="Overall Total: -")
        self.rubric_grid.build([])

        if self.grade_con is None:
            self._suspend_auto_save = False
            return
        if not self.selected_student_id:
            self._suspend_auto_save = False
            return

        rows, self._rubric_ui_map, self._rubric_max_map = self._build_full_rubric_rows()
        self.rubric_grid.build(rows)

        combined_scores = {}
        combined_notes = {}
        rationale_blocks = []
        for qid in self.question_map.keys():
            score_map, note_map = load_student_scores(self.grade_con, self.selected_student_id, qid)
            for ui_key, (qid2, col_key) in self._rubric_ui_map.items():
                if qid2 == qid:
                    combined_scores[ui_key] = score_map.get(col_key)
                    combined_notes[ui_key] = note_map.get(col_key, "")

            note_row = load_student_note(self.grade_con, self.selected_student_id, qid)
            if note_row and (note_row[0] or "").strip():
                rationale_blocks.append(f"[{qid}]\n{note_row[0].strip()}")

        self.rubric_grid.set_values(combined_scores, combined_notes)
        if rationale_blocks:
            self.rationale_text.insert("1.0", "\n\n".join(rationale_blocks))

        total = compute_total(self.grade_con, self.selected_student_id, None)
        self.total_lbl.config(text=f"Overall Total: {total:g}")
        self._suspend_auto_save = False

    # ---- save ----
    def save_scores_and_rationale(self, show_message: bool = True):
        if not self.require_grading_db():
            return
        if not self.selected_student_id:
            messagebox.showinfo("Missing", "Select a student first.")
            return
        if not self._rubric_ui_map:
            messagebox.showinfo("Missing", "Load a rubric scheme first.")
            return

        score_raw, note_raw = self.rubric_grid.get_values()

        with self.grade_con:
            for ui_key, raw in score_raw.items():
                raw = (raw or "").strip()
                if raw == "":
                    points = None
                else:
                    try:
                        points = float(raw)
                    except ValueError:
                        messagebox.showerror("Invalid score", f"Score must be numeric or blank.\nBad value for:\n{ui_key}")
                        return
                    points = clamp_points(points, self._rubric_max_map.get(ui_key, points))

                qid, col_key = self._rubric_ui_map[ui_key]
                upsert_score(self.grade_con, self.selected_student_id, qid, col_key, points, note_raw.get(ui_key, ""), commit=False)

            rationale = self.rationale_text.get("1.0", tk.END).strip()
            total = compute_total(self.grade_con, self.selected_student_id, None)
            for qid in self.question_map.keys():
                upsert_student_note(self.grade_con, self.selected_student_id, qid, rationale, overall_grade=total, commit=False)
            upsert_grading_progress(self.grade_con, self.selected_student_id, self.selected_question_id, mark_graded=True, reviewed=True, commit=False)

        self.total_lbl.config(text=f"Overall Total: {total:g}")
        self.refresh_summary()
        self.refresh_progress_tab()
        if self.selected_student_id:
            graded, reviewed, first_g, last_u, last_q = load_grading_progress(self.grade_con, self.selected_student_id)
            self.grade_meta_var.set(
                f"Graded: {'YES' if int(graded or 0)==1 else 'NO'} | Reviewed: {'YES' if int(reviewed or 0)==1 else 'NO'} | First graded: {first_g or '-'} | Last updated: {last_u or '-'} | Last question: {last_q or '-'}"
            )
        if show_message:
            messagebox.showinfo("Saved", "Saved scores + rationale for all rubric questions.")


    # ---- Optional auto grade (separate component) ----
    def auto_grade_optional(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id or not self.selected_question_id:
            messagebox.showinfo("Missing", "Select a student and a question first.")
            return
        cols = fetch_columns_for_question(self.grade_con, self.selected_question_id)
        rubric_items = [{"col_key": col_key, "group": (group or ""), "criterion": text, "min_points": 0.0, "max_points": float(mx)}
                        for col_key, group, text, mx in cols]
        merged_code = merge_student_code(self.sub_con, self.selected_student_id)
        theme = self._current_theme_instructions()

        try:
            self._refresh_gpt_client()
            res = self.auto_grader.auto_grade(question_id=self.selected_question_id, question_title=(self.question_map.get(self.selected_question_id, self.selected_question_id) or self.selected_question_id), merged_code=merged_code, rubric_items=rubric_items, theme_text=theme, leniency_level=float(self.leniency_level_var.get()))
            self._capture_auto_grade_trace(self.selected_question_id)
        except Exception as e:
            messagebox.showerror("Auto grade failed", str(e))
            return

        # Apply
        score_map = {x["col_key"]: x.get("points", 0.0) for x in res.get("scores", []) if "col_key" in x}
        note_map = {x["col_key"]: x.get("note", "") for x in res.get("scores", []) if "col_key" in x}

        max_map = {k: float(mx) for (k, _g, _t, mx) in cols}
        with self.grade_con:
            for col_key in max_map.keys():
                pts = float(score_map.get(col_key, 0.0))
                pts = clamp_points(pts, max_map[col_key])
                upsert_score(self.grade_con, self.selected_student_id, self.selected_question_id, col_key, pts, note_map.get(col_key, ""), commit=False)

            rationale = (res.get("rationale") or "").strip()
            total = compute_total(self.grade_con, self.selected_student_id, self.selected_question_id)
            upsert_student_note(self.grade_con, self.selected_student_id, self.selected_question_id, rationale, overall_grade=total, commit=False)
            upsert_grading_progress(self.grade_con, self.selected_student_id, self.selected_question_id, mark_graded=True, reviewed=False, commit=False)

        self.load_student_question_view()
        self.refresh_summary()
        self.refresh_progress_tab()
        messagebox.showinfo("Auto grade", "Draft applied and saved. Review/edit if needed.")

    # ---- Autofill/Test: iterate all students × assigned questions (kept; now purely a loop) ----
    def make_all_assigned(self):
        if not self.require_grading_db():
            return
        if not self.question_map:
            messagebox.showinfo("No scheme", "Load Scheme CSV first.")
            return

        students = list(self.student_ids)
        if not students:
            messagebox.showinfo("No students", "Scan/Commit first so students exist.")
            return

        # Here: intentionally does NOT call any external/AI.
        # Placeholder pass that can be extended for validation checks.
        self.refresh_summary()
        messagebox.showinfo("Done", f"Checked {len(students)} students. Scheme-driven grading is active.")

    # ---- Export Excel (all) ----
    def save_all_excel(self):
        if not self.require_grading_db():
            return
        out = filedialog.asksaveasfilename(
            title="Export all grades (Excel)",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not out:
            return
        export_all_to_excel(self.sub_con, self.grade_con, Path(out))
        messagebox.showinfo("Exported", f"Saved:\n{out}")

    def export_selected_excel(self):
        if not self.require_grading_db():
            return
        if not self.selected_student_id:
            messagebox.showinfo("Select", "Select a student first.")
            return

        out = filedialog.asksaveasfilename(
            title="Export selected student's grades (Excel)",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not out:
            return

        export_all_to_excel(
            self.sub_con,
            self.grade_con,
            Path(out),
            student_filter={self.selected_student_id},
        )
        messagebox.showinfo("Exported", f"Saved selected student grade export:\n{out}")

    # ---- Summary + class stats + histogram ----
    def compute_class_values(self):
        """
        Returns list of overall raw totals for non-FULL students.
        """
        if self.grade_con is None:
            return []
        rows = self.sub_con.execute("""
          SELECT student_id, student_name FROM students
          WHERE LOWER(student_id) <> 'full' AND COALESCE(included,1)=1
        """).fetchall()
        vals = []
        display_qids = fetch_display_question_ids(self.grade_con)
        for sid, sname in rows:
            if not has_required_student_fields(sid, sname):
                continue
            vals.append(sum(compute_total_by_display_id(self.grade_con, sid, qid) for qid in display_qids))
        return vals

    def compute_class_stats_text(self):
        vals = sorted(self.compute_class_values())
        if not vals:
            return "Class Stats: (not available)"
        n = len(vals)
        avg = sum(vals) / n
        mn = vals[0]
        mx = vals[-1]
        med = vals[n//2] if n % 2 else (vals[n//2 - 1] + vals[n//2]) / 2.0
        q1 = vals[n//4]
        q3 = vals[(3*n)//4]
        variance = sum((v - avg) ** 2 for v in vals) / n
        std = math.sqrt(variance)
        overall_max = sum(compute_question_max(self.grade_con, qid) for qid in fetch_display_question_ids(self.grade_con)) if self.grade_con else 0.0
        zeros = sum(1 for v in vals if v <= 0.0001)
        perfects = sum(1 for v in vals if overall_max > 0 and abs(v - overall_max) < 1e-6)
        target_avg = 75.0
        curve_factor = (target_avg / avg) if avg > 0 else 1.0
        return (
            f"Class Stats: avg {avg:.2f} | median {med:.2f} | std {std:.2f} | min {mn:.2f} | max {mx:.2f} "
            f"| Q1 {q1:.2f} | Q3 {q3:.2f} | zeros {zeros} | perfect {perfects} | suggested curve× {curve_factor:.3f}"
        )

    def refresh_histogram(self):
        if self.hist_canvas is None or self.hist_ax is None:
            return
        vals = sorted(self.compute_class_values())
        self.hist_ax.clear()
        if not vals:
            self.hist_ax.set_title("No class data")
            self.hist_canvas.draw()
            return
        import numpy as np
        counts, bins = np.histogram(vals, bins=min(12, max(5, len(vals)//2)))
        centers = (bins[:-1] + bins[1:]) / 2
        self.hist_ax.plot(centers, counts, marker="o", label="distribution")
        if len(counts) >= 3:
            kernel = np.array([0.25, 0.5, 0.25])
            smoothed = np.convolve(counts, kernel, mode="same")
            self.hist_ax.plot(centers, smoothed, linestyle="--", label="smoothed")
        k = float(self.curve_preview_var.get())
        curved_vals = [max(0.0, v * k) for v in vals]
        c_counts, c_bins = np.histogram(curved_vals, bins=bins)
        c_centers = (c_bins[:-1] + c_bins[1:]) / 2
        self.hist_ax.plot(c_centers, c_counts, alpha=0.75, label=f"curve preview ×{k:.2f}")
        self.hist_ax.set_title("Score distribution (line curve)")
        self.hist_ax.set_xlabel("Overall total")
        self.hist_ax.set_ylabel("Count")
        self.hist_ax.legend()
        self.hist_canvas.draw()

    def refresh_summary(self):
        if self.grade_con is None:
            for item in self.sum_tree.get_children():
                self.sum_tree.delete(item)
            self.class_stats_lbl.config(text="Class: avg - | min - | max - | curve -")
            self.summary_stats_lbl.config(text="Assessed 0/0 | Left 0")
            self._update_student_progress_labels(0, 0, 0, 0)
            self.refresh_histogram()
            self.refresh_progress_tab()
            return

        for item in self.sum_tree.get_children():
            self.sum_tree.delete(item)

        question_ids = fetch_display_question_ids(self.grade_con)
        parts = fetch_rubric_parts(self.grade_con)
        q_display_map = build_question_display_map(self.grade_con)
        cols = ["student_id", "student_name", "lab_id"] + [f"{q_display_map.get(qid, qid)}:{ck}" for qid, ck, _g, _t, _m in parts] + [f"{qid}_total" for qid in question_ids] + ["overall", "overall_curved"]
        self.sum_tree.configure(columns=cols)
        for c in cols:
            self.sum_tree.heading(c, text=c)
            self.sum_tree.column(c, width=110 if c not in {"student_name"} else 180, anchor="w")

        students = self.sub_con.execute("""
          SELECT student_id, student_name, COALESCE(lab_id,'')
          FROM students
          WHERE LOWER(student_id) <> 'full' AND COALESCE(included,1)=1
          ORDER BY student_id
        """).fetchall()

        curve_k = float(self.curve_preview_var.get())
        for sid, sname, lab in students:
            if not has_required_student_fields(sid, sname):
                continue
            row = [sid, sname, lab]
            score_cache = {}
            for qid, ck, _g, _t, _m in parts:
                if qid not in score_cache:
                    score_cache[qid] = load_student_scores(self.grade_con, sid, qid)[0]
                v = score_cache[qid].get(ck)
                row.append("" if v is None else f"{v:g}")
            for qid in question_ids:
                row.append(f"{compute_total_by_display_id(self.grade_con, sid, qid):g}")
            overall = sum(compute_total_by_display_id(self.grade_con, sid, qid) for qid in question_ids)
            row.append(f"{overall:g}")
            row.append(f"{max(0.0, overall * curve_k):.2f}")
            self.sum_tree.insert("", "end", values=row)

        stats_text = self.compute_class_stats_text()
        self.class_stats_lbl.config(text=stats_text.replace("Class Stats:", "Class:"))
        total, assessed, reviewed_count, left = self._compute_progress_counts()
        self.summary_stats_lbl.config(text=f"Assessed {assessed}/{total} | Left {left}")
        self._update_student_progress_labels(total, assessed, reviewed_count, left)
        self.refresh_histogram()
        self.refresh_progress_tab()

    # ---- PDF Exports ----
    def export_student_pdf(self):
        if not self.require_grading_db():
            return
        if SimpleDocTemplate is None:
            messagebox.showinfo("PDF missing", "reportlab not installed. Install: pip install reportlab")
            return
        if not self.selected_student_id:
            messagebox.showinfo("Select", "Select a student first.")
            return

        out = filedialog.asksaveasfilename(
            title="Export student PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not out:
            return

        exporter = PDFExporter(self.sub_con, self.grade_con, self.question_map, self._student_pdf_options())
        try:
            exporter.export_student_pdf(self.selected_student_id, Path(out))
        except Exception as e:
            messagebox.showerror("PDF export failed", str(e))
            return
        messagebox.showinfo("PDF exported", f"Saved:\n{out}")

    def export_summary_pdf(self):
        if not self.require_grading_db():
            return
        if SimpleDocTemplate is None:
            messagebox.showinfo("PDF missing", "reportlab not installed. Install: pip install reportlab")
            return

        out = filedialog.asksaveasfilename(
            title="Export summary PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not out:
            return

        exporter = PDFExporter(self.sub_con, self.grade_con, self.question_map, self._student_pdf_options())
        try:
            exporter.export_summary_pdf(Path(out), self.compute_class_stats_text())
        except Exception as e:
            messagebox.showerror("PDF export failed", str(e))
            return
        messagebox.showinfo("PDF exported", f"Saved:\n{out}")

    def export_all_students_pdfs(self, report_tag_override: str | None = None):
        if not self.require_grading_db():
            return
        if SimpleDocTemplate is None:
            messagebox.showinfo("PDF missing", "reportlab not installed. Install: pip install reportlab")
            return

        out_dir = filedialog.askdirectory(title="Choose output folder for all student PDFs")
        if not out_dir:
            return

        report_tag = report_tag_override
        if report_tag is None:
            report_tag = simpledialog.askstring(
                "Report tag",
                "Enter report tag to include in filename (example: Midterm):",
                initialvalue="Midterm"
            )
            if report_tag is None:
                return
        report_tag = report_tag.strip() or "Midterm"

        exporter = PDFExporter(self.sub_con, self.grade_con, self.question_map, self._student_pdf_options())

        prog = tk.Toplevel(self.root)
        prog.title("Exporting PDFs...")
        prog.geometry("520x140")
        lbl = ttk.Label(prog, text="Starting...", style="Pastel.TLabel")
        lbl.pack(pady=10)
        pb = ttk.Progressbar(prog, orient="horizontal", length=480, mode="determinate")
        pb.pack(pady=10)

        def progress_cb(i, n, sid, ok, err):
            pb["maximum"] = n
            pb["value"] = i
            if ok:
                lbl.config(text=f"[{i}/{n}] Exported: {sid}")
            else:
                lbl.config(text=f"[{i}/{n}] Failed: {sid} ({err})")
            prog.update_idletasks()

        try:
            exporter.export_all_students_pdfs(Path(out_dir), report_tag=report_tag, progress_cb=progress_cb)
        except Exception as e:
            messagebox.showerror("Batch export failed", str(e))
            prog.destroy()
            return

        prog.destroy()
        messagebox.showinfo("Done", f"Saved PDFs to:\n{out_dir}")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
